from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.common.exceptions import NoSuchElementException
from pageObjects.Loginpage_sender import LoginsenderClass
from pageObjects.sender_manage_profile import SenderProfileClass
from pageObjects.Sender_user_ach_setup import SenderUserachSetupClass
import boto3
import os
from pageObjects.Sending_message_process import SendingClass
from dotenv import load_dotenv
load_dotenv()
aws_access_key_id = os.getenv('AWS_ACCESS_KEY')
aws_secret_access_key = os.getenv('AWS_SECRET_KEY')
bucket_name = os.getenv('AWS_BUCKET_NAME')
from utilities import XLUtils
from datetime import date
date = date.today()
from openpyxl.workbook import Workbook
from openpyxl.styles import Font
import openpyxl
from time import sleep

class Test_002_Sender_User:

    def excelReader_test_cases(self, method, sheet):
        path = "assets"
        fileName = "test_cases.xlsx"
        getFolderPath = os.path.join(path, fileName)
        self.rows = XLUtils.getRowCount(getFolderPath, sheet)
        for r in range(2, self.rows + 1):
            self.testCase = XLUtils.readData(getFolderPath, sheet, r, 1)
            self.action = XLUtils.readData(getFolderPath, sheet, r, 3)
            if self.testCase == method and self.action == 'Y':
                return "success"

    def test_excel_file_data_clear(self):
        book = openpyxl.load_workbook('Reports/result_sender_cases_output.xlsx')
        sheet = book['Sender-Testcases-Output']
        sheet.delete_rows(2, 20)
        book.save('Reports/result_sender_cases_output.xlsx')




    def check_sender_login(self):
        self.driver = webdriver.Chrome(ChromeDriverManager().install())
        self.driver.maximize_window()
        sleep(1)
        path = "assets"
        fileName = "data.xlsx"
        getFolderPath = os.path.join(path, fileName)
        self.lp = LoginsenderClass(self.driver)
        self.rows = XLUtils.getRowCount(getFolderPath, 'Sender')
        for r in range(4, self.rows + 1):
            self.login_url = XLUtils.readData(getFolderPath, 'Sender', r, 1)
            self.email = XLUtils.readData(getFolderPath, 'Sender', r, 2)
            self.password = XLUtils.readData(getFolderPath, 'Sender', r, 3)
            self.driver.get(self.login_url)
            sleep(1)
            self.lp.setsenderEmailName(self.email)
            sleep(2)
            self.lp.setsenderPasswordName(self.password)
            sleep(2)
            self.lp.clicksenderLogin()
            sleep(10)
            act_title = self.driver.title
            return act_title


    def test_sender_login(self):
        if self.excelReader_test_cases('test_sender_login', 'test_cases') == "success":
            act_title = self.check_sender_login()
            exp_title = "Start Sending Message"
            if act_title == exp_title:
                sleep(8)
                wb = Workbook()
                ws = wb.active
                ws['A1'] = "Test Cases"
                ws['A1'].font = Font(name="Arial", b=True)
                ws['B1'] = "Status"
                ws['B1'].font = Font(name="Arial", b=True)
                ws.title = "Sender-Testcases-Output"
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                wb.save('Reports/result_sender_cases_output.xlsx')
                #self.add_s3bucket_file_upload()
                self.driver.close()
                assert True
            else:
                wb = Workbook()
                ws = wb.active
                ws['A1'] = "Test Cases"
                ws['A1'].font = Font(name="Arial", b=True)
                ws['B1'] = "Status"
                ws['B1'].font = Font(name="Arial", b=True)
                ws.title = "Sender-Testcases-Output"
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                wb.save('Reports/result_sender_cases_output.xlsx')
                #self.add_s3bucket_file_upload()
                self.driver.close()

    def test_sending_process_twilio(self):
        if self.excelReader_test_cases('test_sending_process_twilio', 'test_cases') == "success":
            act_title = self.check_sender_login()
            exp_title = "Start Sending Message"
            if act_title == exp_title:
                book = openpyxl.load_workbook('Reports/result_sender_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_sender_cases_output.xlsx')
                self.sending_obj = SendingClass(self.driver)
                self.sending_obj.clicksenderbtn()
                sleep(2)
                self.sending_obj.Sending_Process_view_btn_class()
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_sender_cases_output.xlsx')
                ws = book.active
                ws['A3'] = "test-sending-process-twilio"
                ws['B3'] = "Pass"
                book.save('Reports/result_sender_cases_output.xlsx')
                self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_sender_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_sender_cases_output.xlsx')
                self.driver.close()


    def test_sender_profile_update(self):
        if self.excelReader_test_cases('test_sender_profile_update', 'test_cases') == "success":
            act_title = self.check_sender_login()
            exp_title = "Start Sending Message"
            if act_title == exp_title:
                book = openpyxl.load_workbook('Reports/result_sender_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_sender_cases_output.xlsx')
                path = "assets"
                fileName = "data.xlsx"
                getFolderPath = os.path.join(path, fileName)
                self.lp = LoginsenderClass(self.driver)
                self.rows = XLUtils.getRowCount(getFolderPath, 'Sender')
                self.sender_profile_obj = SenderProfileClass(self.driver)
                for r in range(9, self.rows + 1):
                    self.first_name = XLUtils.readData(getFolderPath, 'Sender', r, 1)
                    self.last_name = XLUtils.readData(getFolderPath, 'Sender', r, 2)
                    self.phone_number = XLUtils.readData(getFolderPath, 'Sender', r, 3)
                    self.address_line1 = XLUtils.readData(getFolderPath, 'Sender', r, 4)
                    self.address_line2 = XLUtils.readData(getFolderPath, 'Sender', r, 5)
                    self.city = XLUtils.readData(getFolderPath, 'Sender', r, 6)
                    self.zipcode = XLUtils.readData(getFolderPath, 'Sender', r, 7)
                    self.security_number = XLUtils.readData(getFolderPath, 'Sender', r, 8)
                    self.paypal_email = XLUtils.readData(getFolderPath, 'Sender', r, 9)
                    self.current_password = XLUtils.readData(getFolderPath, 'Sender', r, 10)
                    self.new_password = XLUtils.readData(getFolderPath, 'Sender', r, 11)
                    self.confirm_password = XLUtils.readData(getFolderPath, 'Sender', r, 12)
                    self.sender_profile_obj.clicksenderProfile()
                    self.sender_profile_obj.setsenderProfileFirstName(self.first_name)
                    self.sender_profile_obj.setsenderProfileLastName(self.last_name)
                    self.sender_profile_obj.setsenderProfilePhoneNumber(self.phone_number)
                    self.sender_profile_obj.clickSelectsenderProfileState()
                    self.sender_profile_obj.setsenderProfileAddressLine1(self.address_line1)
                    self.sender_profile_obj.setsenderProfileAddressLine2(self.address_line2)
                    self.sender_profile_obj.setsenderProfilecity(self.city)
                    self.sender_profile_obj.setsenderProfilezipcode(self.zipcode)
                    self.sender_profile_obj.setsenderProfilesecurityNumber(self.security_number)
                    self.sender_profile_obj.setsenderProfile_paypal_Email(self.paypal_email)
                    self.sender_profile_obj.clicksenderuploadimage()
                    self.sender_profile_obj.setsenderProfilecurrentPassword(self.current_password)
                    self.sender_profile_obj.setsenderProfilenewPassword(self.new_password)
                    self.sender_profile_obj.setsenderProfileconfirmPassword(self.confirm_password)
                    self.sender_profile_obj.clickupdate_sender_profilebtn()
                    sleep(5)
                    try:
                        success_toastr = self.driver.find_element_by_class_name("toast-error")
                        if success_toastr:
                            book = openpyxl.load_workbook('Reports/result_sender_cases_output.xlsx')
                            ws = book.active
                            ws['A4'] = "test-sender-profile"
                            ws['B4'] = "Fail"
                            book.save('Reports/result_sender_cases_output.xlsx')
                            self.driver.close()
                    except NoSuchElementException:
                        sleep(2)
                        book = openpyxl.load_workbook('Reports/result_sender_cases_output.xlsx')
                        ws = book.active
                        ws['A4'] = "test-sender-profile"
                        ws['B4'] = "Pass"
                        book.save('Reports/result_sender_cases_output.xlsx')
                        self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_sender_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_sender_cases_output.xlsx')
                self.driver.close()

    def test_sender_Achsetup(self):
        if self.excelReader_test_cases('test_sender_Achsetup', 'test_cases') == "success":
            act_title = self.check_sender_login()
            exp_title = "Start Sending Message"
            if act_title == exp_title:
                book = openpyxl.load_workbook('Reports/result_sender_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_sender_cases_output.xlsx')
                #self.add_s3bucket_file_upload()
                path = "assets"
                fileName = "data.xlsx"
                getFolderPath = os.path.join(path, fileName)
                self.lp = LoginsenderClass(self.driver)
                self.rows = XLUtils.getRowCount(getFolderPath, 'Sender')
                self.sender_ach_setup_obj = SenderUserachSetupClass(self.driver)
                for r in range(14, self.rows + 1):
                    self.first_name = XLUtils.readData(getFolderPath, 'Sender', r, 1)
                    self.middle_name = XLUtils.readData(getFolderPath, 'Sender', r, 2)
                    self.last_name = XLUtils.readData(getFolderPath, 'Sender', r, 3)
                    self.suffix = XLUtils.readData(getFolderPath, 'Sender', r, 4)
                    self.email = XLUtils.readData(getFolderPath, 'Sender', r, 5)
                    self.phone_number = XLUtils.readData(getFolderPath, 'Sender', r, 6)
                    self.address_line1 = XLUtils.readData(getFolderPath, 'Sender', r, 7)
                    self.address_line2 = XLUtils.readData(getFolderPath, 'sending_user_ach', r, 8)
                    self.city = XLUtils.readData(getFolderPath, 'Sender', r, 9)
                    self.postal_code = XLUtils.readData(getFolderPath, 'Sender', r, 10)
                    self.country = XLUtils.readData(getFolderPath, 'Sender', r, 11)
                    self.security_number = XLUtils.readData(getFolderPath, 'Sender', r, 12)
                    self.account_number = XLUtils.readData(getFolderPath, 'Sender', r, 13)
                    self.routing_number = XLUtils.readData(getFolderPath, 'Sender', r, 14)
                    self.confirm_routing_number = XLUtils.readData(getFolderPath, 'Sender', r, 15)
                    self.sender_ach_setup_obj.clicksenderProfile()
                    self.sender_ach_setup_obj.clicksenderAchbtn()
                    self.sender_ach_setup_obj.setsenderAchFirstName(self.first_name)
                    self.sender_ach_setup_obj.setsenderAchMiddleWare(self.middle_name)
                    self.sender_ach_setup_obj.setsenderAchLastName(self.last_name)
                    self.sender_ach_setup_obj.setsenderAchsuffix(self.suffix)
                    self.sender_ach_setup_obj.setsenderAchEmail(self.email)
                    self.sender_ach_setup_obj.setsenderAchPhonenumber(self.phone_number)
                    sleep(2)
                    self.sender_ach_setup_obj.setsenderAchAddressLine1(self.address_line1)
                    sleep(2)
                    self.sender_ach_setup_obj.setsenderAchAddressLine2(self.address_line2)
                    sleep(2)
                    self.sender_ach_setup_obj.setsenderAchcity(self.city)
                    sleep(2)
                    self.sender_ach_setup_obj.clickSelectsenderAchState()
                    sleep(2)
                    self.sender_ach_setup_obj.setsenderAchpostalcode(self.postal_code)
                    sleep(2)
                    self.sender_ach_setup_obj.setsenderAchcountry(self.country)
                    sleep(2)
                    self.sender_ach_setup_obj.setsenderAchsocialsecurityNumber(self.security_number)
                    sleep(2)
                    self.sender_ach_setup_obj.click_sender_ach_submit_btn()
                    sleep(2)
                    try:
                        success_toastr = self.driver.find_element_by_class_name("toast-error")
                        if success_toastr:
                            book = openpyxl.load_workbook('Reports/result_sender_cases_output.xlsx')
                            ws = book.active
                            ws['A5'] = "test-sender-ach-setup"
                            ws['B5'] = "Fail"
                            book.save('Reports/result_sender_cases_output.xlsx')
                            #self.add_s3bucket_file_upload()
                            self.driver.close()
                    except NoSuchElementException:
                        self.sender_ach_setup_obj.clickSelectsenderAchType()
                        sleep(2)
                        self.sender_ach_setup_obj.setsenderAchAccountNumber(self.account_number)
                        sleep(2)
                        self.sender_ach_setup_obj.setsenderAchRoutingNumber(self.routing_number)
                        sleep(2)
                        self.sender_ach_setup_obj.setsenderAchConfirmRoutingNumber(self.confirm_routing_number)
                        sleep(2)
                        self.sender_ach_setup_obj.click_sender_update_ach_submit_btn()
                        sleep(5)
                        try:
                            success_toastr = self.driver.find_element_by_class_name("toast-error")
                            if success_toastr:
                                book = openpyxl.load_workbook('Reports/result_sender_cases_output.xlsx')
                                ws = book.active
                                ws['A5'] = "test-sender-ach-setup"
                                ws['B5'] = "Fail"
                                book.save('Reports/result_sender_cases_output.xlsx')
                                #self.add_s3bucket_file_upload()
                                self.driver.close()
                        except NoSuchElementException:
                            sleep(2)
                            book = openpyxl.load_workbook('Reports/result_sender_cases_output.xlsx')
                            ws = book.active
                            ws['A5'] = "test-sender-ach-setup"
                            ws['B5'] = "Pass"
                            book.save('Reports/result_sender_cases_output.xlsx')
                            #self.add_s3bucket_file_upload()
                            self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_sender_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_sender_cases_output.xlsx')
                self.driver.close()



    def test_excel_empty_row_remove(self):
        book = openpyxl.load_workbook('Reports/result_sender_cases_output.xlsx')
        ws = book.active
        index_row = []
        # loop each row in column A
        for i in range(1, ws.max_row):
            # define emptiness of cell
            if ws.cell(i, 1).value is None:
                # collect indexes of rows
                index_row.append(i)
        # loop each index value
        for row_del in range(len(index_row)):
            ws.delete_rows(idx=index_row[row_del], amount=1)
            # exclude offset of rows through each iteration
            index_row = list(map(lambda k: k - 1, index_row))
        book.save('Reports/result_sender_cases_output.xlsx')


    def test_s3bucket_file_upload(self):
        client_s3 = boto3.client('s3', aws_access_key_id=aws_access_key_id,
                                 aws_secret_access_key=aws_secret_access_key)
        with open('Reports/result_sender_cases_output.xlsx', 'rb') as data:
            client_s3.upload_fileobj(data, bucket_name, 'Reports/result_sender' + str(date) + '.xlsx')



