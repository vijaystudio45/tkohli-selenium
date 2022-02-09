from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.common.exceptions import NoSuchElementException
from pageObjects.orgadmin_user_Loginpage import OrgadminLoginClass
from pageObjects.orgadmin_view_report import Reports_orgadmin_Class
from pageObjects.orgadmin_past_report import Orgadmin_PastReportClass
from pageObjects.orgadmin_report_history import ReportorgadminSchedulerHistoryClass
from pageObjects.orgadmin_clicks import Orgadmin_clickClass
from pageObjects.orgadmin_todaycampaign import orgadminTodayCampaignsClass
from pageObjects.orgadmin_campaign import CampaignorgadminClass
from pageObjects.orgadmin_organization import orgadmin_OrganizationsClass
from pageObjects.orgadmin_vcard import orgadmin_vcardclass
from pageObjects.orgadmin_api_key import orgadminApiKeyClass
from pageObjects.orgadmin_business_units import orgadminBusinessUnitClass
from pageObjects.orgadmin_message_merge import orgadminMessagemergeClass
from pageObjects.orgadmin_conversation_history import orgadminConversationhistoryClass
from pageObjects.orgadmin_message_merge_parser import orgadminMessagemergeparserClass
from pageObjects.orgadmin_common_response import orgadminCommonResponseClass
from pageObjects.orgadmin_new_response import orgadminnewresponseClass
from pageObjects.orgadmin_past_response import orgadminpastresponseClass
from pageObjects.orgadmin_ignore_response import orgadminignoreresponseClass
from pageObjects.orgadmin_manage_profile import orgadminManageProfileClass
from utilities import XLUtils
from datetime import date
date = date.today()
from openpyxl.workbook import Workbook
from openpyxl.styles import Font
import openpyxl
from time import sleep
import boto3
import os
from dotenv import load_dotenv
load_dotenv()
aws_access_key_id = os.getenv('AWS_ACCESS_KEY')
aws_secret_access_key = os.getenv('AWS_SECRET_KEY')
bucket_name = os.getenv('AWS_BUCKET_NAME')

class Test_004_orgadmin_User:


    def excelReader_test_cases(self, method, sheet):
        path = "assets"
        fileName = "test_cases.xlsx"
        getFolderPath = os.path.join(path, fileName)
        self.rows = XLUtils.getRowCount(getFolderPath, sheet)
        for r in range(2, self.rows + 1):
            self.testCase = XLUtils.readData(getFolderPath, sheet, r, 1)
            self.action = XLUtils.readData(getFolderPath, sheet, r, 3)
            if self.testCase == method and self.action == 'Y':
                return "success"


    def check_org_admin_user_login(self):
        self.driver = webdriver.Chrome(ChromeDriverManager().install())
        self.driver.maximize_window()
        sleep(1)
        path = "assets"
        fileName = "data.xlsx"
        getFolderPath = os.path.join(path, fileName)
        self.lp = OrgadminLoginClass(self.driver)
        self.rows = XLUtils.getRowCount(getFolderPath, 'Org-Admin')
        for r in range(4, self.rows + 1):
            self.login_url = XLUtils.readData(getFolderPath, 'Org-Admin', r, 1)
            self.email = XLUtils.readData(getFolderPath, 'Org-Admin', r, 2)
            self.password = XLUtils.readData(getFolderPath, 'Org-Admin', r, 3)
            self.driver.get(self.login_url)
            sleep(1)
            self.lp.setorgadminuserEmailName(self.email)
            sleep(2)
            self.lp.setorgadminuserPassword(self.password)
            sleep(2)
            self.lp.clickLogin()
            sleep(10)
            act_title = self.driver.title
            return act_title

    def test_excel_file_data_clear(self):
        book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
        sheet = book['Org-Admin-Testcases-Output']
        sheet.delete_rows(2, 70)
        book.save('Reports/result_org_admin_cases_output.xlsx')

    def test_org_admin_user_login(self):
        if self.excelReader_test_cases('test_org_admin_user_login', 'test_cases') == "success":
            act_title = self.check_org_admin_user_login()
            exp_title = "User Dashboard"
            if act_title == exp_title:
                sleep(8)
                wb = Workbook()
                ws = wb.active
                ws['A1'] = "Test Cases"
                ws['A1'].font = Font(name="Arial", b=True)
                ws['B1'] = "Status"
                ws['B1'].font = Font(name="Arial", b=True)
                ws.title = "Org-Admin-Testcases-Output"
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                wb.save('Reports/result_org_admin_cases_output.xlsx')
                self.driver.close()
                assert True
            else:
                wb = Workbook()
                ws = wb.active
                ws['A1'] = "Test Cases"
                ws['A1'].font = Font(name="Arial", b=True)
                ws['B1'] = "Status"
                ws['B1'].font = Font(name="Arial", b=True)
                ws.title = "Org-Admin-Testcases-Output"
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                wb.save('Reports/result_org_admin_cases_output.xlsx')
                self.driver.close()


    def test_org_admin_view_report_campaign(self):
        if self.excelReader_test_cases('test_org_admin_view_report_campaign', 'test_cases') == "success":
            act_title = self.check_org_admin_user_login()
            exp_title = "User Dashboard"
            if act_title == exp_title:
                book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_org_admin_cases_output.xlsx')
                self.org_admin_report = Reports_orgadmin_Class(self.driver)
                sleep(2)
                self.org_admin_report.clickReportbtn()
                try:
                    self.org_admin_report.clickvewReportcampaignbtn()
                    sleep(3)
                    book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                    ws = book.active
                    ws['A3'] = "test-view-report-campaign"
                    ws['B3'] = "Pass"
                    book.save('Reports/result_org_admin_cases_output.xlsx')
                    self.driver.close()
                except NoSuchElementException:
                    sleep(4)
                    book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                    ws = book.active
                    ws['A3'] = "test-view-report-campaign"
                    ws['B3'] = "No data Found"
                    book.save('Reports/result_org_admin_cases_output.xlsx')
                    self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_org_admin_cases_output.xlsx')
                self.driver.close()

    def test_org_admin_view_report_unsigned_campaign(self):
        if self.excelReader_test_cases('test_org_admin_view_report_unsigned_campaign', 'test_cases') == "success":
            act_title = self.check_org_admin_user_login()
            exp_title = "User Dashboard"
            if act_title == exp_title:
                book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_org_admin_cases_output.xlsx')
                self.org_admin_report = Reports_orgadmin_Class(self.driver)
                sleep(2)
                self.org_admin_report.clickReportbtn()
                try:
                    self.org_admin_report.clickvewReportunsignedcampaignbtn()
                    sleep(3)
                    book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                    ws = book.active
                    ws['A4'] = "test-view-report-unsigned-campaign"
                    ws['B4'] = "Pass"
                    book.save('Reports/result_org_admin_cases_output.xlsx')
                    self.driver.close()
                except NoSuchElementException:
                    sleep(4)
                    book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                    ws = book.active
                    ws['A4'] = "test-view-report-unsigned-campaign"
                    ws['B4'] = "No data Found"
                    book.save('Reports/result_org_admin_cases_output.xlsx')
                    self.driver.close()

            else:
                book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_org_admin_cases_output.xlsx')
                self.driver.close()



    def test_org_admin_view_report_response(self):
        if self.excelReader_test_cases('test_org_admin_view_report_response', 'test_cases') == "success":
            act_title = self.check_org_admin_user_login()
            exp_title = "User Dashboard"
            if act_title == exp_title:
                book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_org_admin_cases_output.xlsx')
                self.org_admin_report = Reports_orgadmin_Class(self.driver)
                sleep(2)
                self.org_admin_report.clickReportbtn()
                try:
                    self.org_admin_report.clickvewReportresponsebtn()
                    sleep(3)
                    book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                    ws = book.active
                    ws['A5'] = "test-view-report-response"
                    ws['B5'] = "Pass"
                    book.save('Reports/result_org_admin_cases_output.xlsx')
                    self.driver.close()
                except NoSuchElementException:
                    sleep(4)
                    book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                    ws = book.active
                    ws['A5'] = "test-view-report-response"
                    ws['B5'] = "No data found"
                    book.save('Reports/result_org_admin_cases_output.xlsx')
                    self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_org_admin_cases_output.xlsx')
                self.driver.close()





    def test_org_admin_past_report_view(self):
        if self.excelReader_test_cases('test_org_admin_past_report_view', 'test_cases') == "success":
            act_title = self.check_org_admin_user_login()
            exp_title = "User Dashboard"
            if act_title == exp_title:
                book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_org_admin_cases_output.xlsx')
                self.org_admin_past_report = Orgadmin_PastReportClass(self.driver)
                sleep(2)
                self.org_admin_past_report.click_past_report()
                sleep(2)
                try:
                    self.org_admin_past_report.check_download_btn()
                    sleep(3)
                    try:
                        success_toastr = self.driver.find_element_by_class_name("toast-error")
                        if success_toastr:
                            book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                            ws = book.active
                            ws['A6'] = "test-past-report-download"
                            ws['B6'] = "Fail"
                            book.save('Reports/result_org_admin_cases_output.xlsx')
                            self.driver.close()
                    except NoSuchElementException:
                        book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A6'] = "test-past-report-download"
                        ws['B6'] = "Pass"
                        book.save('Reports/result_org_admin_cases_output.xlsx')
                        self.driver.close()
                except NoSuchElementException:
                    book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                    ws = book.active
                    ws['A6'] = "test-past-report-download"
                    ws['B6'] = "No data Found"
                    book.save('Reports/result_org_admin_cases_output.xlsx')
                    self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_org_admin_cases_output.xlsx')
                self.driver.close()




    def test_org_admin_report_history(self):
        if self.excelReader_test_cases('test_org_admin_report_history', 'test_cases') == "success":
            act_title = self.check_org_admin_user_login()
            exp_title = "User Dashboard"
            if act_title == exp_title:
                book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_org_admin_cases_output.xlsx')
                self.org_admin_report_history = ReportorgadminSchedulerHistoryClass(self.driver)
                sleep(2)
                self.org_admin_report_history.clickreportbtn()
                try:
                    self.org_admin_report_history.clickDownloadReport()
                    sleep(6)
                    book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                    ws = book.active
                    ws['A7'] = "test-report-history-download"
                    ws['B7'] = "Pass"
                    book.save('Reports/result_org_admin_cases_output.xlsx')
                    self.driver.close()
                except NoSuchElementException:
                    book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                    ws = book.active
                    ws['A7'] = "test-report-history-download"
                    ws['B7'] = "Fail"
                    book.save('Reports/result_org_admin_cases_output.xlsx')
                    self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_org_admin_cases_output.xlsx')
                self.driver.close()

    def test_org_admin_clicks(self):
        if self.excelReader_test_cases('test_org_admin_clicks', 'test_cases') == "success":
            act_title = self.check_org_admin_user_login()
            exp_title = "User Dashboard"
            if act_title == exp_title:
                book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_org_admin_cases_output.xlsx')
                self.org_admin_click = Orgadmin_clickClass(self.driver)
                sleep(2)
                self.org_admin_click.click_page_link()
                sleep(3)
                book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                ws = book.active
                ws['A8'] = "test-clicks-data"
                ws['B8'] = "Pass"
                book.save('Reports/result_org_admin_cases_output.xlsx')
                self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_org_admin_cases_output.xlsx')
                self.driver.close()



    def test_org_admin_insert_campaigns(self):
        if self.excelReader_test_cases('test_org_admin_insert_campaigns', 'test_cases') == "success":
            act_title = self.check_org_admin_user_login()
            exp_title = "User Dashboard"
            if act_title == exp_title:
                book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_org_admin_cases_output.xlsx')
                self.org_admin_campaigns = CampaignorgadminClass(self.driver)
                sleep(2)
                self.org_admin_campaigns.clickCampaigns()
                sleep(2)
                path = "assets"
                fileName = "data.xlsx"
                getFolderPath = os.path.join(path, fileName)
                self.rows = XLUtils.getRowCount(getFolderPath, 'Org-Admin')
                for r in range(9, self.rows + 1):
                    self.campaign_name = XLUtils.readData(getFolderPath, 'Org-Admin', r, 1)
                    self.campaign_message = XLUtils.readData(getFolderPath, 'Org-Admin', r, 2)
                    self.org_admin_campaigns.clickCreateCampaigns()
                    sleep(2)
                    self.org_admin_campaigns.setCampaignName(self.campaign_name)
                    sleep(2)
                    self.org_admin_campaigns.clickcampaigncsv()
                    sleep(2)
                    self.org_admin_campaigns.clickstartdate()
                    sleep(2)
                    self.org_admin_campaigns.click_Selectstart_time()
                    sleep(2)
                    self.org_admin_campaigns.clickenddate()
                    sleep(2)
                    self.org_admin_campaigns.click_Select_end_time()
                    sleep(2)
                    self.org_admin_campaigns.clickCampaignContinue()
                    sleep(2)
                    self.org_admin_campaigns.setCampaignMessage(self.campaign_message)
                    self.org_admin_campaigns.clickCampaignsbtn()
                    sleep(10)
                    act_title_create_campaigns = self.driver.title
                    exp_title_create_campaigns = "Campaigns"
                    if act_title_create_campaigns == exp_title_create_campaigns:
                        book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A10'] = "test-insert-campaigns"
                        ws['B10'] = "Pass"
                        book.save('Reports/result_org_admin_cases_output.xlsx')
                        self.driver.close()
                        assert True
                    else:
                        book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A10'] = "test-insert-campaigns"
                        ws['B10'] = "Fail"
                        book.save('Reports/result_org_admin_cases_output.xlsx')
                        self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_org_admin_cases_output.xlsx')
                self.driver.close()

    def test_org_admin_update_campaigns(self):
        if self.excelReader_test_cases('test_org_admin_update_campaigns', 'test_cases') == "success":
            act_title = self.check_org_admin_user_login()
            exp_title = "User Dashboard"
            if act_title == exp_title:
                book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_org_admin_cases_output.xlsx')
                self.org_admin_campaigns = CampaignorgadminClass(self.driver)
                sleep(2)
                self.org_admin_campaigns.clickCampaigns()
                sleep(2)
                path = "assets"
                fileName = "data.xlsx"
                getFolderPath = os.path.join(path, fileName)
                self.rows = XLUtils.getRowCount(getFolderPath, 'Org-Admin')
                for r in range(14, self.rows + 1):
                    self.campaign_name = XLUtils.readData(getFolderPath, 'Org-Admin', r, 1)
                    self.campaign_message = XLUtils.readData(getFolderPath, 'Org-Admin', r, 2)
                    try:
                        if self.driver.find_element_by_class_name("campaign_sec").text == 'Automation-campaign-add':
                            self.org_admin_campaigns.clickCampaignEditBtn()
                        else:
                            sleep(4)
                            book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                            ws = book.active
                            ws['A11'] = "test-update-campaigns"
                            ws['B11'] = "No data found"
                            book.save('Reports/result_org_admin_cases_output.xlsx')
                            # self.add_s3bucket_file_upload()
                            self.driver.close()
                    except NoSuchElementException:
                        sleep(4)
                        ws = book.active
                        ws['A11'] = "test-update-campaigns"
                        ws['B11'] = "No data found"
                        book.save('Reports/result_org_admin_cases_output.xlsx')
                        self.driver.close()
                    sleep(2)
                    self.org_admin_campaigns.setCampaignName(self.campaign_name)
                    sleep(2)
                    self.org_admin_campaigns.clickcampaigncsv()
                    sleep(2)
                    self.org_admin_campaigns.clickstartdate()
                    sleep(2)
                    self.org_admin_campaigns.click_edit_Selectstart_time()
                    sleep(2)
                    self.org_admin_campaigns.clickenddate()
                    sleep(2)
                    self.org_admin_campaigns.click_edit_Select_end_time()
                    sleep(2)
                    self.org_admin_campaigns.setCampaignMessage(self.campaign_message)
                    self.org_admin_campaigns.click_campaigns_update_btn_class()
                    sleep(10)
                    act_title_update_campaigns = self.driver.title
                    exp_title_update_campaigns = "Campaigns"
                    if act_title_update_campaigns == exp_title_update_campaigns:
                        ws = book.active
                        ws['A11'] = "test-update-campaigns"
                        ws['B11'] = "Pass"
                        book.save('Reports/result_org_admin_cases_output.xlsx')
                        self.driver.close()
                    else:
                        ws = book.active
                        ws['A11'] = "test-update-campaigns"
                        ws['B11'] = "Fail"
                        book.save('Reports/result_org_admin_cases_output.xlsx')
                        self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_org_admin_cases_output.xlsx')
                self.driver.close()

    def test_org_admin_delete_campaigns(self):
        if self.excelReader_test_cases('test_org_admin_delete_campaigns', 'test_cases') == "success":
            act_title = self.check_org_admin_user_login()
            exp_title = "User Dashboard"
            if act_title == exp_title:
                book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_org_admin_cases_output.xlsx')
                self.org_admin_campaigns = CampaignorgadminClass(self.driver)
                sleep(2)
                self.org_admin_campaigns.clickCampaigns()
                try:
                    if self.driver.find_element_by_class_name("campaign_sec").text == 'Automation-campaign-update':
                        self.org_admin_campaigns.clickRemoveCampaign()
                    else:
                        sleep(4)
                        book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A12'] = "test-delete-campaigns"
                        ws['B12'] = "No data found"
                        book.save('Reports/result_org_admin_cases_output.xlsx')
                        # self.add_s3bucket_file_upload()
                        self.driver.close()
                    sleep(2)
                    self.org_admin_campaigns.clickConfirmDelete()
                    sleep(3)
                    book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                    ws = book.active
                    ws['A12'] = "test-delete-campaigns"
                    ws['B12'] = "Pass"
                    book.save('Reports/result_org_admin_cases_output.xlsx')
                    self.driver.close()
                except NoSuchElementException:
                    book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                    ws = book.active
                    ws['A12'] = "test-delete-campaigns"
                    ws['B12'] = "No data found"
                    book.save('Reports/result_org_admin_cases_output.xlsx')
                    self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_org_admin_cases_output.xlsx')
                self.driver.close()



    def test_org_admin_today_campaigns(self):
        if self.excelReader_test_cases('test_org_admin_today_campaigns', 'test_cases') == "success":
            act_title = self.check_org_admin_user_login()
            exp_title = "User Dashboard"
            if act_title == exp_title:
                book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_org_admin_cases_output.xlsx')
                self.org_admin_today_campaigns = orgadminTodayCampaignsClass(self.driver)
                sleep(3)
                self.org_admin_today_campaigns.clickTodayCampaigns()
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                ws = book.active
                ws['A13'] = "test-today-campaigns-view"
                ws['B13'] = "Pass"
                book.save('Reports/result_org_admin_cases_output.xlsx')
                self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_org_admin_cases_output.xlsx')
                self.driver.close()

    def test_org_admin_Organizations_update(self):
     if self.excelReader_test_cases('test_org_admin_Organizations_update', 'test_cases') == "success":
        act_title = self.check_org_admin_user_login()
        exp_title = "User Dashboard"
        if act_title == exp_title:
            book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
            ws = book.active
            ws['A2'] = "test-login"
            ws['B2'] = "Pass"
            book.save('Reports/result_org_admin_cases_output.xlsx')
            path = "assets"
            fileName = "data.xlsx"
            getFolderPath = os.path.join(path, fileName)
            self.Org_obj = orgadmin_OrganizationsClass(self.driver)
            self.rows = XLUtils.getRowCount(getFolderPath, 'Org-Admin')
            for r in range(19, self.rows + 1):
                self.org_name = XLUtils.readData(getFolderPath, 'Org-Admin', r, 1)
                self.org_website = XLUtils.readData(getFolderPath, 'Org-Admin', r, 2)
                self.org_phone_number = XLUtils.readData(getFolderPath, 'Org-Admin', r, 3)
                self.org_first_name = XLUtils.readData(getFolderPath, 'Org-Admin', r, 4)
                self.org_last_name = XLUtils.readData(getFolderPath, 'Org-Admin', r, 5)
                self.org_contact_phone_number = XLUtils.readData(getFolderPath, 'Org-Admin', r, 6)
                self.Org_obj.clickorganization()
                sleep(2)
                try:
                    self.Org_obj.clickeditorgbtn()
                except NoSuchElementException:
                    sleep(4)
                    self.driver.close()
                sleep(2)
                self.Org_obj.setorgadmin_organizationWebsite(self.org_website)
                sleep(2)
                self.Org_obj.setorgadmin_organizationPhone_Number(self.org_phone_number)
                sleep(2)
                self.Org_obj.clickSelecttimezone()
                sleep(2)
                self.Org_obj.setorgadmin_organization_first_name(self.org_first_name)
                sleep(2)
                self.Org_obj.setorgadmin_organization_last_name(self.org_last_name)
                sleep(2)
                self.Org_obj.setorgadmin_organization_contact_phone_number(self.org_contact_phone_number)
                sleep(2)
                self.Org_obj.clickupdateorgbtn()
                sleep(4)
                act_title_org_update = self.driver.title
                exp_title_org_update = "Organizations"
                if act_title_org_update == exp_title_org_update:
                    book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                    ws = book.active
                    ws['A14'] = "test-organizations-update"
                    ws['B14'] = "Pass"
                    book.save('Reports/result_org_admin_cases_output.xlsx')
                    self.driver.close()
                else:
                    book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                    ws = book.active
                    ws['A14'] = "test-organizations-update"
                    ws['B14'] = "Fail"
                    book.save('Reports/result_org_admin_cases_output.xlsx')
                    self.driver.close()
        else:
            book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
            ws = book.active
            ws['A2'] = "test-login"
            ws['B2'] = "Fail"
            book.save('Reports/result_org_admin_cases_output.xlsx')
            self.driver.close()


    def test_org_admin_vcard_add(self):
        if self.excelReader_test_cases('test_org_admin_vcard_add', 'test_cases') == "success":
         act_title = self.check_org_admin_user_login()
         exp_title = "User Dashboard"
         if act_title == exp_title:
             book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
             ws = book.active
             ws['A2'] = "test-login"
             ws['B2'] = "Pass"
             book.save('Reports/result_org_admin_cases_output.xlsx')
             path = "assets"
             fileName = "data.xlsx"
             getFolderPath = os.path.join(path, fileName)
             self.vcard_obj = orgadmin_vcardclass(self.driver)
             self.rows = XLUtils.getRowCount(getFolderPath, 'Org-Admin')
             for r in range(24, self.rows + 1):
                 self.vcard_name = XLUtils.readData(getFolderPath, 'Org-Admin', r, 1)
                 self.vcard_display_name = XLUtils.readData(getFolderPath, 'Org-Admin', r, 2)
                 self.vcard_website = XLUtils.readData(getFolderPath, 'Org-Admin', r, 3)
                 self.vcard_phone_number = XLUtils.readData(getFolderPath, 'Org-Admin', r, 4)
                 self.vcard_first_name = XLUtils.readData(getFolderPath, 'Org-Admin', r, 5)
                 self.vcard_last_name = XLUtils.readData(getFolderPath, 'Org-Admin', r, 6)
                 self.vcard_email = XLUtils.readData(getFolderPath, 'Org-Admin', r, 7)
                 self.vcard_address = XLUtils.readData(getFolderPath, 'Org-Admin', r, 8)
                 self.vcard_obj.clickvcard()
                 sleep(2)
                 self.vcard_obj.clickcreatevcard()
                 sleep(2)
                 self.vcard_obj.set_org_admin_vcard_name(self.vcard_name)
                 sleep(2)
                 self.vcard_obj.set_org_admin_display_vcard_name(self.vcard_display_name)
                 sleep(2)
                 self.vcard_obj.set_org_admin_vcard_website(self.vcard_website)
                 sleep(2)
                 self.vcard_obj.set_org_admin_vcard_telephone(self.vcard_phone_number)
                 sleep(2)
                 self.vcard_obj.set_org_admin_vcard_first_name(self.vcard_first_name)
                 sleep(2)
                 self.vcard_obj.set_org_admin_vcard_last_name(self.vcard_last_name)
                 sleep(2)
                 self.vcard_obj.set_org_admin_vcard_email(self.vcard_email)
                 sleep(2)
                 self.vcard_obj.set_org_admin_vcard_Address(self.vcard_address)
                 self.vcard_obj.clickcreatevcardbtn()
                 sleep(4)
                 act_title_vcard_add = self.driver.title
                 exp_title_vcard_add = "Vcard"
                 if act_title_vcard_add == exp_title_vcard_add:
                     book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                     ws = book.active
                     ws['A15'] = "test-vcard-add"
                     ws['B15'] = "Pass"
                     book.save('Reports/result_org_admin_cases_output.xlsx')
                     self.driver.close()
                 else:
                     book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                     ws = book.active
                     ws['A15'] = "test-vcard-add"
                     ws['B15'] = "Fail"
                     book.save('Reports/result_org_admin_cases_output.xlsx')
                     self.driver.close()
         else:
             book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
             ws = book.active
             ws['A2'] = "test-login"
             ws['B2'] = "Fail"
             book.save('Reports/result_org_admin_cases_output.xlsx')
             self.driver.close()





    def test_org_admin_vcard_update(self):
        if self.excelReader_test_cases('test_org_admin_vcard_update', 'test_cases') == "success":
         act_title = self.check_org_admin_user_login()
         exp_title = "User Dashboard"
         if act_title == exp_title:
             book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
             ws = book.active
             ws['A2'] = "test-login"
             ws['B2'] = "Pass"
             book.save('Reports/result_org_admin_cases_output.xlsx')
             path = "assets"
             fileName = "data.xlsx"
             getFolderPath = os.path.join(path, fileName)
             self.vcard_obj = orgadmin_vcardclass(self.driver)
             self.rows = XLUtils.getRowCount(getFolderPath, 'Org-Admin')
             for r in range(29, self.rows + 1):
                 self.vcard_name = XLUtils.readData(getFolderPath, 'Org-Admin', r, 1)
                 self.vcard_display_name = XLUtils.readData(getFolderPath, 'Org-Admin', r, 2)
                 self.vcard_website = XLUtils.readData(getFolderPath, 'Org-Admin', r, 3)
                 self.vcard_phone_number = XLUtils.readData(getFolderPath, 'Org-Admin', r, 4)
                 self.vcard_first_name = XLUtils.readData(getFolderPath, 'Org-Admin', r, 5)
                 self.vcard_last_name = XLUtils.readData(getFolderPath, 'Org-Admin', r, 6)
                 self.vcard_address = XLUtils.readData(getFolderPath, 'Org-Admin', r, 8)
                 self.vcard_obj.clickvcard()
                 sleep(2)
                 try:
                     if self.driver.find_element_by_class_name("name").text == 'Automation-vcard-add':
                         self.vcard_obj.clickeditvcardbtn()
                     else:
                         sleep(4)
                         book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                         ws = book.active
                         ws['A16'] = "test-vcard-update"
                         ws['B16'] = "No data found"
                         book.save('Reports/result_org_admin_cases_output.xlsx')
                         # self.add_s3bucket_file_upload()
                         self.driver.close()
                 except NoSuchElementException:
                     book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                     sleep(4)
                     ws = book.active
                     ws['A16'] = "test-vcard-update"
                     ws['B16'] = "No data found"
                     book.save('Reports/result_org_admin_cases_output.xlsx')
                     self.driver.close()
                 sleep(2)
                 self.vcard_obj.set_org_admin_vcard_name(self.vcard_name)
                 sleep(2)
                 self.vcard_obj.set_org_admin_display_vcard_name(self.vcard_display_name)
                 sleep(2)
                 self.vcard_obj.set_org_admin_vcard_website(self.vcard_website)
                 sleep(2)
                 self.vcard_obj.set_org_admin_vcard_telephone(self.vcard_phone_number)
                 sleep(2)
                 self.vcard_obj.set_org_admin_vcard_first_name(self.vcard_first_name)
                 sleep(2)
                 self.vcard_obj.set_org_admin_vcard_last_name(self.vcard_last_name)
                 sleep(2)
                 self.vcard_obj.set_org_admin_vcard_Address(self.vcard_address)
                 self.vcard_obj.clickupdatevcardbtn()
                 sleep(4)
                 act_title_vcard_update = self.driver.title
                 exp_title_vcard_update = "Vcard"
                 if act_title_vcard_update == exp_title_vcard_update:
                     book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                     ws = book.active
                     ws['A16'] = "test-vcard-update"
                     ws['B16'] = "Pass"
                     book.save('Reports/result_org_admin_cases_output.xlsx')
                     self.driver.close()
                 else:
                     book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                     ws = book.active
                     ws['A16'] = "test-vcard-update"
                     ws['B16'] = "Fail"
                     book.save('Reports/result_org_admin_cases_output.xlsx')
                     self.driver.close()
         else:
             book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
             ws = book.active
             ws['A2'] = "test-login"
             ws['B2'] = "Fail"
             book.save('Reports/result_org_admin_cases_output.xlsx')
             self.driver.close()

    def test_org_admin_vcard_delete(self):
        if self.excelReader_test_cases('test_org_admin_vcard_delete', 'test_cases') == "success":
            act_title = self.check_org_admin_user_login()
            exp_title = "User Dashboard"
            if act_title == exp_title:
                book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_org_admin_cases_output.xlsx')
                self.vcard_obj = orgadmin_vcardclass(self.driver)
                sleep(2)
                self.vcard_obj.clickvcard()
                try:
                    if self.driver.find_element_by_class_name("name").text == 'Automation-vcard-update':
                        self.vcard_obj.clickRemovevcard()
                    else:
                        sleep(4)
                        book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A17'] = "test-vcard-delete"
                        ws['B17'] = "No data found"
                        book.save('Reports/result_org_admin_cases_output.xlsx')
                        # self.add_s3bucket_file_upload()
                        self.driver.close()
                    sleep(2)
                    self.vcard_obj.clickConfirmDelete()
                    sleep(4)
                    book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                    ws = book.active
                    ws['A17'] = "test-vcard-delete"
                    ws['B17'] = "Pass"
                    book.save('Reports/result_org_admin_cases_output.xlsx')
                    self.driver.close()
                except NoSuchElementException:
                    book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                    ws = book.active
                    ws['A17'] = "test-vcard-delete"
                    ws['B17'] = "Fail"
                    book.save('Reports/result_org_admin_cases_output.xlsx')
                    self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_org_admin_cases_output.xlsx')
                self.driver.close()

    def test_org_admin_api_key_add(self):
        if self.excelReader_test_cases('test_org_admin_api_key_add', 'test_cases') == "success":
            act_title = self.check_org_admin_user_login()
            exp_title = "User Dashboard"
            if act_title == exp_title:
                book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_org_admin_cases_output.xlsx')
                path = "assets"
                fileName = "data.xlsx"
                getFolderPath = os.path.join(path, fileName)
                self.org_api_key = orgadminApiKeyClass(self.driver)
                self.rows = XLUtils.getRowCount(getFolderPath, 'Org-Admin')
                for r in range(34, self.rows + 1):
                    self.api_key = XLUtils.readData(getFolderPath, 'Org-Admin', r, 1)
                    self.org_api_key.clickapibtn()
                    sleep(2)
                    self.org_api_key.clickcreatebtn()
                    self.org_api_key.setapikey(self.api_key)
                    sleep(2)
                    self.org_api_key.clickdatepickertest()
                    sleep(2)
                    self.org_api_key.clickActiveCheckbox()
                    self.org_api_key.clickapikeycreatebtn()
                    sleep(4)
                    act_title_create_api_key = self.driver.title
                    exp_title_create_api_key = "Api Key"
                    if act_title_create_api_key == exp_title_create_api_key:
                        book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A18'] = "test-apikey-add"
                        ws['B18'] = "Pass"
                        book.save('Reports/result_org_admin_cases_output.xlsx')
                        self.driver.close()
                    else:
                        book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A18'] = "test-apikey-add"
                        ws['B18'] = "Fail"
                        book.save('Reports/result_org_admin_cases_output.xlsx')
                        self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_org_admin_cases_output.xlsx')
                self.driver.close()


    def test_org_admin_api_key_update(self):
        if self.excelReader_test_cases('test_org_admin_api_key_update', 'test_cases') == "success":
            act_title = self.check_org_admin_user_login()
            exp_title = "User Dashboard"
            if act_title == exp_title:
                book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_org_admin_cases_output.xlsx')
                path = "assets"
                fileName = "data.xlsx"
                getFolderPath = os.path.join(path, fileName)
                self.org_api_key = orgadminApiKeyClass(self.driver)
                self.rows = XLUtils.getRowCount(getFolderPath, 'Org-Admin')
                for r in range(39, self.rows + 1):
                    self.api_key = XLUtils.readData(getFolderPath, 'Org-Admin', r, 1)
                    self.org_api_key.clickapibtn()
                    sleep(2)
                    try:
                        if self.driver.find_element_by_class_name("key").text == 'Automation-api-key-add':
                            self.org_api_key.clickapikeyeditbtn()
                        else:
                            sleep(4)
                            book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                            ws = book.active
                            ws['A19'] = "test-apikey-update"
                            ws['B19'] = "No data found"
                            book.save('Reports/result_org_admin_cases_output.xlsx')
                            # self.add_s3bucket_file_upload()
                            self.driver.close()
                    except NoSuchElementException:
                        sleep(4)
                        book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A19'] = "test-apikey-update"
                        ws['B19'] = "No data found"
                        book.save('Reports/result_org_admin_cases_output.xlsx')
                        self.driver.close()
                    self.org_api_key.setapikey(self.api_key)
                    sleep(2)
                    self.org_api_key.clickdatepickertest()
                    sleep(2)
                    self.org_api_key.clickActiveCheckbox()
                    sleep(2)
                    self.org_api_key.clickapikeyupdatebtn()
                    sleep(4)
                    act_title_update_api_key = self.driver.title
                    exp_title_update_api_key = "Api Key"
                    if act_title_update_api_key == exp_title_update_api_key:
                        book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A19'] = "test-apikey-update"
                        ws['B19'] = "Pass"
                        book.save('Reports/result_org_admin_cases_output.xlsx')
                        self.driver.close()
                    else:
                        book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A19'] = "test-apikey-update"
                        ws['B19'] = "Fail"
                        book.save('Reports/result_org_admin_cases_output.xlsx')
                        self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_org_admin_cases_output.xlsx')
                self.driver.close()

    def test_org_admin_api_key_delete(self):
        if self.excelReader_test_cases('test_org_admin_api_key_delete', 'test_cases') == "success":
            act_title = self.check_org_admin_user_login()
            exp_title = "User Dashboard"
            if act_title == exp_title:
                book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_org_admin_cases_output.xlsx')
                self.org_api_key = orgadminApiKeyClass(self.driver)
                self.org_api_key.clickapibtn()
                sleep(2)
                try:
                    if self.driver.find_element_by_class_name("key").text == 'Automation-api-key-update':
                        self.org_api_key.clickRemoveApikeyLink()
                    else:
                        sleep(4)
                        book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A19'] = "test-apikey-update"
                        ws['B19'] = "No data found"
                        book.save('Reports/result_org_admin_cases_output.xlsx')
                        # self.add_s3bucket_file_upload()
                        self.driver.close()
                    sleep(2)
                    self.org_api_key.clickConfirmDelete()
                    sleep(2)
                    book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                    ws = book.active
                    ws['A20'] = "test-apikey-delete"
                    ws['B20'] = "Pass"
                    book.save('Reports/result_org_admin_cases_output.xlsx')
                    self.driver.close()
                except NoSuchElementException:
                    book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                    ws = book.active
                    ws['A20'] = "test-apikey-delete"
                    ws['B20'] = "No data found"
                    book.save('Reports/result_org_admin_cases_output.xlsx')
                    self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_org_admin_cases_output.xlsx')
                self.driver.close()

    def test_org_admin_business_unit_add(self):
        if self.excelReader_test_cases('test_org_admin_business_unit_add', 'test_cases') == "success":
            act_title = self.check_org_admin_user_login()
            exp_title = "User Dashboard"
            if act_title == exp_title:
                book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_org_admin_cases_output.xlsx')
                path = "assets"
                fileName = "data.xlsx"
                getFolderPath = os.path.join(path, fileName)
                self.org_admin_unit = orgadminBusinessUnitClass(self.driver)
                self.rows = XLUtils.getRowCount(getFolderPath, 'Org-Admin')
                for r in range(44, self.rows + 1):
                    self.unit_name = XLUtils.readData(getFolderPath, 'Org-Admin', r, 1)
                    self.unit_website = XLUtils.readData(getFolderPath, 'Org-Admin', r, 2)
                    self.unit_phone_number = XLUtils.readData(getFolderPath, 'Org-Admin', r, 3)
                    self.unit_first_name = XLUtils.readData(getFolderPath, 'Org-Admin', r, 4)
                    self.unit_last_name = XLUtils.readData(getFolderPath, 'Org-Admin', r, 5)
                    self.unit_email = XLUtils.readData(getFolderPath, 'Org-Admin', r, 6)
                    self.unit_contact_phone_number = XLUtils.readData(getFolderPath, 'Org-Admin', r, 7)
                    self.org_admin_unit.clickBuisnessUnit()
                    sleep(2)
                    self.org_admin_unit.clickcreateBuisnessUnit()
                    sleep(2)
                    self.org_admin_unit.setUnitName(self.unit_name)
                    sleep(2)
                    self.org_admin_unit.setUnitWebsite(self.unit_website)
                    sleep(2)
                    self.org_admin_unit.setUnitPhoneno(self.unit_phone_number)
                    sleep(2)
                    self.org_admin_unit.setUnitFirstName(self.unit_first_name)
                    sleep(2)
                    self.org_admin_unit.setUnitLastName(self.unit_last_name)
                    sleep(2)
                    self.org_admin_unit.setUnitEmail(self.unit_email)
                    sleep(2)
                    self.org_admin_unit.setUnitContactPhoneNo(self.unit_contact_phone_number)
                    self.org_admin_unit.clickunitActiveCheckbox()
                    self.org_admin_unit.clickcreatbusinessUnitbtn()
                    sleep(6)
                    act_title_create_business_units = self.driver.title
                    exp_title_create_business_units = "Business units"
                    if act_title_create_business_units == exp_title_create_business_units:
                        book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A21'] = "test-business-units-add"
                        ws['B21'] = "Pass"
                        book.save('Reports/result_org_admin_cases_output.xlsx')
                        self.driver.close()
                    else:
                        book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A21'] = "test-business-units-add"
                        ws['B21'] = "Fail"
                        book.save('Reports/result_org_admin_cases_output.xlsx')
                        self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_org_admin_cases_output.xlsx')
                self.driver.close()


    def test_org_admin_business_unit_update(self):
        if self.excelReader_test_cases('test_org_admin_business_unit_update', 'test_cases') == "success":
            act_title = self.check_org_admin_user_login()
            exp_title = "User Dashboard"
            if act_title == exp_title:
                book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_org_admin_cases_output.xlsx')
                path = "assets"
                fileName = "data.xlsx"
                getFolderPath = os.path.join(path, fileName)
                self.org_admin_unit = orgadminBusinessUnitClass(self.driver)
                self.rows = XLUtils.getRowCount(getFolderPath, 'Org-Admin')
                for r in range(49, self.rows + 1):
                    self.unit_name = XLUtils.readData(getFolderPath, 'Org-Admin', r, 1)
                    self.unit_website = XLUtils.readData(getFolderPath, 'Org-Admin', r, 2)
                    self.unit_phone_number = XLUtils.readData(getFolderPath, 'Org-Admin', r, 3)
                    self.unit_first_name = XLUtils.readData(getFolderPath, 'Org-Admin', r, 4)
                    self.unit_last_name = XLUtils.readData(getFolderPath, 'Org-Admin', r, 5)
                    self.unit_email = XLUtils.readData(getFolderPath, 'Org-Admin', r, 6)
                    self.unit_contact_phone_number = XLUtils.readData(getFolderPath, 'Org-Admin', r, 7)
                    self.org_admin_unit.clickBuisnessUnit()
                    sleep(2)
                    try:
                        if self.driver.find_element_by_class_name("unit").text == 'Automation-business-unit-add':
                            self.org_admin_unit.clickunitEditBtn()
                        else:
                            sleep(4)
                            book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                            ws = book.active
                            ws['A22'] = "test-business-units-update"
                            ws['B22'] = "No data found"
                            book.save('Reports/result_admin_cases_output.xlsx')
                            # self.add_s3bucket_file_upload()
                            self.driver.close()
                    except NoSuchElementException:
                        sleep(4)
                        book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A22'] = "test-business-units-update"
                        ws['B22'] = "No data found"
                        book.save('Reports/result_org_admin_cases_output.xlsx')
                        self.driver.close()
                    sleep(2)
                    self.org_admin_unit.setUnitName(self.unit_name)
                    sleep(2)
                    self.org_admin_unit.setUnitWebsite(self.unit_website)
                    sleep(2)
                    self.org_admin_unit.setUnitPhoneno(self.unit_phone_number)
                    sleep(2)
                    self.org_admin_unit.setUnitFirstName(self.unit_first_name)
                    sleep(2)
                    self.org_admin_unit.setUnitLastName(self.unit_last_name)
                    sleep(2)
                    self.org_admin_unit.setUnitContactPhoneNo(self.unit_contact_phone_number)
                    self.org_admin_unit.clickunitActiveCheckbox()
                    self.org_admin_unit.clickunitUpdateBtn()
                    sleep(6)
                    act_title_update_business_units = self.driver.title
                    exp_title_update_business_units = "Business units"
                    if act_title_update_business_units == exp_title_update_business_units:
                        book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A22'] = "test-business-units-update"
                        ws['B22'] = "Pass"
                        book.save('Reports/result_org_admin_cases_output.xlsx')
                        self.driver.close()
                    else:
                        book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A22'] = "test-business-units-update"
                        ws['B22'] = "Fail"
                        book.save('Reports/result_org_admin_cases_output.xlsx')
                        self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_org_admin_cases_output.xlsx')
                self.driver.close()

    def test_org_admin_business_unit_delete(self):
      if self.excelReader_test_cases('test_org_admin_business_unit_delete', 'test_cases') == "success":
        act_title = self.check_org_admin_user_login()
        exp_title = "User Dashboard"
        if act_title == exp_title:
            book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
            ws = book.active
            ws['A2'] = "test-login"
            ws['B2'] = "Pass"
            book.save('Reports/result_org_admin_cases_output.xlsx')
            self.org_admin_unit = orgadminBusinessUnitClass(self.driver)
            sleep(2)
            self.org_admin_unit.clickBuisnessUnit()
            try:
                if self.driver.find_element_by_class_name("unit").text == 'Automation-business-unit-update':
                    self.org_admin_unit.clickRemoveunitLink()
                else:
                    sleep(4)
                    book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                    ws = book.active
                    ws['A23'] = "test-business-units-delete"
                    ws['B23'] = "No data Found"
                    book.save('Reports/result_admin_cases_output.xlsx')
                    # self.add_s3bucket_file_upload()
                    self.driver.close()
                sleep(2)
                self.org_admin_unit.clickConfirmDelete()
                sleep(4)
                book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                ws = book.active
                ws['A23'] = "test-business-units-delete"
                ws['B23'] = "Pass"
                book.save('Reports/result_org_admin_cases_output.xlsx')
                self.driver.close()
            except NoSuchElementException:
                book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                ws = book.active
                ws['A23'] = "test-business-units-delete"
                ws['B23'] = "No data Found"
                book.save('Reports/result_org_admin_cases_output.xlsx')
                self.driver.close()
        else:
            book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
            ws = book.active
            ws['A2'] = "test-login"
            ws['B2'] = "Fail"
            book.save('Reports/result_org_admin_cases_output.xlsx')
            self.driver.close()

    def test_org_admin_message_merge_add(self):
     if self.excelReader_test_cases('test_org_admin_message_merge_add', 'test_cases') == "success":
        act_title = self.check_org_admin_user_login()
        exp_title = "User Dashboard"
        if act_title == exp_title:
            book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
            ws = book.active
            ws['A2'] = "test-login"
            ws['B2'] = "Pass"
            book.save('Reports/result_org_admin_cases_output.xlsx')
            path = "assets"
            fileName = "data.xlsx"
            getFolderPath = os.path.join(path, fileName)
            self.org_admin_merge = orgadminMessagemergeClass(self.driver)
            self.rows = XLUtils.getRowCount(getFolderPath, 'Org-Admin')
            for r in range(54, self.rows + 1):
                self.title = XLUtils.readData(getFolderPath, 'Org-Admin', r, 1)
                self.short_description = XLUtils.readData(getFolderPath, 'Org-Admin', r, 2)
                self.full_description = XLUtils.readData(getFolderPath, 'Org-Admin', r, 3)
                self.org_admin_merge.clickMessageMerge()
                sleep(2)
                self.org_admin_merge.clickCreateMessageMerge()
                sleep(2)
                self.org_admin_merge.setMessageMergeKey(self.title)
                sleep(2)
                self.org_admin_merge.setMessageShortDes(self.short_description)
                sleep(2)
                self.org_admin_merge.setMessageFulltDes(self.full_description)
                sleep(2)
                self.org_admin_merge.clickMessagemergesbtn()
                sleep(4)
                act_title_add_message_merge = self.driver.title
                exp_title_add_message_merge = "Message Merges"
                if act_title_add_message_merge == exp_title_add_message_merge:
                    book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                    ws = book.active
                    ws['A24'] = "test-message-merge-add"
                    ws['B24'] = "Pass"
                    book.save('Reports/result_org_admin_cases_output.xlsx')
                    self.driver.close()
                else:
                    book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                    ws = book.active
                    ws['A24'] = "test-message-merge-add"
                    ws['B24'] = "Fail"
                    book.save('Reports/result_org_admin_cases_output.xlsx')
                    self.driver.close()
        else:
            book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
            ws = book.active
            ws['A2'] = "test-login"
            ws['B2'] = "Fail"
            book.save('Reports/result_org_admin_cases_output.xlsx')
            self.driver.close()



    def test_org_admin_message_merge_update(self):
     if self.excelReader_test_cases('test_org_admin_message_merge_update', 'test_cases') == "success":
        act_title = self.check_org_admin_user_login()
        exp_title = "User Dashboard"
        if act_title == exp_title:
            book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
            ws = book.active
            ws['A2'] = "test-login"
            ws['B2'] = "Pass"
            book.save('Reports/result_org_admin_cases_output.xlsx')
            path = "assets"
            fileName = "data.xlsx"
            getFolderPath = os.path.join(path, fileName)
            self.org_admin_merge = orgadminMessagemergeClass(self.driver)
            self.rows = XLUtils.getRowCount(getFolderPath, 'Org-Admin')
            for r in range(59, self.rows + 1):
                self.title = XLUtils.readData(getFolderPath, 'Org-Admin', r, 1)
                self.short_description = XLUtils.readData(getFolderPath, 'Org-Admin', r, 2)
                self.full_description = XLUtils.readData(getFolderPath, 'Org-Admin', r, 3)
                self.org_admin_merge.clickMessageMerge()
                sleep(2)
                try:
                    if self.driver.find_element_by_class_name("title").text == 'Automation-message-merge-add':
                        self.org_admin_merge.clickMessageMergeEditBtn()
                    else:
                        sleep(4)
                        book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A25'] = "test-message-merge-update"
                        ws['B25'] = "No data found"
                        book.save('Reports/result_org_admin_cases_output.xlsx')
                        # self.add_s3bucket_file_upload()
                        self.driver.close()
                except NoSuchElementException:
                    sleep(4)
                    book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                    ws = book.active
                    ws['A25'] = "test-message-merge-update"
                    ws['B25'] = "No data found"
                    book.save('Reports/result_org_admin_cases_output.xlsx')
                    self.driver.close()
                sleep(2)
                self.org_admin_merge.setMessageMergeKey(self.title)
                sleep(2)
                self.org_admin_merge.setMessageShortDes(self.short_description)
                sleep(2)
                self.org_admin_merge.setMessageFulltDes(self.full_description)
                sleep(2)
                self.org_admin_merge.click_msg_merge_update_btn_class()
                sleep(6)
                act_title_update_message_merge = self.driver.title
                exp_title_update_message_merge = "Message Merges"
                if act_title_update_message_merge == exp_title_update_message_merge:
                    book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                    ws = book.active
                    ws['A25'] = "test-message-merge-update"
                    ws['B25'] = "Pass"
                    book.save('Reports/result_org_admin_cases_output.xlsx')
                    self.driver.close()
                else:
                    book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                    ws = book.active
                    ws['A25'] = "test-message-merge-update"
                    ws['B25'] = "Fail"
                    book.save('Reports/result_org_admin_cases_output.xlsx')
                    self.driver.close()
        else:
            book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
            ws = book.active
            ws['A2'] = "test-login"
            ws['B2'] = "Fail"
            book.save('Reports/result_org_admin_cases_output.xlsx')
            self.driver.close()

    def test_org_admin_message_merge_delete(self):
        if self.excelReader_test_cases('test_org_admin_message_merge_delete', 'test_cases') == "success":
            act_title = self.check_org_admin_user_login()
            exp_title = "User Dashboard"
            if act_title == exp_title:
                book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_org_admin_cases_output.xlsx')
                self.org_admin_merge = orgadminMessagemergeClass(self.driver)
                sleep(2)
                self.org_admin_merge.clickMessageMerge()
                try:
                    if self.driver.find_element_by_class_name("title").text == 'Automation-message-merge-update':
                        self.org_admin_merge.clickRemoveMessageMerge()
                    else:
                        sleep(4)
                        book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A26'] = "test-message-merge-delete"
                        ws['B26'] = "No data found"
                        book.save('Reports/result_org_admin_cases_output.xlsx')
                        # self.add_s3bucket_file_upload()
                        self.driver.close()
                    self.org_admin_merge.clickConfirmDelete()
                    sleep(6)
                    book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                    ws = book.active
                    ws['A26'] = "test-message-merge-delete"
                    ws['B26'] = "Pass"
                    book.save('Reports/result_org_admin_cases_output.xlsx')
                    self.driver.close()
                except NoSuchElementException:
                    book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                    ws = book.active
                    ws['A26'] = "test-message-merge-delete"
                    ws['B26'] = "Fail"
                    book.save('Reports/result_org_admin_cases_output.xlsx')
                    self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_org_admin_cases_output.xlsx')
                self.driver.close()



    def test_org_admin_message_merge_parser(self):
      if self.excelReader_test_cases('test_org_admin_message_merge_parser', 'test_cases') == "success":
        act_title = self.check_org_admin_user_login()
        exp_title = "User Dashboard"
        if act_title == exp_title:
            book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
            ws = book.active
            ws['A2'] = "test-login"
            ws['B2'] = "Pass"
            book.save('Reports/result_org_admin_cases_output.xlsx')
            path = "assets"
            fileName = "data.xlsx"
            getFolderPath = os.path.join(path, fileName)
            self.org_admin_parser = orgadminMessagemergeparserClass(self.driver)
            self.rows = XLUtils.getRowCount(getFolderPath, 'Org-Admin')
            for r in range(64, self.rows + 1):
                self.string = XLUtils.readData(getFolderPath, 'Org-Admin', r, 1)
                self.org_admin_parser.clickMessageMergeParser()
                self.org_admin_parser.setMessageMergeParserName(self.string)
                sleep(2)
                self.org_admin_parser.clickMessagemergesparserbtn()
                sleep(5)
                try:
                    success_toastr = self.driver.find_element_by_class_name("toast-success")
                    if success_toastr:
                        book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A27'] = "test-message-merge-parser"
                        ws['B27'] = "Pass"
                        book.save('Reports/result_org_admin_cases_output.xlsx')
                        self.driver.close()
                except NoSuchElementException:
                    book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                    ws = book.active
                    ws['A27'] = "test-message-merge-parser"
                    ws['B27'] = "Fail"
                    book.save('Reports/result_org_admin_cases_output.xlsx')
                    self.driver.close()
        else:
            book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
            ws = book.active
            ws['A2'] = "test-login"
            ws['B2'] = "Fail"
            book.save('Reports/result_org_admin_cases_output.xlsx')
            self.driver.close()






    def test_org_admin_conversation_history(self):
      if self.excelReader_test_cases('test_org_admin_conversation_history', 'test_cases') == "success":
        act_title = self.check_org_admin_user_login()
        exp_title = "User Dashboard"
        if act_title == exp_title:
            book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
            ws = book.active
            ws['A2'] = "test-login"
            ws['B2'] = "Pass"
            book.save('Reports/result_org_admin_cases_output.xlsx')
            self.org_admin_conversation = orgadminConversationhistoryClass(self.driver)
            sleep(2)
            self.org_admin_conversation.clickconversationhistorybtn()
            sleep(2)
            book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
            ws = book.active
            ws['A28'] = "test-conversation-history"
            ws['B28'] = "Pass"
            book.save('Reports/result_org_admin_cases_output.xlsx')
            self.driver.close()
        else:
              book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
              ws = book.active
              ws['A2'] = "test-login"
              ws['B2'] = "Fail"
              book.save('Reports/result_org_admin_cases_output.xlsx')
              self.driver.close()

    def test_org_admin_common_response_add(self):
        if self.excelReader_test_cases('test_org_admin_common_response_add', 'test_cases') == "success":
            act_title = self.check_org_admin_user_login()
            exp_title = "User Dashboard"
            if act_title == exp_title:
                book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_org_admin_cases_output.xlsx')
                path = "assets"
                fileName = "data.xlsx"
                getFolderPath = os.path.join(path, fileName)
                self.org_admin_common_response = orgadminCommonResponseClass(self.driver)
                self.rows = XLUtils.getRowCount(getFolderPath, 'Org-Admin')
                for r in range(69, self.rows + 1):
                    self.title = XLUtils.readData(getFolderPath, 'Org-Admin', r, 1)
                    self.synonyms = XLUtils.readData(getFolderPath, 'Org-Admin', r, 2)
                    self.message = XLUtils.readData(getFolderPath, 'Org-Admin', r, 3)
                    self.org_admin_common_response.clickCommonResponse()
                    sleep(2)
                    self.org_admin_common_response.clickCreateCommonResponse()
                    sleep(2)
                    self.org_admin_common_response.setResponse(self.title)
                    sleep(2)
                    self.org_admin_common_response.setCommonResponsesynonyms(self.synonyms)
                    sleep(2)
                    self.org_admin_common_response.setResponseMessage(self.message)
                    sleep(2)
                    self.org_admin_common_response.clickuploadimage()
                    sleep(2)
                    self.org_admin_common_response.clickCreatecommonResponseBtn()
                    sleep(6)
                    act_title_create_common_responses = self.driver.title
                    exp_title_create_common_responses = "Common Responses"
                    if act_title_create_common_responses == exp_title_create_common_responses:
                        book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A29'] = "test-common-response-add"
                        ws['B29'] = "Pass"
                        book.save('Reports/result_org_admin_cases_output.xlsx')
                        self.driver.close()
                    else:
                        book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A29'] = "test-common-response-add"
                        ws['B29'] = "Fail"
                        book.save('Reports/result_org_admin_cases_output.xlsx')
                        self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_org_admin_cases_output.xlsx')
                self.driver.close()


    def test_org_admin_common_response_update(self):
        if self.excelReader_test_cases('test_org_admin_common_response_update', 'test_cases') == "success":
            act_title = self.check_org_admin_user_login()
            exp_title = "User Dashboard"
            if act_title == exp_title:
                book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_org_admin_cases_output.xlsx')
                path = "assets"
                fileName = "data.xlsx"
                getFolderPath = os.path.join(path, fileName)
                self.org_admin_common_response = orgadminCommonResponseClass(self.driver)
                self.rows = XLUtils.getRowCount(getFolderPath, 'Org-Admin')
                for r in range(74, self.rows + 1):
                    self.title = XLUtils.readData(getFolderPath, 'Org-Admin', r, 1)
                    self.synonyms = XLUtils.readData(getFolderPath, 'Org-Admin', r, 2)
                    self.message = XLUtils.readData(getFolderPath, 'Org-Admin', r, 3)
                    self.org_admin_common_response.clickCommonResponse()
                    sleep(2)
                    try:
                        if self.driver.find_element_by_class_name("response").text == 'Automation-common-responses-add':
                            self.org_admin_common_response.clickEditcommonResponseBtn()
                        else:
                            sleep(4)
                            book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                            ws = book.active
                            ws['A30'] = "test-common-response-update"
                            ws['B30'] = "No data found"
                            book.save('Reports/result_org_admin_cases_output.xlsx')
                            # self.add_s3bucket_file_upload()
                            self.driver.close()
                    except NoSuchElementException:
                        sleep(4)
                        book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A30'] = "test-common-response-update"
                        ws['B30'] = "No data found"
                        book.save('Reports/result_org_admin_cases_output.xlsx')
                        self.driver.close()
                    sleep(2)
                    self.org_admin_common_response.setResponse(self.title)
                    sleep(2)
                    self.org_admin_common_response.setResponseMessage(self.message)
                    sleep(2)
                    self.org_admin_common_response.clickuploadimage()
                    sleep(2)
                    self.org_admin_common_response.clickUpdatecommonResponseBtn()
                    sleep(6)
                    act_title_update_common_responses = self.driver.title
                    exp_title_update_common_responses = "Common Responses"
                    if act_title_update_common_responses == exp_title_update_common_responses:
                        book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A30'] = "test-common-response-update"
                        ws['B30'] = "Pass"
                        book.save('Reports/result_org_admin_cases_output.xlsx')
                        self.driver.close()
                    else:
                        book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A30'] = "test-common-response-update"
                        ws['B30'] = "Fail"
                        book.save('Reports/result_org_admin_cases_output.xlsx')
                        self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_org_admin_cases_output.xlsx')
                self.driver.close()






    def test_org_admin_new_response(self):
     if self.excelReader_test_cases('test_org_admin_new_response', 'test_cases') == "success":
        act_title = self.check_org_admin_user_login()
        exp_title = "User Dashboard"
        if act_title == exp_title:
            book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
            ws = book.active
            ws['A2'] = "test-login"
            ws['B2'] = "Pass"
            book.save('Reports/result_org_admin_cases_output.xlsx')
            self.org_admin_new_response = orgadminnewresponseClass(self.driver)
            sleep(2)
            self.org_admin_new_response.clicknewresponse()
            sleep(3)
            book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
            ws = book.active
            ws['A31'] = "test-new-response"
            ws['B31'] = "Pass"
            book.save('Reports/result_org_admin_cases_output.xlsx')
            self.driver.close()
        else:
            book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
            ws = book.active
            ws['A2'] = "test-login"
            ws['B2'] = "Fail"
            book.save('Reports/result_org_admin_cases_output.xlsx')
            self.driver.close()




    def test_org_admin_past_response(self):
        if self.excelReader_test_cases('test_org_admin_past_response', 'test_cases') == "success":
            act_title = self.check_org_admin_user_login()
            exp_title = "User Dashboard"
            if act_title == exp_title:
                book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_org_admin_cases_output.xlsx')
                self.org_admin_past_response = orgadminpastresponseClass(self.driver)
                sleep(2)
                self.org_admin_past_response.clickpastresponse()
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                ws = book.active
                ws['A32'] = "test-past-response"
                ws['B32'] = "Pass"
                book.save('Reports/result_org_admin_cases_output.xlsx')
                self.driver.close()
        else:
                book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_org_admin_cases_output.xlsx')
                self.driver.close()


    def test_org_admin_ignore_response(self):
      if self.excelReader_test_cases('test_org_admin_ignore_response', 'test_cases') == "success":
        act_title = self.check_org_admin_user_login()
        exp_title = "User Dashboard"
        if act_title == exp_title:
            book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
            ws = book.active
            ws['A2'] = "test-login"
            ws['B2'] = "Pass"
            book.save('Reports/result_org_admin_cases_output.xlsx')
            self.org_admin_ignore_response = orgadminignoreresponseClass(self.driver)
            sleep(2)
            self.org_admin_ignore_response.clickignoreresponse()
            sleep(6)
            book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
            ws = book.active
            ws['A33'] = "test-ignore-response"
            ws['B33'] = "Pass"
            book.save('Reports/result_org_admin_cases_output.xlsx')
            self.driver.close()
        else:
            book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
            ws = book.active
            ws['A2'] = "test-login"
            ws['B2'] = "Fail"
            book.save('Reports/result_org_admin_cases_output.xlsx')
            self.driver.close()

    def test_org_admin_manage_profile(self):
      if self.excelReader_test_cases('test_org_admin_manage_profile', 'test_cases') == "success":
        act_title = self.check_org_admin_user_login()
        exp_title = "User Dashboard"
        if act_title == exp_title:
            book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
            ws = book.active
            ws['A2'] = "test-login"
            ws['B2'] = "Pass"
            book.save('Reports/result_org_admin_cases_output.xlsx')
            self.org_admin_profile = orgadminManageProfileClass(self.driver)
            sleep(2)
            self.org_admin_profile.click_manage_profile_btn()
            path = "assets"
            fileName = "data.xlsx"
            getFolderPath = os.path.join(path, fileName)
            self.rows = XLUtils.getRowCount(getFolderPath, 'Org-Admin')
            for r in range(79, self.rows + 1):
                self.first_name = XLUtils.readData(getFolderPath, 'Org-Admin', r, 1)
                self.last_name = XLUtils.readData(getFolderPath, 'Org-Admin', r, 2)
                self.phone_number = XLUtils.readData(getFolderPath, 'Org-Admin', r, 3)
                self.group_assignment = XLUtils.readData(getFolderPath, 'Org-Admin', r, 4)
                self.current_password = XLUtils.readData(getFolderPath, 'Org-Admin', r, 5)
                self.new_password = XLUtils.readData(getFolderPath, 'Org-Admin', r, 6)
                self.confirm_password = XLUtils.readData(getFolderPath, 'Org-Admin', r, 7)
                self.org_admin_profile.set_manage_profile_name(self.first_name)
                self.org_admin_profile.set_manage_profile_last_name(self.last_name)
                self.org_admin_profile.set_manage_profile_phone_number(self.phone_number)
                self.org_admin_profile.click_manage_profile_dob()
                self.org_admin_profile.set_manage_group_assignment(self.group_assignment)
                self.org_admin_profile.manage_profile_click_state()
                self.org_admin_profile.set_manage_profile_current_password(self.current_password)
                self.org_admin_profile.set_manage_profile_new_password(self.new_password)
                self.org_admin_profile.set_manage_profile_confirm_password(self.confirm_password)
                self.org_admin_profile.click_update_account_btn()
                sleep(4)
                try:
                    success_toastr = self.driver.find_element_by_class_name("toast-error")
                    if success_toastr:
                        book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A34'] = "test-user-profile"
                        ws['B34'] = "No data found"
                        book.save('Reports/result_org_admin_cases_output.xlsx')
                        self.driver.close()
                except NoSuchElementException:
                    book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
                    ws = book.active
                    ws['A34'] = "test-user-profile"
                    ws['B34'] = "Pass"
                    book.save('Reports/result_org_admin_cases_output.xlsx')
                    self.driver.close()
        else:
            book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
            ws = book.active
            ws['A2'] = "test-login"
            ws['B2'] = "Fail"
            book.save('Reports/result_org_admin_cases_output.xlsx')
            self.driver.close()

    def test_excel_empty_row_remove(self):
        book = openpyxl.load_workbook('Reports/result_org_admin_cases_output.xlsx')
        ws = book.active
        index_row = []
        for i in range(1, ws.max_row):
            if ws.cell(i, 1).value is None:
                index_row.append(i)
        for row_del in range(len(index_row)):
            ws.delete_rows(idx=index_row[row_del], amount=1)
            index_row = list(map(lambda k: k - 1, index_row))
        book.save('Reports/result_org_admin_cases_output.xlsx')


    def test_s3bucket_file_upload(self):
        client_s3 = boto3.client('s3', aws_access_key_id=aws_access_key_id,
                                 aws_secret_access_key=aws_secret_access_key)
        with open('Reports/result_org_admin_cases_output.xlsx', 'rb') as data:
            client_s3.upload_fileobj(data, bucket_name, 'Reports/result_org_admin' + str(date) + '.xlsx')