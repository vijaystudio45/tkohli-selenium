import pytest
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from webdriver_manager.chrome import ChromeDriverManager

from pageObjects.LoginPage import LoginClass
from pageObjects.ContractPage import ContractClass
from pageObjects.ReportSchedulerHistory import ReportSchedulerHistoryClass
from pageObjects.ClicksPage import ClickClass
from pageObjects.PastReportPage import PastReportClass
from pageObjects.PastReportErrorItemsPage import PastReportErrorItemsClass
from pageObjects.ReportSchedulerPage import ReportSchedulerClass
from pageObjects.PermissionPage import PermissionClass
from pageObjects.CampaignPage import CampaignClass
from pageObjects.Ignore_link import IgnoreLinkClass
from pageObjects.Shorten_link import Shorten_link_Class
from pageObjects.NewReponsePage import NewResponseClass
from pageObjects.PastResponsePage import PastResponseClass
from pageObjects.IgnoreResponsePage import IgnoreResponseClass
from pageObjects.Organizations import OrganizationsClass
from pageObjects.Vcard import VcardClass
from pageObjects.business_units import BusinessUnitClass
from pageObjects.Longcodes import LongcodesClass
from pageObjects.Cron_Scheduler import CronSchedulerClass
from pageObjects.Profile import ProfileClass
from pageObjects.Cron_jobs import CronJobsClass
from pageObjects.Ach_Setup import AchSetupClass
from pageObjects.Api_keys import ApiKeyClass
from pageObjects.Artiva_import import ArtivaimportClass
from pageObjects.Delimited_import import delimitedimportClass
from pageObjects.Sending_message_process import SendingClass
from pageObjects.Reports import ReportsClass
from pageObjects.MessageMergeParser import MessagemergeparserClass
from pageObjects.MessageMerge import MessagemergeClass
from pageObjects.ImportCsvPage import ImportCsvClass
from pageObjects.LogFilePage import LogFileClass
from pageObjects.SendingDemoPage import SendingDemoClass
from pageObjects.ZipwhipSendingDemoPage import ZipwhipSendingDemoClass
from pageObjects.TelnyxAPIDemoPage import TelnyxApiDemoClass
from pageObjects.AssignMessagesPage import AssignMessageClass
from pageObjects.MessageRecyclingPage import MessageRecyclingClass
from pageObjects.SendingMessagesPage import SendingMessagesClass
from utilities.readProperties import ReadConfig
from time import sleep
from selenium.webdriver.remote.webelement import WebElement
from pageObjects.TodaysCampaignsPage import TodayCampaignsClass
from pageObjects.ProcessedFiles import ProcessedFilesClass
from pageObjects.SenderClaims import SenderClaimsClass
from pageObjects.Mapped_media import MappedmediaClass
from pageObjects.Conversation_history import ConversationhistoryClass
from pageObjects.Manage_messages import ManagemessagesClass
from pageObjects.Friendly_error_messages import FriendlyerrorClass
from pageObjects.Phone_lookup_rules import PhoneLookupRuleClass
from pageObjects.SignalWireDemo import SignalWireDemoClass
from pageObjects.file_import import FileimportClass
from pageObjects.Common_response import CommonResponseClass
from pageObjects.Signup import SignupClass
from pageObjects.Loginpage_sender import LoginsenderClass
from pageObjects.sender_manage_profile import SenderProfileClass
from pageObjects.Sender_user_ach_setup import SenderUserachSetupClass
from pageObjects.simple_user_LoginPage import SimpleuserLoginClass
from pageObjects.Simple_user_profile import SimpleUserprofileClass
from pageObjects.Chat_agent_user_LoginPage import ChatagentLoginClass
from pageObjects.chatagent_new_response import chatagentnewresponseClass
from pageObjects.chatagent_past_response import chatagentpastresponseClass
from pageObjects.chatagent_ignore_response import chatagentignoreresponseClass
from pageObjects.chat_agent_user_profile import chatagentUserprofileClass
from pageObjects.orgadmin_user_Loginpage import OrgadminLoginClass
from pageObjects.orgadmin_organization import orgadmin_OrganizationsClass
from pageObjects.orgadmin_vcard import orgadmin_vcardclass
from pageObjects.MessageSendersPage import MessageSendersClass
from pageObjects.SendingMessagesAdminPage import SendingMessagesAdminClass
from pageObjects.File_Ignore_Page import FileIgnoreClass
import boto3
from utilities.Customlogger import LogGen
import os
from dotenv import load_dotenv

load_dotenv()
# aws detail

aws_access_key_id = os.getenv('AWS_ACCESS_KEY')
aws_secret_access_key = os.getenv('AWS_SECRET_KEY')
bucket_name = os.getenv('AWS_BUCKET_NAME')
from utilities import XLUtils

from datetime import date

date = date.today()
from openpyxl.workbook import Workbook
from openpyxl.styles import Font
import openpyxl


class Test_001_Admin:
    Organizations_telnyx_key = ReadConfig.getOrganizationstelnyxkey()
    logger = LogGen.loggen()

    def excelReader_test_cases(self, method, sheet):
        path = "assets"
        fileName = "test_cases.xlsx"
        getFolderPath = os.path.join(path, fileName)
        self.rows = XLUtils.getRowCount(getFolderPath, sheet)
        for r in range(2, self.rows + 1):
            self.testCase = XLUtils.readData(getFolderPath, sheet, r, 1)
            self.action = XLUtils.readData(getFolderPath, sheet, r, 3)
            if self.testCase == method and self.action == 'Y':
                return "success"

    def test_excel_file_data_clear(self):
        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
        sheet = book['Admin-Testcases-Output']
        sheet.delete_rows(2, 200)
        book.save('Reports/result_admin_cases_output.xlsx')

    def check_login(self):
        self.driver = webdriver.Chrome(ChromeDriverManager().install())
        self.driver.maximize_window()
        sleep(1)
        path = "assets"
        fileName = "data.xlsx"
        getFolderPath = os.path.join(path, fileName)
        self.lp = LoginClass(self.driver)
        self.rows = XLUtils.getRowCount(getFolderPath, 'Admin')
        for r in range(4, self.rows + 1):
            self.login_url = XLUtils.readData(getFolderPath, 'Admin', r, 1)
            self.email = XLUtils.readData(getFolderPath, 'Admin', r, 2)
            self.password = XLUtils.readData(getFolderPath, 'Admin', r, 3)
            self.driver.get(self.login_url)
            sleep(1)
            self.lp.setEmailName(self.email)
            sleep(2)
            self.lp.setPassword(self.password)
            sleep(2)
            self.lp.clickLogin()
            sleep(10)
            act_title = self.driver.title
            return act_title








    def test_login(self):
        if self.excelReader_test_cases('test_login', 'test_cases') == "success":
                act_title = self.check_login()
                exp_title = "Admin Dashboard"
                if act_title == exp_title:
                    sleep(8)
                    wb = Workbook()
                    ws = wb.active
                    ws['A1'] = "Test Cases"
                    ws['A1'].font = Font(name="Arial",b=True)
                    ws['B1'] = "Status"
                    ws['B1'].font = Font(name="Arial",b=True)
                    ws.title = "Admin-Testcases-Output"
                    ws['A2'] = "test-login"
                    ws['B2'] = "Pass"
                    wb.save('Reports/result_admin_cases_output.xlsx')
                    sleep(4)
                    #self.add_s3bucket_file_upload()
                    self.driver.close()
                    assert True
                else:
                    wb = Workbook()
                    ws = wb.active
                    ws['A1'] = "Test Cases"
                    ws['A1'].font = Font(name="Arial", b=True)
                    ws['B1'] = "Status"
                    ws['B1'].font = Font(name="Arial", b=True)
                    ws.title = "Admin-Testcases-Output"
                    ws['A2'] = "test-login"
                    ws['B2'] = "Fail"
                    wb.save('Reports/result_admin_cases_output.xlsx')
                    self.driver.close()
                    #self.add_s3bucket_file_upload()


    def test_add_contract(self):
        if self.excelReader_test_cases('test_add_contract', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #s3_file_upload = self.add_s3bucket_file_upload()
                sleep(8)
                path = "assets"
                fileName = "data.xlsx"
                getFolderPath = os.path.join(path, fileName)
                self.rows = XLUtils.getRowCount(getFolderPath, 'Admin')
                for r in range(9, self.rows + 1):
                    self.contract_name = XLUtils.readData(getFolderPath, 'Admin', r, 1)
                    self.pay_rate = XLUtils.readData(getFolderPath, 'Admin', r, 2)
                    self.text_area = XLUtils.readData(getFolderPath, 'Admin', r, 3)
                    self.cont_obj = ContractClass(self.driver)
                    self.cont_obj.clickviewContract()
                    sleep(2)
                    self.cont_obj.clickCreateContract()
                    sleep(2)
                    self.cont_obj.setContractName(self.contract_name)
                    sleep(2)
                    self.cont_obj.setPayRate(self.pay_rate)
                    sleep(2)
                    self.cont_obj.setContractTextArea(self.text_area)
                    self.cont_obj.clickActiveCheckbox()
                    sleep(2)
                    self.cont_obj.clickSaveContract()
                    sleep(10)
                    act_title_add = self.driver.title
                    exp_title_test = "Contract"
                    if act_title_add == exp_title_test:
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A3'] = "test-add-contract"
                        ws['B3'] = "Pass"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #s3_file_upload = self.add_s3bucket_file_upload()
                        self.driver.close()
                    else:
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A3'] = "test-add-contract"
                        ws['B3'] = "Fail"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #s3_file_upload = self.add_s3bucket_file_upload()
                        self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                #s3_file_upload = self.add_s3bucket_file_upload()
                self.driver.close()

    def test_update_contract(self):
        if self.excelReader_test_cases('test_update_contract', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #self.add_s3bucket_file_upload()
                sleep(8)
                path = "assets"
                fileName = "data.xlsx"
                getFolderPath = os.path.join(path, fileName)
                self.rows = XLUtils.getRowCount(getFolderPath, 'Admin')
                for r in range(14, self.rows + 1):
                    self.contract_name = XLUtils.readData(getFolderPath, 'Admin', r, 1)
                    self.pay_rate = XLUtils.readData(getFolderPath, 'Admin', r, 2)
                    self.text_area = XLUtils.readData(getFolderPath, 'Admin', r, 3)
                    self.cont_obj = ContractClass(self.driver)
                    self.cont_obj.clickviewContract()
                    try:
                        if self.driver.find_element_by_class_name("version_tag").text == 'Automation-Contract-added':
                            self.cont_obj.clickeditContract()
                        else:
                            sleep(4)
                            book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                            ws = book.active
                            ws['A4'] = "test-update-contract"
                            ws['B4'] = "Contract Not found"
                            book.save('Reports/result_admin_cases_output.xlsx')
                            # self.add_s3bucket_file_upload()
                            self.driver.close()
                    except NoSuchElementException:
                        sleep(4)
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A4'] = "test-update-contract"
                        ws['B4'] = "No data found"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #self.add_s3bucket_file_upload()
                        self.driver.close()
                    sleep(2)
                    self.cont_obj.setContractName(self.contract_name)
                    sleep(2)
                    self.cont_obj.setPayRate(self.pay_rate)
                    sleep(2)
                    self.cont_obj.setContractTextArea(self.text_area)
                    self.cont_obj.clickActiveCheckbox()
                    sleep(2)
                    self.cont_obj.clickUpdateContract()
                    sleep(10)
                    act_title_update = self.driver.title
                    exp_title_test = "Contract"
                    if act_title_update == exp_title_test:
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A4'] = "test-update-contract"
                        ws['B4'] = "Pass"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #self.add_s3bucket_file_upload()
                        self.driver.close()
                    else:
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A4'] = "test-update-contract"
                        ws['B4'] = "Fail"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #self.add_s3bucket_file_upload()
                        self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                #self.add_s3bucket_file_upload()
                self.driver.close()
                #assert False




    def test_view_reports(self):
        if self.excelReader_test_cases('test_view_reports', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #s3_file_upload = self.add_s3bucket_file_upload()
                self.view_report_obj = ReportsClass(self.driver)
                sleep(2)
                self.view_report_obj.clickReportbtn()
                sleep(2)
                try:
                    self.view_report_obj.clickvewReportcampaignbtn()
                    book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                    ws = book.active
                    ws['A5'] = "test-view-report-campaign"
                    ws['B5'] = "Pass"
                    book.save('Reports/result_admin_cases_output.xlsx')
                    #s3_file_upload = self.add_s3bucket_file_upload()
                    sleep(2)
                    try:
                        self.view_report_obj.clickReportbtn()
                        sleep(2)
                        self.view_report_obj.clickvewReportunsignedcampaignbtn()
                        sleep(2)
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A6'] = "test-view-report-unsigned-campaign"
                        ws['B6'] = "pass"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #s3_file_upload = self.add_s3bucket_file_upload()
                        self.driver.close()
                    except NoSuchElementException:
                        sleep(2)
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A6'] = "test-view-report-unsigned-campaign"
                        ws['B6'] = "No data found"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #s3_file_upload = self.add_s3bucket_file_upload()
                        self.driver.close()
                except NoSuchElementException:
                    book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                    ws = book.active
                    ws['A5'] = "test-view-report-campaign"
                    ws['B5'] = "No data found"
                    book.save('Reports/result_admin_cases_output.xlsx')
                    #s3_file_upload = self.add_s3bucket_file_upload()
                    sleep(2)
                    self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                #s3_file_upload = self.add_s3bucket_file_upload()
                self.driver.close()




    def test_past_report(self):
        if self.excelReader_test_cases('test_past_report', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #s3_file_upload = self.add_s3bucket_file_upload()
                sleep(8)
                self.past_rpt_obj = PastReportClass(self.driver)
                self.past_rpt_obj.clickviewpastreport()
                try:
                    self.driver.find_element_by_class_name('download-file-past-report-from-bucket')
                    self.past_rpt_obj.clickDownloadPastReport()
                    sleep(4)
                    book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                    ws = book.active
                    ws['A7'] = "test-past-report"
                    ws['B7'] = "Pass"
                    book.save('Reports/result_admin_cases_output.xlsx')
                    self.driver.close()
                except NoSuchElementException:
                    book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                    ws = book.active
                    ws['A7'] = "test-past-report"
                    ws['B7'] = "Fail"
                    book.save('Reports/result_admin_cases_output.xlsx')
                    #s3_file_upload = self.add_s3bucket_file_upload()
                    self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                # s3_file_upload = self.add_s3bucket_file_upload()
                self.driver.close()




    def test_past_report_error_items(self):
        if self.excelReader_test_cases('test_past_report_error_items', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #s3_file_upload = self.add_s3bucket_file_upload()
                self.past_rpt_err_itm_obj = PastReportErrorItemsClass(self.driver)
                sleep(2)
                self.past_rpt_err_itm_obj.clickviewpastreporterroritems()
                sleep(1)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A8'] = "test-past-report-error-items"
                ws['B8'] = "pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                self.driver.close()
                #s3_file_upload = self.add_s3bucket_file_upload()
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                #s3_file_upload = self.add_s3bucket_file_upload()
                self.driver.close()











    def test_Organizations_add(self):
        if self.excelReader_test_cases('test_Organizations_add', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #self.add_s3bucket_file_upload()
                fileName = "data.xlsx"
                path = "assets"
                getFolderPath = os.path.join(path, fileName)
                self.rows = XLUtils.getRowCount(getFolderPath, 'Admin')
                sleep(4)
                for r in range(19, self.rows + 1):
                    self.org_name = XLUtils.readData(getFolderPath, 'Admin', r, 1)
                    self.org_website = XLUtils.readData(getFolderPath, 'Admin', r, 2)
                    self.org_phone_number = XLUtils.readData(getFolderPath, 'Admin', r, 3)
                    self.org_first_name = XLUtils.readData(getFolderPath, 'Admin', r, 4)
                    self.org_last_name = XLUtils.readData(getFolderPath, 'Admin', r, 5)
                    self.org_email = XLUtils.readData(getFolderPath, 'Admin', r, 6)
                    self.org_contact_phone_number = XLUtils.readData(getFolderPath, 'Admin', r, 7)
                    self.org_response_service_email = XLUtils.readData(getFolderPath, 'Admin', r, 8)
                    self.org_qc_phone_number = XLUtils.readData(getFolderPath, 'Admin', r, 9)
                    self.org_sftp_host = XLUtils.readData(getFolderPath, 'Admin', r, 10)
                    self.org_sftp_username = XLUtils.readData(getFolderPath, 'Admin', r, 11)
                    self.org_sftp_password = XLUtils.readData(getFolderPath, 'Admin', r, 12)
                    self.org_sftp_file_location = XLUtils.readData(getFolderPath, 'Admin', r, 13)
                    self.org_import_public_root = XLUtils.readData(getFolderPath, 'Admin', r, 14)
                    self.org_import_private_root = XLUtils.readData(getFolderPath, 'Admin', r, 15)
                    self.org_message_webhook_url = XLUtils.readData(getFolderPath, 'Admin', r, 16)
                    self.org_message_webhook_key = XLUtils.readData(getFolderPath, 'Admin', r, 17)
                    self.org_response_webhook_url = XLUtils.readData(getFolderPath, 'Admin', r, 18)
                    self.org_response_webhook_key = XLUtils.readData(getFolderPath, 'Admin', r, 19)
                    self.org_click_webhook_url = XLUtils.readData(getFolderPath, 'Admin', r, 20)
                    self.org_click_webhook_key = XLUtils.readData(getFolderPath, 'Admin', r, 21)
                    self.Org_obj = OrganizationsClass(self.driver)
                    self.Org_obj.clickOrganizations()
                    sleep(2)
                    self.Org_obj.clickCreateOrganizations()
                    sleep(2)
                    self.Org_obj.setorganizationName(self.org_name)
                    sleep(2)
                    self.Org_obj.setorganizationWebsite(self.org_website)
                    self.Org_obj.setorganizationPhonenumber(self.org_phone_number)
                    self.Org_obj.setorganizationFirstName(self.org_first_name)
                    self.Org_obj.setorganizationLastName(self.org_last_name)
                    self.Org_obj.setorganizationEmail(self.org_email)
                    self.Org_obj.setorganizationContactPhoneNo(self.org_contact_phone_number)
                    self.Org_obj.setorganizationResponseServiceEmail(self.org_response_service_email)
                    self.Org_obj.setFrom_Organization_timezone_Dropdown()
                    sleep(2)
                    self.Org_obj.click_qc_btn_class()
                    self.Org_obj.setorganizationqc_Phoneno(self.org_qc_phone_number)
                    sleep(1)
                    self.Org_obj.setorganization_sftp_host(self.org_sftp_host)
                    sleep(1)
                    self.Org_obj.setorganization_sftp_username(self.org_sftp_username)
                    sleep(1)
                    self.Org_obj.setorganization_sftp_Password(self.org_sftp_password)
                    self.Org_obj.setorganization_sftp_FileLocation(self.org_sftp_file_location)
                    self.Org_obj.setorganization_import_root_public(self.org_import_public_root)
                    self.Org_obj.setorganization_import_root_private(self.org_import_private_root)
                    self.Org_obj.setorganization_message_webhook_url(self.org_message_webhook_url)
                    self.Org_obj.setorganization_message_webhook_key(self.org_message_webhook_key)
                    self.Org_obj.setorganization_response_webhook_url(self.org_response_webhook_url)
                    self.Org_obj.setorganization_response_webhook_key(self.org_response_webhook_key)
                    self.Org_obj.setorganization_click_webhook_Url(self.org_click_webhook_url)
                    self.Org_obj.setorganization_click_webhook_key(self.org_click_webhook_key)
                    self.Org_obj.setorganization_telnyx_key(self.Organizations_telnyx_key)
                    self.Org_obj.clickActiveCheckbox()
                    self.Org_obj.clickQCCheckbox()
                    self.Org_obj.clickcreateorgbtn()
                    sleep(6)
                    act_title_create_organization = self.driver.title
                    exp_title_create_organization = "Organizations"
                    if act_title_create_organization == exp_title_create_organization:
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A9'] = "test-organization-add"
                        ws['B9'] = "Pass"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #s3_file_upload = self.add_s3bucket_file_upload()
                        self.driver.close()
                        assert True
                    else:
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A9'] = "test-organization-add"
                        ws['B9'] = "Fail"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #self.add_s3bucket_file_upload()
                        self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                #self.add_s3bucket_file_upload()
                self.driver.close()






    def test_Organizations_update(self):
        if self.excelReader_test_cases('test_Organizations_update', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #self.add_s3bucket_file_upload()
                fileName = "data.xlsx"
                path = "assets"
                getFolderPath = os.path.join(path, fileName)
                self.rows = XLUtils.getRowCount(getFolderPath, 'Admin')
                sleep(4)
                for r in range(24, self.rows + 1):
                    self.org_name = XLUtils.readData(getFolderPath, 'Admin', r, 1)
                    self.org_website = XLUtils.readData(getFolderPath, 'Admin', r, 2)
                    self.org_phone_number = XLUtils.readData(getFolderPath, 'Admin', r, 3)
                    self.org_first_name = XLUtils.readData(getFolderPath, 'Admin', r, 4)
                    self.org_last_name = XLUtils.readData(getFolderPath, 'Admin', r, 5)
                    self.org_email = XLUtils.readData(getFolderPath, 'Admin', r, 6)
                    self.org_contact_phone_number = XLUtils.readData(getFolderPath, 'Admin', r, 7)
                    self.org_response_service_email = XLUtils.readData(getFolderPath, 'Admin', r, 8)
                    self.org_qc_phone_number = XLUtils.readData(getFolderPath, 'Admin', r, 9)
                    self.org_sftp_host = XLUtils.readData(getFolderPath, 'Admin', r, 10)
                    self.org_sftp_username = XLUtils.readData(getFolderPath, 'Admin', r, 11)
                    self.org_sftp_password = XLUtils.readData(getFolderPath, 'Admin', r, 12)
                    self.org_sftp_file_location = XLUtils.readData(getFolderPath, 'Admin', r, 13)
                    self.org_import_public_root = XLUtils.readData(getFolderPath, 'Admin', r, 14)
                    self.org_import_private_root = XLUtils.readData(getFolderPath, 'Admin', r, 15)
                    self.org_message_webhook_url = XLUtils.readData(getFolderPath, 'Admin', r, 16)
                    self.org_message_webhook_key = XLUtils.readData(getFolderPath, 'Admin', r, 17)
                    self.org_response_webhook_url = XLUtils.readData(getFolderPath, 'Admin', r, 18)
                    self.org_response_webhook_key = XLUtils.readData(getFolderPath, 'Admin', r, 19)
                    self.org_click_webhook_url = XLUtils.readData(getFolderPath, 'Admin', r, 20)
                    self.org_click_webhook_key = XLUtils.readData(getFolderPath, 'Admin', r, 21)
                    self.Org_obj = OrganizationsClass(self.driver)
                    self.Org_obj.clickOrganizations()
                    sleep(2)
                    try:
                        if self.driver.find_element_by_class_name("custom_org_link").text == 'Automation-organization-add':
                            self.Org_obj.clickOrgEditBtn()
                        else:
                            sleep(4)
                            book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                            ws = book.active
                            ws['A10'] = "test-organization-update"
                            ws['B10'] = "no data found"
                            book.save('Reports/result_admin_cases_output.xlsx')
                            # self.add_s3bucket_file_upload()
                            self.driver.close()
                    except NoSuchElementException:
                        sleep(4)
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A10'] = "test-organization-update"
                        ws['B10'] = "no data found"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #s3_file_upload = self.add_s3bucket_file_upload()
                        self.driver.close()
                    sleep(2)
                    self.Org_obj.setorganizationName(self.org_name)
                    sleep(2)
                    self.Org_obj.setorganizationWebsite(self.org_website)
                    self.Org_obj.setorganizationPhonenumber(self.org_phone_number)
                    self.Org_obj.setorganizationFirstName(self.org_first_name)
                    self.Org_obj.setorganizationLastName(self.org_last_name)
                    self.Org_obj.setorganizationContactPhoneNo(self.org_contact_phone_number)
                    self.Org_obj.setorganizationResponseServiceEmail(self.org_response_service_email)
                    self.Org_obj.setFrom_Organization_timezone_Dropdown()
                    sleep(2)
                    self.Org_obj.setorganization_sftp_host(self.org_sftp_host)
                    sleep(1)
                    self.Org_obj.setorganization_sftp_username(self.org_sftp_username)
                    sleep(1)
                    self.Org_obj.setorganization_sftp_Password(self.org_sftp_password)
                    self.Org_obj.setorganization_sftp_FileLocation(self.org_sftp_file_location)
                    self.Org_obj.setorganization_import_root_public(self.org_import_public_root)
                    self.Org_obj.setorganization_import_root_private(self.org_import_private_root)
                    self.Org_obj.setorganization_message_webhook_url(self.org_message_webhook_url)
                    self.Org_obj.setorganization_message_webhook_key(self.org_message_webhook_key)
                    self.Org_obj.setorganization_response_webhook_url(self.org_response_webhook_url)
                    self.Org_obj.setorganization_response_webhook_key(self.org_response_webhook_key)
                    self.Org_obj.setorganization_click_webhook_Url(self.org_click_webhook_url)
                    self.Org_obj.setorganization_click_webhook_key(self.org_click_webhook_key)
                    self.Org_obj.setorganization_telnyx_key(self.Organizations_telnyx_key)
                    self.Org_obj.clickQCCheckbox()
                    self.Org_obj.clickUpdateorgbtn()
                    sleep(6)
                    act_title_update_organization = self.driver.title
                    exp_title_update_organization = "Organizations"
                    if act_title_update_organization == exp_title_update_organization:
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A10'] = "test-organization-update"
                        ws['B10'] = "Pass"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #s3_file_upload = self.add_s3bucket_file_upload()
                        self.driver.close()
                        assert True
                    else:
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A10'] = "test-organization-update"
                        ws['B10'] = "Fail"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #self.add_s3bucket_file_upload()
                        self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                #self.add_s3bucket_file_upload()
                self.driver.close()




    def test_business_unit_add(self):
        if self.excelReader_test_cases('test_business_unit_add', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #self.add_s3bucket_file_upload()
                fileName = "data.xlsx"
                path = "assets"
                getFolderPath = os.path.join(path, fileName)
                self.rows = XLUtils.getRowCount(getFolderPath, 'Admin')
                sleep(4)
                for r in range(29, self.rows + 1):
                    self.unit_name = XLUtils.readData(getFolderPath, 'Admin', r, 1)
                    self.unit_website = XLUtils.readData(getFolderPath, 'Admin', r, 2)
                    self.unit_phone_number = XLUtils.readData(getFolderPath, 'Admin', r, 3)
                    self.unit_qc_phone_number = XLUtils.readData(getFolderPath, 'Admin', r, 4)
                    self.unit_first_name = XLUtils.readData(getFolderPath, 'Admin', r, 5)
                    self.unit_last_name = XLUtils.readData(getFolderPath, 'Admin', r, 6)
                    self.unit_email = XLUtils.readData(getFolderPath, 'Admin', r, 7)
                    self.unit_contact_phone_number = XLUtils.readData(getFolderPath, 'Admin', r, 8)
                    self.unit_obj = BusinessUnitClass(self.driver)
                    self.unit_obj.clickBuisnessUnit()
                    self.unit_obj.clickcreateBuisnessUnit()
                    self.unit_obj.setUnitName(self.unit_name)
                    sleep(2)
                    self.unit_obj.setUnitWebsite(self.unit_website)
                    self.unit_obj.setUnitPhoneno(self.unit_phone_number)
                    try:
                        self.unit_obj.clickSelectOrg()
                    except NoSuchElementException:
                        sleep(4)
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A11'] = "test-add-business-units"
                        ws['B11'] = "This is not a selected organization"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        self.driver.close()
                    sleep(2)
                    self.unit_obj.click_qc_btn_class()
                    self.unit_obj.setUnitQcPhoneno(self.unit_qc_phone_number)
                    self.unit_obj.clickActiveCheckbox()
                    self.unit_obj.setUnitFirstName(self.unit_first_name)
                    self.unit_obj.setUnitLastName(self.unit_last_name)
                    self.unit_obj.setUnitEmail(self.unit_email)
                    self.unit_obj.setUnitContactPhoneNo(self.unit_contact_phone_number)
                    self.unit_obj.clickunitActiveCheckbox()
                    sleep(2)
                    self.unit_obj.clickphonelookupCheckbox()
                    sleep(2)
                    self.unit_obj.clickcreatbusinessUnitbtn()
                    sleep(5)
                    act_title_create_business_units = self.driver.title
                    exp_title_create_business_units = "Business Units"
                    if act_title_create_business_units == exp_title_create_business_units:
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A11'] = "test-add-business-units"
                        ws['B11'] = "Pass"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #self.add_s3bucket_file_upload()
                        self.driver.close()
                    else:
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A11'] = "test-add-business-units"
                        ws['B11'] = "Fail"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #self.add_s3bucket_file_upload()
                        self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                #self.add_s3bucket_file_upload()
                self.driver.close()

    def test_business_unit_update(self):
        if self.excelReader_test_cases('test_business_unit_update', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #self.add_s3bucket_file_upload()
                fileName = "data.xlsx"
                path = "assets"
                getFolderPath = os.path.join(path, fileName)
                self.rows = XLUtils.getRowCount(getFolderPath, 'Admin')
                sleep(4)
                for r in range(34, self.rows + 1):
                    self.unit_name = XLUtils.readData(getFolderPath, 'Admin', r, 1)
                    self.unit_website = XLUtils.readData(getFolderPath, 'Admin', r, 2)
                    self.unit_phone_number = XLUtils.readData(getFolderPath, 'Admin', r, 3)
                    self.unit_qc_phone_number = XLUtils.readData(getFolderPath, 'Admin', r, 4)
                    self.unit_first_name = XLUtils.readData(getFolderPath, 'Admin', r, 5)
                    self.unit_last_name = XLUtils.readData(getFolderPath, 'Admin', r, 6)
                    self.unit_email = XLUtils.readData(getFolderPath, 'Admin', r, 7)
                    self.unit_contact_phone_number = XLUtils.readData(getFolderPath, 'Admin', r, 8)
                    self.unit_obj = BusinessUnitClass(self.driver)
                    self.unit_obj.clickBuisnessUnit()
                    try:
                        if self.driver.find_element_by_class_name("td_hover").text == 'Automation-Business-unit-add':
                            self.unit_obj.clickunitEditBtn()
                        else:
                            sleep(4)
                            book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                            ws = book.active
                            ws['A12'] = "test-update-business-units"
                            ws['B12'] = "No data found"
                            book.save('Reports/result_admin_cases_output.xlsx')
                            # self.add_s3bucket_file_upload()
                            self.driver.close()
                    except NoSuchElementException:
                        sleep(4)
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A12'] = "test-update-business-units"
                        ws['B12'] = "No data found"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #self.add_s3bucket_file_upload()
                        self.driver.close()
                    self.unit_obj.setUnitName(self.unit_name)
                    sleep(2)
                    self.unit_obj.setUnitWebsite(self.unit_website)
                    self.unit_obj.setUnitPhoneno(self.unit_phone_number)
                    try:
                        self.unit_obj.clickSelectOrg()
                    except NoSuchElementException:
                        sleep(4)
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A12'] = "test-update-business-units"
                        ws['B12'] = "This is not a selected organization"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        self.driver.close()
                    sleep(2)
                    self.unit_obj.setUnitFirstName(self.unit_first_name)
                    self.unit_obj.setUnitLastName(self.unit_last_name)
                    self.unit_obj.setUnitContactPhoneNo(self.unit_contact_phone_number)
                    sleep(2)
                    self.unit_obj.clickphonelookupCheckbox()
                    sleep(2)
                    self.unit_obj.clickunitUpdateBtn()
                    sleep(6)
                    act_title_update_business_units = self.driver.title
                    exp_title_update_business_units = "Business Units"
                    if act_title_update_business_units == exp_title_update_business_units:
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A12'] = "test-update-business-units"
                        ws['B12'] = "Pass"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #self.add_s3bucket_file_upload()
                        self.driver.close()
                    else:
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A12'] = "test-update-business-units"
                        ws['B12'] = "Fail"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #self.add_s3bucket_file_upload()
                        self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                #self.add_s3bucket_file_upload()
                self.driver.close()

    def test_campaign_add(self):
        if self.excelReader_test_cases('test_campaign_add', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #s3_file_upload = self.add_s3bucket_file_upload()
                path = "assets"
                fileName = "data.xlsx"
                getFolderPath = os.path.join(path, fileName)
                self.rows = XLUtils.getRowCount(getFolderPath, 'Admin')
                for r in range(39, self.rows + 1):
                    self.campaign_name = XLUtils.readData(getFolderPath, 'Admin', r, 1)
                    self.campaign_message = XLUtils.readData(getFolderPath, 'Admin', r, 2)
                    self.campaign_obj = CampaignClass(self.driver)
                    self.campaign_obj.clickCampaigns()
                    sleep(2)
                    self.campaign_obj.clickCreateCampaigns()
                    self.campaign_obj.setCampaignName(self.campaign_name)
                    sleep(4)
                    self.campaign_obj.clickcampaigncsv()
                    sleep(2)
                    self.campaign_obj.clickstartdate()
                    sleep(2)
                    self.campaign_obj.click_Selectstart_time()
                    sleep(2)
                    self.campaign_obj.clickenddate()
                    sleep(2)
                    self.campaign_obj.click_Select_end_time()
                    sleep(2)
                    self.campaign_obj.clickCampaignContinue()
                    self.campaign_obj.setCampaignMessage(self.campaign_message)
                    try:
                        self.campaign_obj.clickSelectOrg()
                    except NoSuchElementException:
                        sleep(4)
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A13'] = "test-create-campaign"
                        ws['B13'] = "This is not a selected organization"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        self.driver.close()
                    self.campaign_obj.clickCampaignsbtn()
                    sleep(6)
                    act_title_create_campaign = self.driver.title
                    exp_title_create_campaign = "Campaigns"
                    if act_title_create_campaign == exp_title_create_campaign:
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A13'] = "test-create-campaign"
                        ws['B13'] = "Pass"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #s3_file_upload = self.add_s3bucket_file_upload()
                        self.driver.close()
                        assert True
                    else:
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A13'] = "test-create-campaign"
                        ws['B13'] = "Fail"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #s3_file_upload = self.add_s3bucket_file_upload()
                        self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                #s3_file_upload = self.add_s3bucket_file_upload()
                self.driver.close()

    def test_campaign_update(self):
        if self.excelReader_test_cases('test_campaign_update', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #s3_file_upload = self.add_s3bucket_file_upload()
                path = "assets"
                fileName = "data.xlsx"
                getFolderPath = os.path.join(path, fileName)
                self.rows = XLUtils.getRowCount(getFolderPath, 'Admin')
                for r in range(44, self.rows + 1):
                    self.campaign_name = XLUtils.readData(getFolderPath, 'Admin', r, 1)
                    self.campaign_message = XLUtils.readData(getFolderPath, 'Admin', r, 2)
                    self.campaign_obj = CampaignClass(self.driver)
                    self.campaign_obj.clickCampaigns()
                    sleep(2)
                    try:
                        if self.driver.find_element_by_class_name("campaign_sec").text == 'Automation-campaign-add':
                            self.campaign_obj.clickCampaignEditBtn()
                        else:
                            sleep(4)
                            book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                            ws = book.active
                            ws['A14'] = "test-update-campaign"
                            ws['B14'] = "No data found"
                            book.save('Reports/result_admin_cases_output.xlsx')
                            self.driver.close()
                    except NoSuchElementException:
                        sleep(4)
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A14'] = "test-update-campaign"
                        ws['B14'] = "No data found"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #s3_file_upload = self.add_s3bucket_file_upload()
                        self.driver.close()
                    self.campaign_obj.setCampaignName(self.campaign_name)
                    sleep(4)
                    self.campaign_obj.clickcampaigncsv()
                    sleep(2)
                    self.campaign_obj.clickstartdate()
                    sleep(2)
                    self.campaign_obj.clickenddate()
                    sleep(2)
                    self.campaign_obj.setCampaignMessage(self.campaign_message)
                    try:
                        self.campaign_obj.clickSelectOrg()
                    except NoSuchElementException:
                        sleep(4)
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A14'] = "test-update-campaign"
                        ws['B14'] = "This is not a selected organization"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        self.driver.close()
                    self.campaign_obj.click_campaigns_update_btn_class()
                    sleep(6)
                    act_title_update_campaign = self.driver.title
                    exp_title_update_campaign = "Campaigns"
                    if act_title_update_campaign == exp_title_update_campaign:
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A14'] = "test-update-campaign"
                        ws['B14'] = "Pass"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #s3_file_upload = self.add_s3bucket_file_upload()
                        self.driver.close()
                        assert True
                    else:
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A14'] = "test-update-campaign"
                        ws['B14'] = "Fail"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #s3_file_upload = self.add_s3bucket_file_upload()
                        self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                #s3_file_upload = self.add_s3bucket_file_upload()
                self.driver.close()



    def test_add_report_scheduler(self):
        if self.excelReader_test_cases('test_add_report_scheduler', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #s3_file_upload = self.add_s3bucket_file_upload()
                path = "assets"
                fileName = "data.xlsx"
                getFolderPath = os.path.join(path, fileName)
                self.rows = XLUtils.getRowCount(getFolderPath, 'Admin')
                for r in range(49, self.rows + 1):
                    self.title = XLUtils.readData(getFolderPath, 'Admin', r, 1)
                    self.email = XLUtils.readData(getFolderPath, 'Admin', r, 2)
                    self.sql = XLUtils.readData(getFolderPath, 'Admin', r, 3)
                    self.sftp_host = XLUtils.readData(getFolderPath, 'Admin', r, 4)
                    self.sftp_username = XLUtils.readData(getFolderPath, 'Admin', r, 5)
                    self.sftp_password = XLUtils.readData(getFolderPath, 'Admin', r, 6)
                    self.sftp_file_location = XLUtils.readData(getFolderPath, 'Admin', r, 7)
                    self.rpt_schd_obj = ReportSchedulerClass(self.driver)
                    sleep(2)
                    self.rpt_schd_obj.clickReportScheduler()
                    sleep(2)
                    self.rpt_schd_obj.clickCreateReportScheduler()
                    sleep(2)
                    try:
                        self.rpt_schd_obj.clickSelectOrg()
                    except NoSuchElementException:
                        sleep(4)
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A15'] = "test-report-scheduler-add"
                        ws['B15'] = "This is not a selected organization"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        self.driver.close()
                    self.rpt_schd_obj.setFrequencyFromDropdown()
                    sleep(2)
                    self.rpt_schd_obj.setReportTypeFromDropdown()
                    sleep(2)
                    self.rpt_schd_obj.setReportTitle(self.title)
                    self.rpt_schd_obj.setReportSchedulerSQLQuery(self.sql)
                    self.rpt_schd_obj.setReport_sftp_host(self.sftp_host)
                    self.rpt_schd_obj.setReport_sftp_username(self.sftp_username)
                    self.rpt_schd_obj.setReport_sftp_password(self.sftp_password)
                    self.rpt_schd_obj.setReport_sftp_file_location(self.sftp_file_location)
                    sleep(2)
                    self.rpt_schd_obj.clickCreatebtn()
                    sleep(10)
                    act_title_report_scheduler = self.driver.title
                    exp_title_report_scheduler = "Reports Scheduler"
                    if act_title_report_scheduler == exp_title_report_scheduler:
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A15'] = "test-report-scheduler"
                        ws['B15'] = "Pass"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #s3_file_upload = self.add_s3bucket_file_upload()
                        self.driver.close()
                        assert True
                    else:
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A15'] = "test-report-scheduler"
                        ws['B15'] = "Fail"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #s3_file_upload = self.add_s3bucket_file_upload()
                        self.driver.close()
                        # assert False
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
               # s3_file_upload = self.add_s3bucket_file_upload()
                self.driver.close()

    def test_update_report_scheduler(self):
        if self.excelReader_test_cases('test_update_report_scheduler', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #s3_file_upload = self.add_s3bucket_file_upload()
                path = "assets"
                fileName = "data.xlsx"
                getFolderPath = os.path.join(path, fileName)
                self.rows = XLUtils.getRowCount(getFolderPath, 'Admin')
                for r in range(54, self.rows + 1):
                    self.title = XLUtils.readData(getFolderPath, 'Admin', r, 1)
                    self.email = XLUtils.readData(getFolderPath, 'Admin', r, 2)
                    self.sql = XLUtils.readData(getFolderPath, 'Admin', r, 3)
                    self.sftp_host = XLUtils.readData(getFolderPath, 'Admin', r, 4)
                    self.sftp_username = XLUtils.readData(getFolderPath, 'Admin', r, 5)
                    self.sftp_password = XLUtils.readData(getFolderPath, 'Admin', r, 6)
                    self.sftp_file_location = XLUtils.readData(getFolderPath, 'Admin', r, 7)
                    self.rpt_schd_obj = ReportSchedulerClass(self.driver)
                    sleep(2)
                    self.rpt_schd_obj.clickReportScheduler()
                    sleep(2)
                    try:

                        if self.driver.find_element_by_class_name("covert_title").text == 'Automation-report-scheduler-add':
                            self.rpt_schd_obj.clickEditBtn()
                        else:
                            sleep(4)
                            book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                            ws = book.active
                            ws['A16'] = "test-update-report-scheduler"
                            ws['B16'] = "No data found"
                            book.save('Reports/result_admin_cases_output.xlsx')
                            self.driver.close()
                    except NoSuchElementException:
                        sleep(4)
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A16'] = "test-update-report-scheduler"
                        ws['B16'] = "No data found"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        # self.add_s3bucket_file_upload()
                        self.driver.close()
                    sleep(2)
                    try:
                        self.rpt_schd_obj.clickSelectOrg()
                    except NoSuchElementException:
                        sleep(4)
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A16'] = "test-update-report-scheduler"
                        ws['B16'] = "This is not a selected organization"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        self.driver.close()
                    sleep(2)
                    self.rpt_schd_obj.setFrequencyFromDropdown()
                    sleep(2)
                    self.rpt_schd_obj.setReportTypeFromDropdown()
                    sleep(2)
                    self.rpt_schd_obj.setReportTitle(self.title)
                    self.rpt_schd_obj.setReportSchedulerSQLQuery(self.sql)
                    self.rpt_schd_obj.setReport_sftp_host(self.sftp_host)
                    self.rpt_schd_obj.setReport_sftp_username(self.sftp_username)
                    self.rpt_schd_obj.setReport_sftp_password(self.sftp_password)
                    self.rpt_schd_obj.setReport_sftp_file_location(self.sftp_file_location)
                    sleep(2)
                    self.rpt_schd_obj.clickUpdateBtn()
                    sleep(10)
                    act_title_report_scheduler = self.driver.title
                    exp_title_report_scheduler = "Reports Scheduler"
                    if act_title_report_scheduler == exp_title_report_scheduler:
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A16'] = "test-update-report-scheduler"
                        ws['B16'] = "Pass"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #s3_file_upload = self.add_s3bucket_file_upload()
                        self.driver.close()
                        assert True
                    else:
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A16'] = "test-update-report-scheduler"
                        ws['B16'] = "Fail"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #s3_file_upload = self.add_s3bucket_file_upload()
                        self.driver.close()
                        # assert False
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
               # s3_file_upload = self.add_s3bucket_file_upload()
                self.driver.close()

    def test_excicute_report_scheduler(self):
      if self.excelReader_test_cases('test_excicute_report_scheduler', 'test_cases') == "success":
        act_title = self.check_login()
        exp_title = "Admin Dashboard"
        if act_title == exp_title:
            sleep(2)
            book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
            ws = book.active
            ws['A2'] = "test-login"
            ws['B2'] = "Pass"
            book.save('Reports/result_admin_cases_output.xlsx')
            self.rpt_schd_obj = ReportSchedulerClass(self.driver)
            sleep(2)
            self.rpt_schd_obj.clickReportScheduler()
            try:
                self.rpt_schd_obj.clickexcicuteBtn()
                sleep(3)
                success_toastr = self.driver.find_element_by_class_name("toast-error")
                if success_toastr:
                    book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                    ws = book.active
                    ws['A17'] = "test-excicute-report-scheduler"
                    ws['B17'] = "Fail"
                    book.save('Reports/result_admin_cases_output.xlsx')
                    self.driver.close()
            except NoSuchElementException:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A17'] = "test-excicute-report-scheduler"
                ws['B17'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                sleep(8)
                self.driver.close()
        else:
            book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
            ws = book.active
            ws['A2'] = "test-login"
            ws['B2'] = "Fail"
            book.save('Reports/result_admin_cases_output.xlsx')
            # s3_file_upload = self.add_s3bucket_file_upload()
            self.driver.close()

    def test_download_report_scheduler_History(self):
        if self.excelReader_test_cases('test_download_report_scheduler_History', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #s3_file_upload = self.add_s3bucket_file_upload()
                self.report_obj = ReportSchedulerHistoryClass(self.driver)
                sleep(2)
                self.report_obj.clickreportbtn()
                sleep(2)
                try:
                    sleep(2)
                    self.report_obj.clickDownloadReport()
                    sleep(4)
                    book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                    ws = book.active
                    ws['A18'] = "test-download-report-scheduler-history"
                    ws['B18'] = "Pass"
                    book.save('Reports/result_admin_cases_output.xlsx')
                    #s3_file_upload = self.add_s3bucket_file_upload()
                    self.driver.close()
                    assert True
                except NoSuchElementException:
                    book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                    ws = book.active
                    ws['A18'] = "test-download-report-scheduler-history"
                    ws['B18'] = "Fail"
                    book.save('Reports/result_admin_cases_output.xlsx')
                    #s3_file_upload = self.add_s3bucket_file_upload()
                    self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                #s3_file_upload = self.add_s3bucket_file_upload()
                self.driver.close()

    def test_report_scheduler_error_items(self):
        if self.excelReader_test_cases('test_report_scheduler_error_items', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                self.rpt_schd_obj = ReportSchedulerClass(self.driver)
                self.rpt_schd_obj.clickreporthistoryerrorbtn()
                sleep(4)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A19'] = "test-report-scheduler-error-items"
                ws['B19'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #s3_file_upload = self.add_s3bucket_file_upload()
                self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                #s3_file_upload = self.add_s3bucket_file_upload()
                self.driver.close()





    def test_today_campaigns_add(self):
        if self.excelReader_test_cases('test_today_campaigns_add', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #s3_file_upload = self.add_s3bucket_file_upload()
                self.todays_camp_obj = TodayCampaignsClass(self.driver)
                self.todays_camp_obj.clickTodayCampaigns()
                sleep(3)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A20'] = "test-today-campaigns"
                ws['B20'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #s3_file_upload = self.add_s3bucket_file_upload()
                self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                #s3_file_upload = self.add_s3bucket_file_upload()
                self.driver.close()



    def test_signup(self):
        if self.excelReader_test_cases('test_signup', 'test_cases') == "success":
            self.driver = webdriver.Chrome(ChromeDriverManager().install())
            self.driver.maximize_window()
            sleep(1)
            path = "assets"
            fileName = "data.xlsx"
            getFolderPath = os.path.join(path, fileName)
            self.lp = LoginClass(self.driver)
            self.rows = XLUtils.getRowCount(getFolderPath, 'Admin')
            self.signup_obj = SignupClass(self.driver)
            for r in range(59, self.rows + 1):
                self.signup_url = XLUtils.readData(getFolderPath, 'Admin', r, 1)
                self.signup_first_name = XLUtils.readData(getFolderPath, 'Admin', r, 2)
                self.signup_last_name = XLUtils.readData(getFolderPath, 'Admin', r, 3)
                self.signup_email = XLUtils.readData(getFolderPath, 'Admin', r, 4)
                self.signup_password = XLUtils.readData(getFolderPath, 'Admin', r, 5)
                self.signup_confirm_password = XLUtils.readData(getFolderPath, 'Admin', r, 6)
                self.driver.get(self.signup_url)
                self.signup_obj.setSignupFirstName(self.signup_first_name)
                self.signup_obj.setSignupLastName(self.signup_last_name)
                self.signup_obj.setSignupemail(self.signup_email)
                self.signup_obj.setSignuppassword(self.signup_password)
                self.signup_obj.setSignupconfirmpassword(self.signup_confirm_password)
                self.signup_obj.clickCreateSignupBtn()
                sleep(6)
                act_title_signup = self.driver.title
                exp_title_signup = "Signin"
                if act_title_signup == exp_title_signup:
                    book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                    ws = book.active
                    ws['A21'] = "test-signup"
                    ws['B21'] = "Pass"
                    book.save('Reports/result_admin_cases_output.xlsx')
                    #self.add_s3bucket_file_upload()
                    self.driver.close()
                    assert True
                else:
                    book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                    ws = book.active
                    ws['A21'] = "test-signup"
                    ws['B21'] = "Fail"
                    book.save('Reports/result_admin_cases_output.xlsx')
                    #self.add_s3bucket_file_upload()
                    self.driver.close()

    def test_permission_role_assign_user(self):
        if self.excelReader_test_cases('test_permission_role_assign_user', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #s3_file_upload = self.add_s3bucket_file_upload()
                sleep(2)
                self.permission_obj = PermissionClass(self.driver)
                sleep(2)
                self.permission_obj.clickpermisionbtn()
                try:
                    if self.driver.find_element_by_class_name("test_email").text == 'automation-user@gmail.com':
                        self.permission_obj.clickEditPermission()
                    else:
                        sleep(4)
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A22'] = "test-permission-role-assign-user"
                        ws['B22'] = "No Data Found"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        # self.add_s3bucket_file_upload()
                        self.driver.close()
                    self.permission_obj.clickOnboarding()
                    try:
                        self.permission_obj.clickselectOrg_permission()
                    except NoSuchElementException:
                        sleep(4)
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A22'] = "test-permission-role-assign-user"
                        ws['B22'] = "This is not a selected organization"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        self.driver.close()

                    self.permission_obj.clickRole()
                    self.permission_obj.clickActiveChk()
                    self.permission_obj.clickUpdateBtn()
                    sleep(2)
                    book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                    ws = book.active
                    ws['A22'] = "test-permission"
                    ws['B22'] = "Pass"
                    book.save('Reports/result_admin_cases_output.xlsx')
                    #s3_file_upload = self.add_s3bucket_file_upload()
                    self.driver.close()
                    assert True
                except NoSuchElementException:
                    book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                    ws = book.active
                    ws['A22'] = "test-permission"
                    ws['B22'] = "Fail"
                    book.save('Reports/result_admin_cases_output.xlsx')
                    #s3_file_upload = self.add_s3bucket_file_upload()
                    self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                #s3_file_upload = self.add_s3bucket_file_upload()
                self.driver.close()




    def test_click_records(self):
        if self.excelReader_test_cases('test_click_records', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #s3_file_upload = self.add_s3bucket_file_upload()
                self.clicObj = ClickClass(self.driver)
                sleep(2)
                self.clicObj.clickpagebtn()
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A24'] = "test-clicks-records"
                ws['B24'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #s3_file_upload = self.add_s3bucket_file_upload()
                self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                #s3_file_upload = self.add_s3bucket_file_upload()
                self.driver.close()

    def test_mapped_media_add(self):
        if self.excelReader_test_cases('test_mapped_media_add', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #s3_file_upload = self.add_s3bucket_file_upload()
                path = "assets"
                fileName = "data.xlsx"
                getFolderPath = os.path.join(path, fileName)
                self.rows = XLUtils.getRowCount(getFolderPath, 'Admin')
                for r in range(64, self.rows + 1):
                    self.file_name = XLUtils.readData(getFolderPath, 'Admin', r, 1)
                    self.mapped_location = XLUtils.readData(getFolderPath, 'Admin', r, 2)
                    self.mapped_obj = MappedmediaClass(self.driver)
                    sleep(2)
                    self.mapped_obj.clickmappedmedia()
                    sleep(2)
                    self.mapped_obj.clickcreatemappedmedia()
                    sleep(2)
                    try:
                        self.mapped_obj.clickSelectOrg()
                    except NoSuchElementException:
                        sleep(4)
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A25'] = "test-mapped-media-add"
                        ws['B25'] = "This is not a selected organization"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        self.driver.close()
                    self.mapped_obj.setfilename(self.file_name)
                    sleep(2)
                    self.mapped_obj.setmappedlocation(self.mapped_location)
                    sleep(2)
                    self.mapped_obj.click_create_mapped_media_btn_class()
                    sleep(4)
                    act_title_mapped_media = self.driver.title
                    exp_title_mapped_media = "Mapped Media"
                    if act_title_mapped_media == exp_title_mapped_media:
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A25'] = "test-mapped-media-add"
                        ws['B25'] = "Pass"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #s3_file_upload = self.add_s3bucket_file_upload()
                        self.driver.close()
                        assert True
                    else:
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A25'] = "test-mapped-media-add"
                        ws['B25'] = "Fail"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #s3_file_upload = self.add_s3bucket_file_upload()
                        self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                #s3_file_upload = self.add_s3bucket_file_upload()
                self.driver.close()

    def test_mapped_media_update(self):
        if self.excelReader_test_cases('test_mapped_media_update', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #self.add_s3bucket_file_upload()
                path = "assets"
                fileName = "data.xlsx"
                getFolderPath = os.path.join(path, fileName)
                self.rows = XLUtils.getRowCount(getFolderPath, 'Admin')
                for r in range(69, self.rows + 1):
                    self.file_name = XLUtils.readData(getFolderPath, 'Admin', r, 1)
                    self.mapped_location = XLUtils.readData(getFolderPath, 'Admin', r, 2)
                    self.mapped_obj = MappedmediaClass(self.driver)
                    sleep(2)
                    self.mapped_obj.clickmappedmedia()
                    sleep(2)
                    try:
                        if self.driver.find_element_by_class_name("file_name").text == 'Automation-mapped-media-add':
                            self.mapped_obj.click_edit_mapped_media_btn_class()
                        else:
                            sleep(4)
                            book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                            ws = book.active
                            ws['A26'] = "test-mapped-media-update"
                            ws['B26'] = "No data found"
                            book.save('Reports/result_admin_cases_output.xlsx')
                            self.driver.close()
                    except NoSuchElementException:
                        sleep(4)
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A26'] = "test-mapped-media-update"
                        ws['B26'] = "No data found"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #self.add_s3bucket_file_upload()
                        self.driver.close()
                    sleep(2)
                    try:
                        self.mapped_obj.clickSelectOrg()
                    except NoSuchElementException:
                        sleep(4)
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A26'] = "test-mapped-media-update"
                        ws['B26'] = "This is not a selected organization"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        self.driver.close()
                    self.mapped_obj.setfilename(self.file_name)
                    sleep(2)
                    self.mapped_obj.setmappedlocation(self.mapped_location)
                    sleep(2)
                    self.mapped_obj.click_update_mapped_media_btn_class()
                    sleep(4)
                    act_title_update_mapped_media = self.driver.title
                    exp_title_update_mapped_media = "Mapped Media"
                    if act_title_update_mapped_media == exp_title_update_mapped_media:
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A26'] = "test-mapped-media-update"
                        ws['B26'] = "Pass"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #self.add_s3bucket_file_upload()
                        self.driver.close()
                        assert True
                    else:
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A26'] = "test-mapped-media-update"
                        ws['B26'] = "Fail"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #self.add_s3bucket_file_upload()
                        self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                #self.add_s3bucket_file_upload()
                self.driver.close()

    def test_create_phone_lookup_rule(self):
        if self.excelReader_test_cases('test_create_phone_lookup_rule', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #s3_file_upload = self.add_s3bucket_file_upload()
                path = "assets"
                fileName = "data.xlsx"
                getFolderPath = os.path.join(path, fileName)
                self.rows = XLUtils.getRowCount(getFolderPath, 'Admin')
                for r in range(74, self.rows + 1):
                    self.title = XLUtils.readData(getFolderPath, 'Admin', r, 1)
                    self.message = XLUtils.readData(getFolderPath, 'Admin', r, 2)
                    self.error_trigger = XLUtils.readData(getFolderPath, 'Admin', r, 3)
                    self.phnlookup_obj = PhoneLookupRuleClass(self.driver)
                    self.phnlookup_obj.clickPhoneLookup()
                    self.phnlookup_obj.clickCreatePhoneLookup()
                    self.phnlookup_obj.setTitleName(self.title)
                    self.phnlookup_obj.setPhoneLookUpMessage(self.message)
                    self.phnlookup_obj.setPhoneLookUpErrorTriger(self.error_trigger)
                    self.phnlookup_obj.clickuploadimage()
                    try:
                        self.phnlookup_obj.clickSelectOrg()
                    except NoSuchElementException:
                        sleep(4)
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A27'] = "test-create-phone-lookup"
                        ws['B27'] = "This is not a selected organization"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        self.driver.close()

                    self.phnlookup_obj.clickActiveCheckbox()
                    self.phnlookup_obj.clickMessageCheckbox()
                    self.phnlookup_obj.clickMediaActiveCheckbox()
                    self.phnlookup_obj.clickErrorActiveCheckbox()
                    self.phnlookup_obj.clickCreatePhoneLookupbtn()
                    sleep(8)
                    act_title_create_phone_lookup = self.driver.title
                    exp_title_create_phone_lookup = "Phone Lookup"
                    if act_title_create_phone_lookup == exp_title_create_phone_lookup:
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A27'] = "test-create-phone-lookup"
                        ws['B27'] = "Pass"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #s3_file_upload = self.add_s3bucket_file_upload()
                        self.driver.close()
                        assert True
                    else:
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A27'] = "test-create-phone-lookup"
                        ws['B27'] = "Fail"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #s3_file_upload = self.add_s3bucket_file_upload()
                        self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                #self.add_s3bucket_file_upload()
                self.driver.close()

    def test_update_phone_lookup_rule(self):
        if self.excelReader_test_cases('test_update_phone_lookup_rule', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #s3_file_upload = self.add_s3bucket_file_upload()
                path = "assets"
                fileName = "data.xlsx"
                getFolderPath = os.path.join(path, fileName)
                self.rows = XLUtils.getRowCount(getFolderPath, 'Admin')
                for r in range(79, self.rows + 1):
                    self.title = XLUtils.readData(getFolderPath, 'Admin', r, 1)
                    self.message = XLUtils.readData(getFolderPath, 'Admin', r, 2)
                    self.error_trigger = XLUtils.readData(getFolderPath, 'Admin', r, 3)
                    self.phnlookup_obj = PhoneLookupRuleClass(self.driver)
                    self.phnlookup_obj.clickPhoneLookup()

                    try:
                        if self.driver.find_element_by_class_name("title").text == 'Automation-phone-lookup-add':
                            self.phnlookup_obj.clickPhoneLookupEditBtn()
                        else:
                            sleep(4)
                            book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                            ws = book.active
                            ws['A28'] = "test-update-phone-lookup"
                            ws['B28'] = "No data found"
                            book.save('Reports/result_admin_cases_output.xlsx')
                            self.driver.close()
                    except NoSuchElementException:
                        sleep(4)
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A28'] = "test-update-phone-lookup"
                        ws['B28'] = "No data found"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #s3_file_upload = self.add_s3bucket_file_upload()
                        self.driver.close()
                    self.phnlookup_obj.setTitleName(self.title)
                    self.phnlookup_obj.setPhoneLookUpMessage(self.message)
                    self.phnlookup_obj.setPhoneLookUpErrorTriger(self.error_trigger)
                    self.phnlookup_obj.clickuploadimage()
                    sleep(3)
                    try:
                        self.phnlookup_obj.clickSelectOrg()
                    except NoSuchElementException:
                        sleep(4)
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A28'] = "test-update-phone-lookup"
                        ws['B28'] = "This is not a selected organization"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        self.driver.close()
                    self.phnlookup_obj.clickActiveCheckbox()
                    self.phnlookup_obj.clickMessageCheckbox()
                    self.phnlookup_obj.clickMediaActiveCheckbox()
                    self.phnlookup_obj.clickErrorActiveCheckbox()
                    self.phnlookup_obj.clickUpdatePhoneLookupbtn()
                    sleep(8)
                    act_title_update_phone_lookup = self.driver.title
                    exp_title_update_phone_lookup = "Phone Lookup"
                    if act_title_update_phone_lookup == exp_title_update_phone_lookup:
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A28'] = "test-update-phone-lookup"
                        ws['B28'] = "Pass"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #s3_file_upload = self.add_s3bucket_file_upload()
                        self.driver.close()
                        assert True
                    else:
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A28'] = "test-update-phone-lookup"
                        ws['B28'] = "Fail"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #s3_file_upload = self.add_s3bucket_file_upload()
                        self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                # self.add_s3bucket_file_upload()
                self.driver.close()

    def test_common_reponse_add(self):
        if self.excelReader_test_cases('test_common_reponse_add', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #s3_file_upload = self.add_s3bucket_file_upload()
                path = "assets"
                fileName = "data.xlsx"
                getFolderPath = os.path.join(path, fileName)
                self.rows = XLUtils.getRowCount(getFolderPath, 'Admin')
                for r in range(84, self.rows + 1):
                    self.response = XLUtils.readData(getFolderPath, 'Admin', r, 1)
                    self.synonyms = XLUtils.readData(getFolderPath, 'Admin', r, 2)
                    self.message = XLUtils.readData(getFolderPath, 'Admin', r, 3)
                    self.response_obj = CommonResponseClass(self.driver)
                    self.response_obj.clickCommonResponse()
                    self.response_obj.clickCreateCommonResponse()
                    sleep(2)
                    self.response_obj.setResponse(self.response)
                    sleep(2)
                    self.response_obj.setCommonResponsesynonyms(self.synonyms)
                    sleep(2)
                    self.response_obj.setResponseMessage(self.message)
                    sleep(2)
                    self.response_obj.clickuploadimage()
                    sleep(4)
                    try:
                        self.response_obj.clickSelectOrg()
                    except NoSuchElementException:
                        sleep(4)
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A29'] = "test-add-common-responses"
                        ws['B29'] = "This is not a selected organization"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        self.driver.close()
                    sleep(4)
                    self.response_obj.clickCreatecommonResponseBtn()
                    sleep(5)
                    act_title_create_common_response = self.driver.title
                    exp_title_create_common_response = "Common Responses"
                    if act_title_create_common_response == exp_title_create_common_response:
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A29'] = "test-add-common-responses"
                        ws['B29'] = "Pass"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #s3_file_upload = self.add_s3bucket_file_upload()
                        self.driver.close()
                        assert True
                    else:
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A29'] = "test-add-common-responses"
                        ws['B29'] = "Fail"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #s3_file_upload = self.add_s3bucket_file_upload()
                        self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                # self.add_s3bucket_file_upload()
                self.driver.close()

    def test_common_reponse_update(self):
        if self.excelReader_test_cases('test_common_reponse_update', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
               # s3_file_upload = self.add_s3bucket_file_upload()
                path = "assets"
                fileName = "data.xlsx"
                getFolderPath = os.path.join(path, fileName)
                self.rows = XLUtils.getRowCount(getFolderPath, 'Admin')
                for r in range(89, self.rows + 1):
                    self.response = XLUtils.readData(getFolderPath, 'Admin', r, 1)
                    self.synonyms = XLUtils.readData(getFolderPath, 'Admin', r, 2)
                    self.message = XLUtils.readData(getFolderPath, 'Admin', r, 3)
                    self.response_obj = CommonResponseClass(self.driver)
                    self.response_obj.clickCommonResponse()
                    try:
                        if self.driver.find_element_by_class_name("td_hover").text == 'Automation-common-responses-add':
                            self.response_obj.clickEditcommonResponseBtn()
                        else:
                            sleep(4)
                            book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                            ws = book.active
                            ws['A30'] = "test-update-common-responses"
                            ws['B30'] = "no data found"
                            book.save('Reports/result_admin_cases_output.xlsx')
                            self.driver.close()
                    except NoSuchElementException:
                        sleep(4)
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A30'] = "test-update-common-responses"
                        ws['B30'] = "no data found"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        self.driver.close()
                    sleep(2)
                    self.response_obj.setResponse(self.response)
                    sleep(2)
                    self.response_obj.setResponseMessage(self.message)
                    sleep(2)
                    self.response_obj.clickuploadimage()
                    sleep(4)
                    try:
                        self.response_obj.clickSelectOrg()
                    except NoSuchElementException:
                        sleep(4)
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A30'] = "test-update-common-responses"
                        ws['B30'] = "This is not a selected organization"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        self.driver.close()
                    sleep(4)
                    self.response_obj.clickUpdatecommonResponseBtn()
                    sleep(5)
                    act_title_update_common_response = self.driver.title
                    exp_title_update_common_response = "Common Responses"
                    if act_title_update_common_response == exp_title_update_common_response:
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A30'] = "test-update-common-responses"
                        ws['B30'] = "Pass"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #s3_file_upload = self.add_s3bucket_file_upload()
                        self.driver.close()
                        assert True
                    else:
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A30'] = "test-update-common-responses"
                        ws['B30'] = "Fail"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #s3_file_upload = self.add_s3bucket_file_upload()
                        self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                # self.add_s3bucket_file_upload()
                self.driver.close()

    def test_message_merge_add(self):
        if self.excelReader_test_cases('test_message_merge_add', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #s3_file_upload = self.add_s3bucket_file_upload()
                path = "assets"
                fileName = "data.xlsx"
                getFolderPath = os.path.join(path, fileName)
                self.rows = XLUtils.getRowCount(getFolderPath, 'Admin')
                for r in range(94, self.rows + 1):
                    self.merge_key = XLUtils.readData(getFolderPath, 'Admin', r, 1)
                    self.short_description = XLUtils.readData(getFolderPath, 'Admin', r, 2)
                    self.full_description = XLUtils.readData(getFolderPath, 'Admin', r, 3)
                    self.msgmerge_obj = MessagemergeClass(self.driver)
                    self.msgmerge_obj.clickMessageMerge()
                    sleep(2)
                    self.msgmerge_obj.clickCreateMessageMerge()
                    sleep(2)
                    self.msgmerge_obj.setMessageMergeKey(self.merge_key)
                    self.msgmerge_obj.setMessageShortDes(self.short_description)
                    self.msgmerge_obj.setMessageFulltDes(self.full_description)
                    self.msgmerge_obj.clickSelectMessageOrg()

                    try:
                        self.msgmerge_obj.clickSelectMessageOrg()
                    except NoSuchElementException:
                        sleep(4)
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A31'] = "test-message-merge-add"
                        ws['B31'] = "This is not a selected organization"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        self.driver.close()

                    self.msgmerge_obj.clickMessageMergeCheckbox()
                    self.msgmerge_obj.clickMessagemergesbtn()
                    sleep(6)
                    act_title_create_message_merge = self.driver.title
                    exp_title_create_message_merge = "Message Merges"
                    if act_title_create_message_merge == exp_title_create_message_merge:
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A31'] = "test-message-merge-add"
                        ws['B31'] = "Pass"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #s3_file_upload = self.add_s3bucket_file_upload()
                        self.driver.close()
                        assert True
                    else:
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A31'] = "test-message-merge-add"
                        ws['B31'] = "Fail"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #s3_file_upload = self.add_s3bucket_file_upload()
                        self.driver.close()

            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                # self.add_s3bucket_file_upload()
                self.driver.close()

    def test_message_merge_update(self):
        if self.excelReader_test_cases('test_message_merge_update', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #s3_file_upload = self.add_s3bucket_file_upload()
                path = "assets"
                fileName = "data.xlsx"
                getFolderPath = os.path.join(path, fileName)
                self.rows = XLUtils.getRowCount(getFolderPath, 'Admin')
                for r in range(99, self.rows + 1):
                    self.merge_key = XLUtils.readData(getFolderPath, 'Admin', r, 1)
                    self.short_description = XLUtils.readData(getFolderPath, 'Admin', r, 2)
                    self.full_description = XLUtils.readData(getFolderPath, 'Admin', r, 3)
                    self.msgmerge_obj = MessagemergeClass(self.driver)
                    self.msgmerge_obj.clickMessageMerge()
                    sleep(2)
                    try:
                        if self.driver.find_element_by_class_name("td_hover").text == 'Automation-message-merge-add':
                            self.msgmerge_obj.clickMessageMergeEditBtn()
                        else:
                            sleep(4)
                            book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                            ws = book.active
                            ws['A32'] = "test-message-merge-update"
                            ws['B32'] = "no data found"
                            book.save('Reports/result_admin_cases_output.xlsx')
                            # self.add_s3bucket_file_upload()
                            self.driver.close()
                    except NoSuchElementException:
                        sleep(4)
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A32'] = "test-message-merge-update"
                        ws['B32'] = "no data found"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #s3_file_upload = self.add_s3bucket_file_upload()
                        self.driver.close()
                    sleep(2)
                    self.msgmerge_obj.setMessageMergeKey(self.merge_key)
                    self.msgmerge_obj.setMessageShortDes(self.short_description)
                    self.msgmerge_obj.setMessageFulltDes(self.full_description)
                    try:
                        self.msgmerge_obj.clickSelectMessageOrg()
                    except NoSuchElementException:
                        sleep(4)
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A32'] = "test-message-merge-update"
                        ws['B32'] = "This is not a selected organization"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        self.driver.close()
                    self.msgmerge_obj.clickMessageMergeCheckbox()
                    self.msgmerge_obj.click_msg_merge_update_btn_class()
                    sleep(6)
                    act_title_update_message_merge = self.driver.title
                    exp_title_create_message_merge = "Message Merges"
                    if act_title_update_message_merge == exp_title_create_message_merge:
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A32'] = "test-message-merge-update"
                        ws['B32'] = "Pass"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #s3_file_upload = self.add_s3bucket_file_upload()
                        self.driver.close()
                        assert True
                    else:
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A32'] = "test-message-merge-update"
                        ws['B32'] = "Fail"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #s3_file_upload = self.add_s3bucket_file_upload()
                        self.driver.close()

            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                # self.add_s3bucket_file_upload()
                self.driver.close()


    def test_message_merge_parser_add(self):
        if self.excelReader_test_cases('test_message_merge_parser_add', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #self.add_s3bucket_file_upload()
                fileName = "data.xlsx"
                path = "assets"
                getFolderPath = os.path.join(path, fileName)
                self.rows = XLUtils.getRowCount(getFolderPath, 'Admin')
                sleep(4)
                for r in range(104, self.rows + 1):
                    self.string = XLUtils.readData(getFolderPath, 'Admin', r, 1)
                    self.msgmergeparser_obj = MessagemergeparserClass(self.driver)
                    sleep(2)
                    self.msgmergeparser_obj.clickMessageMergeParser()
                    sleep(2)
                    self.msgmergeparser_obj.clickSelectMessageMergeParser()
                    sleep(2)
                    self.msgmergeparser_obj.setMessageMergeParserName(self.string)
                    self.msgmergeparser_obj.clickMessagemergesparserbtn()
                    sleep(4)
                    success_toastr = self.driver.find_element_by_class_name("toast-message")
                    success_toastr_text = success_toastr.text
                    if 'Message merge successfully' == success_toastr_text:
                        sleep(8)
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A33'] = "test-message-merge-parser-add"
                        ws['B33'] = "Pass"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #s3_file_upload = self.add_s3bucket_file_upload()
                        self.driver.close()
                        assert True
                    else:
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A33'] = "test-message-merge-parser-add"
                        ws['B33'] = "Fail"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #self.add_s3bucket_file_upload()
                        self.driver.close()
                        #assert False
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                # self.add_s3bucket_file_upload()
                self.driver.close()

    def test_shorten_link_add(self):
        if self.excelReader_test_cases('test_shorten_link_add', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #s3_file_upload = self.add_s3bucket_file_upload()
                fileName = "data.xlsx"
                path = "assets"
                getFolderPath = os.path.join(path, fileName)
                self.rows = XLUtils.getRowCount(getFolderPath, 'Admin')
                sleep(4)
                for r in range(109, self.rows + 1):
                    self.shorten_url = XLUtils.readData(getFolderPath, 'Admin', r, 1)
                    self.shorten_obj = Shorten_link_Class(self.driver)
                    sleep(2)
                    self.shorten_obj.clickCreateShortenUrl()
                    sleep(2)
                    self.shorten_obj.clickSelectOrg()
                    self.shorten_obj.setShortenurlName(self.shorten_url)
                    self.shorten_obj.clicksetExpireTime()
                    self.shorten_obj.clickShortenUrl()
                    sleep(4)
                    try:
                        success_toastr = self.driver.find_element_by_class_name("toast-success")
                        if success_toastr:
                            book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                            ws = book.active
                            ws['A34'] = "test-shorten-link"
                            ws['B34'] = "Pass"
                            book.save('Reports/result_admin_cases_output.xlsx')
                            #s3_file_upload = self.add_s3bucket_file_upload()
                            self.driver.close()
                            assert True
                    except NoSuchElementException:
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A34'] = "test-shorten-link"
                        ws['B34'] = "Fail"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #self.add_s3bucket_file_upload()
                        self.driver.close()

            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                # self.add_s3bucket_file_upload()
                self.driver.close()



    def test_ignore_add(self):
        if self.excelReader_test_cases('test_ignore_add', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #self.add_s3bucket_file_upload()
                fileName = "data.xlsx"
                path = "assets"
                getFolderPath = os.path.join(path, fileName)
                self.rows = XLUtils.getRowCount(getFolderPath, 'Admin')
                sleep(4)
                for r in range(114, self.rows + 1):
                    self.domain_name = XLUtils.readData(getFolderPath, 'Admin', r, 1)
                    self.Ignore_obj = IgnoreLinkClass(self.driver)
                    self.Ignore_obj.clickIgnoreLink()
                    self.Ignore_obj.clickCreateIgnoreLink()
                    sleep(2)
                    self.Ignore_obj.setDomainName(self.domain_name)
                    sleep(2)
                    try:
                        self.Ignore_obj.clickSelectOrg()
                    except NoSuchElementException:
                        sleep(4)
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A35'] = "test-ignore-add"
                        ws['B35'] = "This is not a selected organization"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        self.driver.close()
                    sleep(2)
                    self.Ignore_obj.clickActiveCheckbox()
                    self.Ignore_obj.IgnoreclickActiveCheckbox()
                    self.Ignore_obj.clickIgnorebtn()
                    sleep(4)
                    act_title_create_ignore_link = self.driver.title
                    exp_title_create_ignore_link = "Manage Ignore Link"
                    if act_title_create_ignore_link == exp_title_create_ignore_link:
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A35'] = "test-ignore-add"
                        ws['B35'] = "Pass"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #s3_file_upload = self.add_s3bucket_file_upload()
                        self.driver.close()
                        assert True
                    else:
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A35'] = "test-ignore-add"
                        ws['B35'] = "Fail"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #self.add_s3bucket_file_upload()
                        self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                # self.add_s3bucket_file_upload()
                self.driver.close()

    def test_ignore_update(self):
        if self.excelReader_test_cases('test_ignore_update', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #self.add_s3bucket_file_upload()
                fileName = "data.xlsx"
                path = "assets"
                getFolderPath = os.path.join(path, fileName)
                self.rows = XLUtils.getRowCount(getFolderPath, 'Admin')
                sleep(4)
                for r in range(119, self.rows + 1):
                    self.domain_name = XLUtils.readData(getFolderPath, 'Admin', r, 1)
                    self.Ignore_obj = IgnoreLinkClass(self.driver)
                    self.Ignore_obj.clickIgnoreLink()
                    try:
                        if self.driver.find_element_by_class_name("td_hover").text == 'Add-automation.com':
                            self.Ignore_obj.clickIgnorelinkEditBtn()
                        else:
                            sleep(4)
                            book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                            ws = book.active
                            ws['A36'] = "test-ignore-update"
                            ws['B36'] = "no data found"
                            book.save('Reports/result_admin_cases_output.xlsx')
                            # self.add_s3bucket_file_upload()
                            self.driver.close()
                    except NoSuchElementException:
                        sleep(4)
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A36'] = "test-ignore-update"
                        ws['B36'] = "no data found"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #s3_file_upload = self.add_s3bucket_file_upload()
                        self.driver.close()
                    sleep(2)
                    self.Ignore_obj.setDomainName(self.domain_name)
                    sleep(2)
                    try:
                        self.Ignore_obj.clickSelectOrg()
                    except NoSuchElementException:
                        sleep(4)
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A36'] = "test-ignore-update"
                        ws['B36'] = "This is not a selected organization"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        self.driver.close()
                    sleep(2)
                    self.Ignore_obj.clickActiveCheckbox()
                    self.Ignore_obj.IgnoreclickActiveCheckbox()
                    self.Ignore_obj.click_update_Ignore_link_btn_class()
                    sleep(4)
                    act_title_update_ignore_link = self.driver.title
                    exp_title_update_ignore_link = "Manage Ignore Link"
                    if act_title_update_ignore_link == exp_title_update_ignore_link:
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A36'] = "test-ignore-update"
                        ws['B36'] = "Pass"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #s3_file_upload = self.add_s3bucket_file_upload()
                        self.driver.close()
                        assert True
                    else:
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A36'] = "test-ignore-update"
                        ws['B36'] = "Fail"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #self.add_s3bucket_file_upload()
                        self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                # self.add_s3bucket_file_upload()
                self.driver.close()

    def test_conversation_history_add(self):
        if self.excelReader_test_cases('test_conversation_history_add', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                self.history_obj = ConversationhistoryClass(self.driver)
                sleep(2)
                self.history_obj.clickconversationhistorybtn()
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A37'] = "test-conversation-history-add"
                ws['B37'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                self.driver.close()

    def test_new_response(self):
        if self.excelReader_test_cases('test_new_response', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                self.newResp_obj = NewResponseClass(self.driver)
                sleep(2)
                self.newResp_obj.clicknewresponseLink()
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A38'] = "test-new-responser"
                ws['B38'] = "pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                self.driver.close()

    def test_past_response(self):
        if self.excelReader_test_cases('test_past_response', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                self.pastResp_obj = PastResponseClass(self.driver)
                self.pastResp_obj.clickpastresponseLink()
                sleep(4)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A39'] = "test-past-responser"
                ws['B39'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #s3_file_upload = self.add_s3bucket_file_upload()
                self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                # self.add_s3bucket_file_upload()
                self.driver.close()

    def test_Ignore_response(self):
        if self.excelReader_test_cases('test_Ignore_response', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                self.Ignore_Resp_obj = IgnoreResponseClass(self.driver)
                self.Ignore_Resp_obj.clickignoreresponseLink()
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A40'] = "test-ignore-responser"
                ws['B40'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #self.add_s3bucket_file_upload()
                self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                # self.add_s3bucket_file_upload()
                self.driver.close()

    def test_api_key_add(self):
        if self.excelReader_test_cases('test_api_key_add', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #self.add_s3bucket_file_upload()
                fileName = "data.xlsx"
                path = "assets"
                getFolderPath = os.path.join(path, fileName)
                self.rows = XLUtils.getRowCount(getFolderPath, 'Admin')
                sleep(4)
                for r in range(124, self.rows + 1):
                    self.key = XLUtils.readData(getFolderPath, 'Admin', r, 1)
                    self.apikey_obj = ApiKeyClass(self.driver)
                    self.apikey_obj.clickorgbtn()
                    try:
                        self.apikey_obj.clickorgapibtn()
                    except NoSuchElementException:
                        sleep(4)
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A41'] = "test-insert-api-key"
                        ws['B41'] = "Organization not found"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        self.driver.close()
                    self.apikey_obj.clickcreateapikeybtn()
                    self.apikey_obj.setapikey(self.key)
                    sleep(2)
                    self.apikey_obj.clickdatepickertest()
                    sleep(2)
                    self.apikey_obj.clickActiveCheckbox()
                    sleep(2)
                    self.apikey_obj.clickapikeycreatebtn()
                    sleep(6)
                    act_title_create_api_key = self.driver.title
                    exp_title_create_api_key = "Api Key"
                    if act_title_create_api_key == exp_title_create_api_key:
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A41'] = "test-insert-api-key"
                        ws['B41'] = "Pass"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #s3_file_upload = self.add_s3bucket_file_upload()
                        self.driver.close()
                        assert True
                    else:
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A41'] = "test-insert-api-key"
                        ws['B41'] = "Fail"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #self.add_s3bucket_file_upload()
                        self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                # self.add_s3bucket_file_upload()
                self.driver.close()

    def test_api_key_update(self):
        if self.excelReader_test_cases('test_api_key_update', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #self.add_s3bucket_file_upload()
                fileName = "data.xlsx"
                path = "assets"
                getFolderPath = os.path.join(path, fileName)
                self.rows = XLUtils.getRowCount(getFolderPath, 'Admin')
                sleep(4)
                for r in range(129, self.rows + 1):
                    self.key = XLUtils.readData(getFolderPath, 'Admin', r, 1)
                    self.apikey_obj = ApiKeyClass(self.driver)
                    self.apikey_obj.clickorgbtn()
                    try:
                        self.apikey_obj.clickorgapibtn()
                    except NoSuchElementException:
                        sleep(4)
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A42'] = "test-update-api-key"
                        ws['B42'] = "Organization not found"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        self.driver.close()
                    try:
                        if self.driver.find_element_by_xpath("//*[@id='tableData_api']/tbody/tr[1]/td[1]").text == 'Automation-api-key-add':
                            self.apikey_obj.clickapikeyeditbtn()
                        else:
                            book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                            ws = book.active
                            ws['A42'] = "test-update-api-key"
                            ws['B42'] = "No data found"
                            book.save('Reports/result_admin_cases_output.xlsx')
                            # self.add_s3bucket_file_upload()
                            self.driver.close()
                    except NoSuchElementException:
                        sleep(4)
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A42'] = "test-update-api-key"
                        ws['B42'] = "no data found"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #s3_file_upload = self.add_s3bucket_file_upload()
                        self.driver.close()
                    self.apikey_obj.setapikey(self.key)
                    sleep(2)
                    self.apikey_obj.clickdatepickertest()
                    sleep(2)
                    self.apikey_obj.clickActiveCheckbox()
                    sleep(2)
                    self.apikey_obj.clickapikeyupdatebtn()
                    sleep(6)
                    act_title_update_api_key = self.driver.title
                    exp_title_update_api_key = "Api Key"
                    if act_title_update_api_key == exp_title_update_api_key:
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A42'] = "test-update-api-key"
                        ws['B42'] = "Pass"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #s3_file_upload = self.add_s3bucket_file_upload()
                        self.driver.close()
                        assert True
                    else:
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A42'] = "test-update-api-key"
                        ws['B42'] = "Fail"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #self.add_s3bucket_file_upload()
                        self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                # self.add_s3bucket_file_upload()
                self.driver.close()



    def test_vcard_add(self):
        if self.excelReader_test_cases('test_vcard_add', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #self.add_s3bucket_file_upload()
                fileName = "data.xlsx"
                path = "assets"
                getFolderPath = os.path.join(path, fileName)
                self.rows = XLUtils.getRowCount(getFolderPath, 'Admin')
                sleep(4)
                for r in range(134, self.rows + 1):
                    self.vcard_name = XLUtils.readData(getFolderPath, 'Admin', r, 1)
                    self.display_vcard_name = XLUtils.readData(getFolderPath, 'Admin', r, 2)
                    self.vcard_website = XLUtils.readData(getFolderPath, 'Admin', r, 3)
                    self.vcard_phone_number = XLUtils.readData(getFolderPath, 'Admin', r, 4)
                    self.vcard_first_name = XLUtils.readData(getFolderPath, 'Admin', r, 5)
                    self.vcard_last_name = XLUtils.readData(getFolderPath, 'Admin', r, 6)
                    self.vcard_email = XLUtils.readData(getFolderPath, 'Admin', r, 7)
                    self.vcard_address = XLUtils.readData(getFolderPath, 'Admin', r, 8)
                    self.vcard_obj = VcardClass(self.driver)
                    self.vcard_obj.clickVcard()
                    self.vcard_obj.clickcreateVcard()
                    self.vcard_obj.setVcardName(self.vcard_name)
                    self.vcard_obj.displayVcardName(self.display_vcard_name)
                    self.vcard_obj.displayVcardWebsite(self.vcard_website)
                    self.vcard_obj.displayVcardTelephone(self.vcard_phone_number)
                    self.vcard_obj.displayVcard_f_name(self.vcard_first_name)
                    self.vcard_obj.displayVcard_l_name(self.vcard_last_name)
                    self.vcard_obj.displayVcard_email(self.vcard_email)
                    self.vcard_obj.displayVcard_Address(self.vcard_address)
                    sleep(2)
                    try:
                        self.vcard_obj.clickSelectOrg()
                    except NoSuchElementException:
                        sleep(4)
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A43'] = "test-vcard-add"
                        ws['B43'] = "This is not a selected organization"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        self.driver.close()
                    sleep(2)
                    self.vcard_obj.clickcreatevcardbtn()
                    sleep(6)
                    act_title_create_vcard = self.driver.title
                    exp_title_create_vcard = "Vcard"
                    if act_title_create_vcard == exp_title_create_vcard:
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A43'] = "test-vcard-add"
                        ws['B43'] = "Pass"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #self.add_s3bucket_file_upload()
                        self.driver.close()
                        assert True
                    else:
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A43'] = "test-vcard-add"
                        ws['B43'] = "Fail"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #self.add_s3bucket_file_upload()
                        self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                self.driver.close()

    def test_vcard_update(self):
        if self.excelReader_test_cases('test_vcard_update', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                # self.add_s3bucket_file_upload()
                fileName = "data.xlsx"
                path = "assets"
                getFolderPath = os.path.join(path, fileName)
                self.rows = XLUtils.getRowCount(getFolderPath, 'Admin')
                sleep(4)
                for r in range(139, self.rows + 1):
                    self.vcard_name = XLUtils.readData(getFolderPath, 'Admin', r, 1)
                    self.display_vcard_name = XLUtils.readData(getFolderPath, 'Admin', r, 2)
                    self.vcard_website = XLUtils.readData(getFolderPath, 'Admin', r, 3)
                    self.vcard_phone_number = XLUtils.readData(getFolderPath, 'Admin', r, 4)
                    self.vcard_first_name = XLUtils.readData(getFolderPath, 'Admin', r, 5)
                    self.vcard_last_name = XLUtils.readData(getFolderPath, 'Admin', r, 6)
                    self.vcard_email = XLUtils.readData(getFolderPath, 'Admin', r, 7)
                    self.vcard_address = XLUtils.readData(getFolderPath, 'Admin', r, 8)
                    self.vcard_obj = VcardClass(self.driver)
                    self.vcard_obj.clickVcard()
                    try:
                        if self.driver.find_element_by_class_name("td_hover").text == 'Automation-vcard-add':
                            self.vcard_obj.clickVcardEditBtn()
                        else:
                            sleep(2)
                            book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                            ws = book.active
                            ws['A44'] = "test-vcard-update"
                            ws['B44'] = "No data found"
                            book.save('Reports/result_admin_cases_output.xlsx')
                            # self.add_s3bucket_file_upload()
                            self.driver.close()
                    except NoSuchElementException:
                        sleep(4)
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A44'] = "test-vcard-update"
                        ws['B44'] = "No data found"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        self.driver.close()
                    self.vcard_obj.setVcardName(self.vcard_name)
                    self.vcard_obj.displayVcardName(self.display_vcard_name)
                    self.vcard_obj.displayVcardWebsite(self.vcard_website)
                    self.vcard_obj.displayVcardTelephone(self.vcard_phone_number)
                    self.vcard_obj.displayVcard_f_name(self.vcard_first_name)
                    self.vcard_obj.displayVcard_l_name(self.vcard_last_name)
                    self.vcard_obj.displayVcard_Address(self.vcard_address)
                    sleep(4)
                    try:
                        self.vcard_obj.clickSelectOrg()
                    except NoSuchElementException:
                        sleep(4)
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A43'] = "test-vcard-add"
                        ws['B43'] = "This is not a selected organization"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        self.driver.close()
                    self.vcard_obj.clickVcardUpdateBtn()
                    sleep(6)
                    act_title_update_vcard = self.driver.title
                    exp_title_update_vcard = "Vcard"
                    if act_title_update_vcard == exp_title_update_vcard:
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A44'] = "test-vcard-update"
                        ws['B44'] = "Pass"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        # self.add_s3bucket_file_upload()
                        self.driver.close()
                        assert True
                    else:
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A44'] = "test-vcard-update"
                        ws['B44'] = "Fail"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        # self.add_s3bucket_file_upload()
                        self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                # self.add_s3bucket_file_upload()
                self.driver.close()

    def test_longcodes_add(self):
        if self.excelReader_test_cases('test_longcodes_add', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #self.add_s3bucket_file_upload()
                fileName = "data.xlsx"
                path = "assets"
                getFolderPath = os.path.join(path, fileName)
                self.rows = XLUtils.getRowCount(getFolderPath, 'Admin')
                sleep(4)
                for r in range(144, self.rows + 1):
                    self.longcode_phone_number = XLUtils.readData(getFolderPath, 'Admin', r, 1)
                    self.longcode_notes = XLUtils.readData(getFolderPath, 'Admin', r, 2)
                    self.longcode_provider = XLUtils.readData(getFolderPath, 'Admin', r, 3)
                    self.longcode_obj = LongcodesClass(self.driver)
                    self.longcode_obj.clickLongcode()
                    try:
                        self.longcode_obj.clickOrgLongcodeBtn()
                    except NoSuchElementException:
                        sleep(4)
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A45'] = "test-longcode-add"
                        ws['B45'] = "No organizations found"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        self.driver.close()
                    self.longcode_obj.clickCreateLongcode()
                    self.longcode_obj.setlongcodeNumber(self.longcode_phone_number)
                    self.longcode_obj.setlongcodeNotes(self.longcode_notes)
                    self.longcode_obj.setlongcodeProvider(self.longcode_provider)
                    sleep(2)
                    self.longcode_obj.clickActiveCheckbox()
                    sleep(2)
                    self.longcode_obj.clickcreatelongcodebtn()
                    sleep(6)
                    act_title_create_long_codes = self.driver.title
                    exp_title_create_long_codes = "Manage Longcodes"
                    if act_title_create_long_codes == exp_title_create_long_codes:
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A45'] = "test-longcode-add"
                        ws['B45'] = "Pass"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #self.add_s3bucket_file_upload()
                        self.driver.close()
                        assert True
                    else:
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A45'] = "test-longcode-add"
                        ws['B45'] = "Fail"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #self.add_s3bucket_file_upload()
                        self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                # self.add_s3bucket_file_upload()
                self.driver.close()



    def test_longcodes_update(self):
        if self.excelReader_test_cases('test_longcodes_update', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #self.add_s3bucket_file_upload()
                fileName = "data.xlsx"
                path = "assets"
                getFolderPath = os.path.join(path, fileName)
                self.rows = XLUtils.getRowCount(getFolderPath, 'Admin')
                sleep(4)
                for r in range(149, self.rows + 1):
                    self.longcode_phone_number = XLUtils.readData(getFolderPath, 'Admin', r, 1)
                    self.longcode_notes = XLUtils.readData(getFolderPath, 'Admin', r, 2)
                    self.longcode_provider = XLUtils.readData(getFolderPath, 'Admin', r, 3)
                    self.longcode_obj = LongcodesClass(self.driver)
                    self.longcode_obj.clickLongcode()
                    try:
                        self.longcode_obj.clickOrgLongcodeBtn()
                    except NoSuchElementException:
                        sleep(4)
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A46'] = "test-longcode-update"
                        ws['B46'] = "No organizations found"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        self.driver.close()
                    try:
                        if self.driver.find_element_by_class_name("td_hover").text == '14346604959':
                            self.driver.find_element_by_link_text("Edit").click()
                        else:
                            sleep(4)
                            book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                            ws = book.active
                            ws['A46'] = "test-longcode-update"
                            ws['B46'] = "No data Found"
                            book.save('Reports/result_admin_cases_output.xlsx')
                            # self.add_s3bucket_file_upload()
                            self.driver.close()
                        self.longcode_obj.setlongcodeNumber(self.longcode_phone_number)
                        self.longcode_obj.setlongcodeNotes(self.longcode_notes)
                        self.longcode_obj.setlongcodeProvider(self.longcode_provider)
                        sleep(2)
                        self.longcode_obj.clickActiveCheckbox()
                        sleep(2)
                        self.longcode_obj.clickupdatelongcodebtn()
                        sleep(6)
                        act_title_update_long_codes = self.driver.title
                        exp_title_update_long_codes = "Manage Longcodes"
                        if act_title_update_long_codes == exp_title_update_long_codes:
                            book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                            ws = book.active
                            ws['A46'] = "test-longcode-update"
                            ws['B46'] = "Pass"
                            book.save('Reports/result_admin_cases_output.xlsx')
                            #self.add_s3bucket_file_upload()
                            self.driver.close()
                            assert True
                    except NoSuchElementException:
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A46'] = "test-longcode-update"
                        ws['B46'] = "Fail"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #self.add_s3bucket_file_upload()
                        self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                # self.add_s3bucket_file_upload()
                self.driver.close()



    def test_sending_demo_add(self):
        if self.excelReader_test_cases('test_sending_demo_add', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #self.add_s3bucket_file_upload()
                fileName = "data.xlsx"
                path = "assets"
                getFolderPath = os.path.join(path, fileName)
                self.rows = XLUtils.getRowCount(getFolderPath, 'Admin')
                sleep(4)
                for r in range(154, self.rows + 1):
                    self.sending_demo_phone_number = XLUtils.readData(getFolderPath, 'Admin', r, 1)
                    self.sending_demo_obj = SendingDemoClass(self.driver)
                    sleep(2)
                    self.sending_demo_obj.clickSendingDemo()
                    self.sending_demo_obj.clickDemoSendingCreateBtn()
                    sleep(2)
                    self.sending_demo_obj.setSendPhoneNumberName(self.sending_demo_phone_number)
                    sleep(2)
                    self.sending_demo_obj.clickSendingDemobtn()
                    sleep(4)
                    try:
                        success_toastr = self.driver.find_element_by_class_name("toast-error")
                        if success_toastr:
                            book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                            ws = book.active
                            ws['A47'] = "test-sending-demo-add"
                            ws['B47'] = "Fail"
                            book.save('Reports/result_admin_cases_output.xlsx')
                            #self.add_s3bucket_file_upload()
                            self.driver.close()
                            assert True
                    except NoSuchElementException:
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A47'] = "test-sending-demo-add"
                        ws['B47'] = "Pass"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #self.add_s3bucket_file_upload()
                        self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                # self.add_s3bucket_file_upload()
                self.driver.close()

    def test_zipwhip_sending_demo_add(self):
        if self.excelReader_test_cases('test_zipwhip_sending_demo_add', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #self.add_s3bucket_file_upload()
                fileName = "data.xlsx"
                path = "assets"
                getFolderPath = os.path.join(path, fileName)
                self.rows = XLUtils.getRowCount(getFolderPath, 'Admin')
                sleep(4)
                for r in range(159, self.rows + 1):
                    self.zipwhip_sending_demo_phone_number = XLUtils.readData(getFolderPath, 'Admin', r,1)
                    self.zipwhip_sending_message = XLUtils.readData(getFolderPath, 'Admin', r, 2)
                    self.zipwhip_sending_demo_obj = ZipwhipSendingDemoClass(self.driver)
                    try:
                        sleep(2)
                        self.zipwhip_sending_demo_obj.clickZipwhipSendingDemo()
                        sleep(2)
                        self.zipwhip_sending_demo_obj.setZipwhipSendPhoneNumberId(self.zipwhip_sending_demo_phone_number)
                        sleep(2)
                        self.zipwhip_sending_demo_obj.setZipwhipSendPhoneMessageId(self.zipwhip_sending_message)
                        sleep(2)
                        self.zipwhip_sending_demo_obj.clickZipwhipSendingDemobtn()
                        sleep(8)
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A48'] = "test-zipwhip-sending-demo-add"
                        ws['B48'] = "Pass"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #self.add_s3bucket_file_upload()
                        self.driver.close()
                        assert True
                    except NoSuchElementException:
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A48'] = "test-zipwhip-sending-demo-add"
                        ws['B48'] = "Fail"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #self.add_s3bucket_file_upload()
                        self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                # self.add_s3bucket_file_upload()
                self.driver.close()


    def test_signal_wire_demo(self):
        if self.excelReader_test_cases('test_signal_wire_demo', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                # self.add_s3bucket_file_upload()
                fileName = "data.xlsx"
                path = "assets"
                getFolderPath = os.path.join(path, fileName)
                self.rows = XLUtils.getRowCount(getFolderPath, 'Admin')
                sleep(4)
                for r in range(164, self.rows + 1):
                    self.signalwire_phone_number = XLUtils.readData(getFolderPath, 'Admin', r, 1)
                    self.signalwire_message = XLUtils.readData(getFolderPath, 'Admin', r, 2)
                    self.signalwire_obj = SignalWireDemoClass(self.driver)
                    try:
                        self.signalwire_obj.clickSignalWire()
                        self.signalwire_obj.setSignalWireMobileNumber(self.signalwire_phone_number)
                        self.signalwire_obj.setlongcodeFromDropdown()
                        self.signalwire_obj.setSignalWireMessage(self.signalwire_message)
                        self.signalwire_obj.clickSignalWireSubmitebtn()
                        sleep(4)
                        success_toastr = self.driver.find_element_by_class_name("toast-success")
                        if success_toastr:
                            sleep(6)
                            book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                            ws = book.active
                            ws['A49'] = "test-signalwire-demo"
                            ws['B49'] = "Pass"
                            book.save('Reports/result_admin_cases_output.xlsx')
                            # self.add_s3bucket_file_upload()
                            self.driver.close()
                            assert True
                    except NoSuchElementException:
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A49'] = "test-signalwire-demo"
                        ws['B49'] = "Fail"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        # self.add_s3bucket_file_upload()
                        self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                # self.add_s3bucket_file_upload()
                self.driver.close()

    def test_telnyx_api_demo_add(self):
        if self.excelReader_test_cases('test_telnyx_api_demo_add', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                # self.add_s3bucket_file_upload()
                fileName = "data.xlsx"
                path = "assets"
                getFolderPath = os.path.join(path, fileName)
                self.rows = XLUtils.getRowCount(getFolderPath, 'Admin')
                sleep(4)
                for r in range(169, self.rows + 1):
                    self.telnyx_phone_number = XLUtils.readData(getFolderPath, 'Admin', r, 1)
                    self.telnyx_message = XLUtils.readData(getFolderPath, 'Admin', r, 2)
                    self.telnyx_api_demo_obj = TelnyxApiDemoClass(self.driver)
                    self.telnyx_api_demo_obj.clickTelnyxApiDemo()
                    self.telnyx_api_demo_obj.setTelnyxApiPhoneNumberName(self.telnyx_phone_number)
                    self.telnyx_api_demo_obj.setTelnyxApiMessageName(self.telnyx_message)
                    self.telnyx_api_demo_obj.clickTelnyxApiDemobtn()
                    sleep(4)
                    try:
                        success_toastr = self.driver.find_element_by_class_name("toast-error")
                        if success_toastr:
                            sleep(6)
                            book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                            ws = book.active
                            ws['A50'] = "test-telnyx-api-demo"
                            ws['B50'] = "Fail"
                            book.save('Reports/result_admin_cases_output.xlsx')
                            # self.add_s3bucket_file_upload()
                            self.driver.close()
                            assert True
                    except NoSuchElementException:
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A50'] = "test-telnyx-api-demo"
                        ws['B50'] = "Pass"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        # self.add_s3bucket_file_upload()
                        self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                # self.add_s3bucket_file_upload()
                self.driver.close()





    def test_manage_unsend_message_update(self):
        if self.excelReader_test_cases('test_manage_unsend_message_update', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                self.manage_messages_obj = ManagemessagesClass(self.driver)
                self.manage_messages_obj.clickManagemessagebtn()
                sleep(4)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A51'] = "test-manage-unsend-message-update"
                ws['B51'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #self.add_s3bucket_file_upload()
                self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                self.driver.close()




    def test_manage_unsend_message_state_update(self):
        if self.excelReader_test_cases('test_manage_unsend_message_state_update', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #self.add_s3bucket_file_upload()
                self.manage_messages_obj = ManagemessagesClass(self.driver)
                self.manage_messages_obj.clickManagemessagebtn()
                sleep(4)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A52'] = "test-manage-unsend-message-state-update"
                ws['B52'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #self.add_s3bucket_file_upload()
                self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                # self.add_s3bucket_file_upload()
                self.driver.close()






    def test_sender_claims_update(self):
        if self.excelReader_test_cases('test_sender_claims_update', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #self.add_s3bucket_file_upload()
                self.sender_claims_obj = SenderClaimsClass(self.driver)
                self.sender_claims_obj.clickSenderClaims()
                sleep(3)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A53'] = "test-sender-claimed-update"
                ws['B53'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #self.add_s3bucket_file_upload()
                self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                # self.add_s3bucket_file_upload()
                self.driver.close()





    def test_friendly_messages_add(self):
        if self.excelReader_test_cases('test_friendly_messages_add', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #self.add_s3bucket_file_upload()
                fileName = "data.xlsx"
                path = "assets"
                getFolderPath = os.path.join(path, fileName)
                self.rows = XLUtils.getRowCount(getFolderPath, 'Admin')
                sleep(4)
                for r in range(174, self.rows + 1):
                    self.friendly_id = XLUtils.readData(getFolderPath, 'Admin', r, 1)
                    self.friendly_title = XLUtils.readData(getFolderPath, 'Admin', r, 2)
                    self.friendly_description = XLUtils.readData(getFolderPath, 'Admin', r, 3)
                    self.friendly_provider = XLUtils.readData(getFolderPath, 'Admin', r, 4)
                    self.friendly_message_obj = FriendlyerrorClass(self.driver)
                    self.friendly_message_obj.clickfriendlymessagebtn()
                    self.friendly_message_obj.clickcreatefriendlymessagebtn()
                    self.friendly_message_obj.seterrorid(self.friendly_id)
                    self.friendly_message_obj.seterrorfriendlytitle(self.friendly_title)
                    self.friendly_message_obj.seterrorfriendlydescription(self.friendly_description)
                    self.friendly_message_obj.seterrorfriendlyprovider(self.friendly_provider)
                    self.friendly_message_obj.click_create_friendly_messages_btn_class()
                    sleep(4)
                    act_title_create_friendly_message = self.driver.title
                    exp_title_create_friendly_message = "Friendly Message"
                    if act_title_create_friendly_message == exp_title_create_friendly_message:
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A54'] = "test-friendly-message-add"
                        ws['B54'] = "Pass"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #self.add_s3bucket_file_upload()
                        self.driver.close()
                        assert True
                    else:
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A54'] = "test-friendly-message-add"
                        ws['B54'] = "Fail"
                        book.save('Reports/result_admin_cases_output.xlsx')
                       # self.add_s3bucket_file_upload()
                        self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                # self.add_s3bucket_file_upload()
                self.driver.close()




    def test_friendly_messages_update(self):
        if self.excelReader_test_cases('test_friendly_messages_update', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #self.add_s3bucket_file_upload()
                fileName = "data.xlsx"
                path = "assets"
                getFolderPath = os.path.join(path, fileName)
                self.rows = XLUtils.getRowCount(getFolderPath, 'Admin')
                sleep(4)
                for r in range(179, self.rows + 1):
                    self.friendly_id = XLUtils.readData(getFolderPath, 'Admin', r, 1)
                    self.friendly_title = XLUtils.readData(getFolderPath, 'Admin', r, 2)
                    self.friendly_description = XLUtils.readData(getFolderPath, 'Admin', r, 3)
                    self.friendly_provider = XLUtils.readData(getFolderPath, 'Admin', r, 4)
                    self.friendly_message_obj = FriendlyerrorClass(self.driver)
                    self.friendly_message_obj.clickfriendlymessagebtn()
                    try:
                        if self.driver.find_element_by_xpath("//*[@id='tableData_message']/tbody/tr[1]/td[1]").text == '21603':
                            self.friendly_message_obj.click_edit_friendly_messages_btn_class()
                        else:
                            book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                            ws = book.active
                            ws['A55'] = "test-friendly-message-update"
                            ws['B55'] = "no data found"
                            book.save('Reports/result_admin_cases_output.xlsx')
                            # self.add_s3bucket_file_upload()
                            self.driver.close()

                    except NoSuchElementException:
                        sleep(4)
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A55'] = "test-friendly-message-update"
                        ws['B55'] = "no data found"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #self.add_s3bucket_file_upload()
                        self.driver.close()
                    self.friendly_message_obj.seterrorid(self.friendly_id)
                    self.friendly_message_obj.seterrorfriendlytitle(self.friendly_title)
                    self.friendly_message_obj.seterrorfriendlydescription(self.friendly_description)
                    self.friendly_message_obj.seterrorfriendlyprovider(self.friendly_provider)
                    self.friendly_message_obj.click_update_friendly_messages_btn_class()
                    sleep(4)
                    act_title_update_friendly_message = self.driver.title
                    exp_title_update_friendly_message = "Friendly Message"
                    if act_title_update_friendly_message == exp_title_update_friendly_message:
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A55'] = "test-friendly-message-update"
                        ws['B55'] = "Pass"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #self.add_s3bucket_file_upload()
                        self.driver.close()
                        assert True
                    else:
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A55'] = "test-friendly-message-update"
                        ws['B55'] = "Fail"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #self.add_s3bucket_file_upload()
                        self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                # self.add_s3bucket_file_upload()
                self.driver.close()


    def test_message_recycling(self):
        if self.excelReader_test_cases('test_message_recycling', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #self.add_s3bucket_file_upload()
                self.msg_recycling_obj = MessageRecyclingClass(self.driver)
                self.msg_recycling_obj.clickMessageRecycling()
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A56'] = "test-message-recycling"
                ws['B56'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #self.add_s3bucket_file_upload()
                self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                # self.add_s3bucket_file_upload()
                self.driver.close()




    def test_message_senders(self):
        if self.excelReader_test_cases('test_message_senders', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #self.add_s3bucket_file_upload()
                self.msg_senders_obj = MessageSendersClass(self.driver)
                self.msg_senders_obj.clickMessageSenders()
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A57'] = "test-message-senders"
                ws['B57'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #self.add_s3bucket_file_upload()
                self.driver.close()

            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                # self.add_s3bucket_file_upload()
                self.driver.close()



    def test_sending_messages_admin(self):
        if self.excelReader_test_cases('test_sending_messages_admin', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #self.add_s3bucket_file_upload()
                self.sending_msg_obj = SendingMessagesAdminClass(self.driver)
                self.sending_msg_obj.clickSendingMessages()
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A58'] = "test-sending-messages"
                ws['B58'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #self.add_s3bucket_file_upload()
                self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                # self.add_s3bucket_file_upload()
                self.driver.close()



    def test_assign_messages(self):
        if self.excelReader_test_cases('test_assign_messages', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                # self.add_s3bucket_file_upload()
                self.assign_msg_obj = AssignMessageClass(self.driver)
                self.assign_msg_obj.clickAssignMessage()
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A59'] = "test_assign_messages"
                ws['B59'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #self.add_s3bucket_file_upload()
                self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                # self.add_s3bucket_file_upload()
                self.driver.close()




    def test_profile_update(self):
        if self.excelReader_test_cases('test_profile_update', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #self.add_s3bucket_file_upload()
                fileName = "data.xlsx"
                path = "assets"
                getFolderPath = os.path.join(path, fileName)
                self.rows = XLUtils.getRowCount(getFolderPath, 'Admin')
                sleep(4)
                for r in range(184, self.rows + 1):
                    self.profile_phone_number = XLUtils.readData(getFolderPath, 'Admin', r, 1)
                    self.profile_current_password = XLUtils.readData(getFolderPath, 'Admin', r, 2)
                    self.profile_new_password = XLUtils.readData(getFolderPath, 'Admin', r, 3)
                    self.profile_confirm_password = XLUtils.readData(getFolderPath, 'Admin', r, 4)
                    self.profile_obj = ProfileClass(self.driver)
                    sleep(2)
                    self.profile_obj.clickProfile()
                    self.profile_obj.setProfilePhoneNumber(self.profile_phone_number)
                    self.profile_obj.clickSelectProfileState()
                    sleep(2)
                    self.profile_obj.clickdatepickertest()
                    sleep(1)
                    self.profile_obj.setProfilecurrentPassword(self.profile_current_password)
                    sleep(2)
                    self.profile_obj.setProfileNewPassword(self.profile_new_password)
                    sleep(2)
                    self.profile_obj.setProfilconfirmPassword(self.profile_confirm_password)
                    sleep(2)
                    self.profile_obj.clickupdateprofilebtn()
                    sleep(4)
                    try:
                        success_toastr = self.driver.find_element_by_class_name("toast-error")
                        if success_toastr:
                            sleep(6)
                            book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                            ws = book.active
                            ws['A60'] = "test-profile-update"
                            ws['B60'] = "Fail"
                            book.save('Reports/result_admin_cases_output.xlsx')
                            #self.add_s3bucket_file_upload()
                            self.driver.close()
                            assert True
                    except NoSuchElementException:
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A60'] = "test-profile-update"
                        ws['B60'] = "Pass"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #self.add_s3bucket_file_upload()
                        self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                # self.add_s3bucket_file_upload()
                self.driver.close()

    def test_ACH_Setup(self):
        if self.excelReader_test_cases('test_ACH_Setup', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                # self.add_s3bucket_file_upload()
                fileName = "data.xlsx"
                path = "assets"
                getFolderPath = os.path.join(path, fileName)
                self.rows = XLUtils.getRowCount(getFolderPath, 'Admin')
                sleep(4)
                for r in range(189, self.rows + 1):
                    self.ach_middle_name = XLUtils.readData(getFolderPath, 'Admin', r, 1)
                    self.ach_suffix = XLUtils.readData(getFolderPath, 'Admin', r, 2)
                    self.ach_address_line1 = XLUtils.readData(getFolderPath, 'Admin', r, 3)
                    self.ach_address_line2 = XLUtils.readData(getFolderPath, 'Admin', r, 4)
                    self.ach_city = XLUtils.readData(getFolderPath, 'Admin', r, 5)
                    self.ach_postal_code = XLUtils.readData(getFolderPath, 'Admin', r, 6)
                    self.ach_security_number = XLUtils.readData(getFolderPath, 'Admin', r, 7)
                    self.ach_account_number = XLUtils.readData(getFolderPath, 'Admin', r, 8)
                    self.ach_routing_number = XLUtils.readData(getFolderPath, 'Admin', r, 9)
                    self.ach_confirm_routing_number = XLUtils.readData(getFolderPath, 'Admin', r, 10)
                    self.ach_phone_number = XLUtils.readData(getFolderPath, 'Admin', r, 11)
                    self.ach_obj = AchSetupClass(self.driver)
                    self.ach_obj.clickProfile()
                    self.ach_obj.clickAchbtn()
                    self.ach_obj.setAchMiddleName(self.ach_middle_name)
                    self.ach_obj.setAchSuffix(self.ach_suffix)
                    self.ach_obj.setAchPhone_number(self.ach_phone_number)
                    self.ach_obj.setAchAddressLine1(self.ach_address_line1)
                    self.ach_obj.setAchAddressLine2(self.ach_address_line2)
                    self.ach_obj.setAchCity(self.ach_city)
                    sleep(2)
                    self.ach_obj.setAchPostalCode(self.ach_postal_code)
                    sleep(2)
                    self.ach_obj.setAchSecurityNumber(self.ach_security_number)
                    sleep(2)
                    self.ach_obj.clicksubmitAchbtn()
                    sleep(4)
                    try:
                        success_toastr = self.driver.find_element_by_class_name("toast-error")
                        if success_toastr:
                            sleep(6)
                            book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                            ws = book.active
                            ws['A61'] = "test-ach-setup"
                            ws['B61'] = "Fail"
                            book.save('Reports/result_admin_cases_output.xlsx')
                            # self.add_s3bucket_file_upload()
                            self.driver.close()
                            assert True
                    except NoSuchElementException:
                        sleep(4)
                        self.ach_obj.clickSelectACHtype()
                        sleep(2)
                        self.ach_obj.setAchaccountNumber(self.ach_account_number)
                        sleep(2)
                        self.ach_obj.setAchRoutingNumber(self.ach_routing_number)
                        sleep(2)
                        self.ach_obj.setAchConfirmRoutingNumber(self.ach_confirm_routing_number)
                        sleep(2)
                        self.ach_obj.clickupdatesubmitAchbtn()
                        sleep(4)
                        try:
                            success_toastr = self.driver.find_element_by_class_name("toast-error")
                            if success_toastr:
                                sleep(6)
                                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                                ws = book.active
                                ws['A61'] = "test-ach-setup"
                                ws['B61'] = "Fail"
                                book.save('Reports/result_admin_cases_output.xlsx')
                                # self.add_s3bucket_file_upload()
                                self.driver.close()
                                assert True
                        except NoSuchElementException:
                            book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                            ws = book.active
                            ws['A61'] = "test-ach-setup"
                            ws['B61'] = "Pass"
                            book.save('Reports/result_admin_cases_output.xlsx')
                            # self.add_s3bucket_file_upload()
                            self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                # self.add_s3bucket_file_upload()
                self.driver.close()



    def test_cron_scheduler_add(self):
        if self.excelReader_test_cases('test_cron_scheduler_add', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #self.add_s3bucket_file_upload()
                sleep(8)
                path = "assets"
                fileName = "data.xlsx"
                getFolderPath = os.path.join(path, fileName)
                self.rows = XLUtils.getRowCount(getFolderPath, 'Admin')
                for r in range(194, self.rows + 1):
                    self.cron_name = XLUtils.readData(getFolderPath, 'Admin', r, 1)
                    self.cron_link = XLUtils.readData(getFolderPath, 'Admin', r, 2)
                    self.cron_time = XLUtils.readData(getFolderPath, 'Admin', r, 3)
                    sleep(2)
                    self.cron_scheduler_obj = CronSchedulerClass(self.driver)
                    sleep(2)
                    self.cron_scheduler_obj.clickCronScheduler()
                    sleep(2)
                    self.cron_scheduler_obj.clickCronSchedulerCreate()
                    sleep(2)
                    self.cron_scheduler_obj.setCronName(self.cron_name)
                    sleep(2)
                    self.cron_scheduler_obj.setCronLink(self.cron_link)
                    sleep(2)
                    self.cron_scheduler_obj.clickSelectCronRun()
                    sleep(2)
                    self.cron_scheduler_obj.setCronTime(self.cron_time)
                    sleep(2)
                    self.cron_scheduler_obj.clickActiveCheckbox()
                    sleep(2)
                    self.cron_scheduler_obj.clickcreatecronschedulerbtn()
                    sleep(12)
                    act_title_add = self.driver.title
                    exp_title_test = "Cron Scheduler"
                    if act_title_add == exp_title_test:
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A62'] = "test-cron-scheduler-add"
                        ws['B62'] = "Pass"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #self.add_s3bucket_file_upload()
                        self.driver.close()
                        assert True
                    else:
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A62'] = "test-cron-scheduler-add"
                        ws['B62'] = "Fail"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #self.add_s3bucket_file_upload()
                        self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                # self.add_s3bucket_file_upload()
                self.driver.close()


    def test_cron_scheduler_update(self):
        if self.excelReader_test_cases('test_cron_scheduler_update', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #self.add_s3bucket_file_upload()
                sleep(8)
                path = "assets"
                fileName = "data.xlsx"
                getFolderPath = os.path.join(path, fileName)
                self.rows = XLUtils.getRowCount(getFolderPath, 'Admin')
                for r in range(199, self.rows + 1):
                    self.cron_name = XLUtils.readData(getFolderPath, 'Admin', r, 1)
                    self.cron_link = XLUtils.readData(getFolderPath, 'Admin', r, 2)
                    self.cron_time = XLUtils.readData(getFolderPath, 'Admin', r, 3)
                    sleep(2)
                    self.cron_scheduler_obj = CronSchedulerClass(self.driver)
                    sleep(2)
                    self.cron_scheduler_obj.clickCronScheduler()
                    sleep(2)
                    try:
                        if self.driver.find_element_by_class_name("td_hover").text == 'Automation-cron-Long Code Cron-add':
                           self.cron_scheduler_obj.clickEditcronschedulerbtn()
                        else:
                            sleep(4)
                            book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                            ws = book.active
                            ws['A63'] = "test-cron-scheduler-update"
                            ws['B63'] = "No data found"
                            book.save('Reports/result_admin_cases_output.xlsx')
                            # self.add_s3bucket_file_upload()
                            self.driver.close()
                    except NoSuchElementException:
                        sleep(4)
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A63'] = "test-cron-scheduler-update"
                        ws['B63'] = "No data found"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #self.add_s3bucket_file_upload()
                        self.driver.close()
                    sleep(2)
                    self.cron_scheduler_obj.setCronName(self.cron_name)
                    sleep(2)
                    self.cron_scheduler_obj.setCronLink(self.cron_link)
                    sleep(2)
                    self.cron_scheduler_obj.clickSelectCronRun()
                    sleep(2)
                    self.cron_scheduler_obj.setCronTime(self.cron_time)
                    sleep(2)
                    self.cron_scheduler_obj.clickActiveCheckbox()
                    sleep(2)
                    self.cron_scheduler_obj.clickupdatecronschedulerbtn()
                    sleep(6)
                    act_title_update_cron_scheduler = self.driver.title
                    exp_title_update_cron_scheduler = "Cron Scheduler"
                    if act_title_update_cron_scheduler == exp_title_update_cron_scheduler:
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A63'] = "test-cron-scheduler-update"
                        ws['B63'] = "Pass"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        #self.add_s3bucket_file_upload()
                        self.driver.close()
                        assert True
                    else:
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A63'] = "test-cron-scheduler-update"
                        ws['B63'] = "Fail"
                        book.save('Reports/result_admin_cases_output.xlsx')
                       # self.add_s3bucket_file_upload()
                        self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                # self.add_s3bucket_file_upload()
                self.driver.close()



    def test_log_file_download_data(self):
        if self.excelReader_test_cases('test_log_file_download_data', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #self.add_s3bucket_file_upload()
                self.log_file_obj = LogFileClass(self.driver)
                self.log_file_obj.clickLogFile()
                try:
                    self.log_file_obj.clickdownloadLogFile()
                    sleep(4)
                    book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                    ws = book.active
                    ws['A64'] = "test-log-file-download-data"
                    ws['B64'] = "Pass"
                    book.save('Reports/result_admin_cases_output.xlsx')
                    #self.add_s3bucket_file_upload()
                    self.driver.close()
                except NoSuchElementException:
                    book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                    ws = book.active
                    ws['A64'] = "test-log-file-download-data"
                    ws['B64'] = "Fail"
                    book.save('Reports/result_admin_cases_output.xlsx')
                    #self.add_s3bucket_file_upload()
                    self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                # self.add_s3bucket_file_upload()
                self.driver.close()



    def test_file_import(self):
        if self.excelReader_test_cases('test_file_import', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #self.add_s3bucket_file_upload()
                self.file_import_obj = FileimportClass(self.driver)
                sleep(2)
                self.file_import_obj.clickfileimportbtn()
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A65'] = "test-file-import-part-1"
                ws['B65'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                # self.add_s3bucket_file_upload()
                self.driver.close()







    def test_file_ignore_process(self):
        if self.excelReader_test_cases('test_file_ignore_process', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #self.add_s3bucket_file_upload()
                self.file_ignore_obj = FileIgnoreClass(self.driver)
                sleep(2)
                self.file_ignore_obj.clickfileignorebtn()
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A67'] = "test-file-ignore-process"
                ws['B67'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #self.add_s3bucket_file_upload()
                self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                # self.add_s3bucket_file_upload()
                self.driver.close()


    def test_processed_file_view(self):
        if self.excelReader_test_cases('test_processed_file_view', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #self.add_s3bucket_file_upload()
                self.processed_file_obj = ProcessedFilesClass(self.driver)
                self.processed_file_obj.clickProcessedFile()
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A68'] = "test-processed-file-view"
                ws['B68'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #self.add_s3bucket_file_upload()
                self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                # self.add_s3bucket_file_upload()
                self.driver.close()





    def test_import_csv_add(self):
        if self.excelReader_test_cases('test_import_csv_add', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #self.add_s3bucket_file_upload()
                self.import_csv_obj = ImportCsvClass(self.driver)
                self.import_csv_obj.clickImportCsv()
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A69'] = "test-import-csv-add"
                ws['B69'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #self.add_s3bucket_file_upload()
                self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                # self.add_s3bucket_file_upload()
                self.driver.close()




    def test_delimited_import_file(self):
        if self.excelReader_test_cases('test_delimited_import_file', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                sleep(2)
                self.delimited_obj = delimitedimportClass(self.driver)
                sleep(2)
                self.delimited_obj.clickdelimitedbtn()
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A70'] = "test-delimited-file-import"
                ws['B70'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #self.add_s3bucket_file_upload()
                self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                # self.add_s3bucket_file_upload()
                self.driver.close()


    def test_artiva_btn(self):
        if self.excelReader_test_cases('test_artiva_btn', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #self.add_s3bucket_file_upload()
                sleep(2)
                self.artiva_obj = ArtivaimportClass(self.driver)
                sleep(2)
                self.artiva_obj.clickartivabtn()
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A71'] = "test-artiva-import"
                ws['B71'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #self.add_s3bucket_file_upload()
                self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                # self.add_s3bucket_file_upload()
                self.driver.close()









    def test_delete_report_scheduler(self):
        if self.excelReader_test_cases('test_delete_report_scheduler', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                # s3_file_upload = self.add_s3bucket_file_upload()
                self.rpt_schd_obj = ReportSchedulerClass(self.driver)
                sleep(2)
                self.rpt_schd_obj.clickReportScheduler()
                try:
                    if self.driver.find_element_by_class_name("covert_title").text == 'Automation-report-scheduler-update':
                        self.rpt_schd_obj.clickRemoveReportScheduler()
                    else:
                        sleep(4)
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A72'] = "test-delete-report-scheduler"
                        ws['B72'] = "No data found"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        self.driver.close()
                    sleep(3)
                    self.rpt_schd_obj.clickConfirmDelete()
                    sleep(6)
                    book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                    ws = book.active
                    ws['A72'] = "test-delete-report-scheduler"
                    ws['B72'] = "Pass"
                    book.save('Reports/result_admin_cases_output.xlsx')
                    # s3_file_upload = self.add_s3bucket_file_upload()
                    self.driver.close()
                    assert True
                except NoSuchElementException:
                    book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                    ws = book.active
                    ws['A72'] = "test-delete-report-scheduler"
                    ws['B72'] = "Fail"
                    book.save('Reports/result_admin_cases_output.xlsx')
                    # s3_file_upload = self.add_s3bucket_file_upload()
                    self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                # s3_file_upload = self.add_s3bucket_file_upload()
                self.driver.close()




    def test_mapped_media_delete(self):
        if self.excelReader_test_cases('test_mapped_media_delete', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #s3_file_upload = self.add_s3bucket_file_upload()
                self.mapped_obj = MappedmediaClass(self.driver)
                sleep(2)
                self.mapped_obj.clickmappedmedia()
                sleep(2)
                try:
                    if self.driver.find_element_by_class_name("file_name").text == 'Automation-mapped-media-updated':
                        self.mapped_obj.clickRemoveMappedMediaLink()
                    else:
                        sleep(4)
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A73'] = "test-mapped-media-delete"
                        ws['B73'] = "No data Found"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        self.driver.close()
                    sleep(4)
                    self.mapped_obj.clickConfirmDelete()
                    sleep(4)
                    book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                    ws = book.active
                    ws['A73'] = "test-mapped-media-delete"
                    ws['B73'] = "Pass"
                    book.save('Reports/result_admin_cases_output.xlsx')
                    #s3_file_upload = self.add_s3bucket_file_upload()
                    self.driver.close()
                    assert True
                except NoSuchElementException:
                    book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                    ws = book.active
                    ws['A73'] = "test-mapped-media-delete"
                    ws['B73'] = "Fail"
                    book.save('Reports/result_admin_cases_output.xlsx')
                    #s3_file_upload = self.add_s3bucket_file_upload()
                    self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                # s3_file_upload = self.add_s3bucket_file_upload()
                self.driver.close()


    def test_delete_phone_lookup_rule(self):
        if self.excelReader_test_cases('test_delete_phone_lookup_rule', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                # s3_file_upload = self.add_s3bucket_file_upload()
                self.phnlookup_obj = PhoneLookupRuleClass(self.driver)
                self.phnlookup_obj.clickPhoneLookup()
                try:
                    if self.driver.find_element_by_class_name("title").text == 'Automation-phone-lookup-update':
                        self.phnlookup_obj.clickDeletePhoneLookup()
                    else:
                        sleep(4)
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A74'] = "test-delete-phone-lookup"
                        ws['B74'] = "No data Found"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        self.driver.close()
                    sleep(4)
                    self.phnlookup_obj.clickConfirmDeletePhnLookup()
                    sleep(4)
                    book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                    ws = book.active
                    ws['A74'] = "test-delete-phone-lookup"
                    ws['B74'] = "Pass"
                    book.save('Reports/result_admin_cases_output.xlsx')
                    #s3_file_upload = self.add_s3bucket_file_upload()
                    self.driver.close()
                    assert True
                except NoSuchElementException:
                    book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                    ws = book.active
                    ws['A74'] = "test-delete-phone-lookup"
                    ws['B74'] = "Fail"
                    book.save('Reports/result_admin_cases_output.xlsx')
                    #s3_file_upload = self.add_s3bucket_file_upload()
                    self.driver.close()
                    assert False
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                # s3_file_upload = self.add_s3bucket_file_upload()
                self.driver.close()



    def test_campaign_delete(self):
        if self.excelReader_test_cases('test_campaign_delete', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #s3_file_upload = self.add_s3bucket_file_upload()
                self.campaign_obj = CampaignClass(self.driver)
                sleep(2)
                self.campaign_obj.clickCampaigns()
                try:
                    if self.driver.find_element_by_class_name("campaign_sec").text == 'Automation-campaign-update':
                        self.campaign_obj.clickRemoveCampaign()
                    else:
                        sleep(4)
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A75'] = "test-delete-campaign"
                        ws['B75'] = "No data found"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        self.driver.close()
                    sleep(3)
                    self.campaign_obj.clickConfirmDelete()
                    sleep(6)
                    book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                    ws = book.active
                    ws['A75'] = "test-delete-campaign"
                    ws['B75'] = "Pass"
                    book.save('Reports/result_admin_cases_output.xlsx')
                    #s3_file_upload = self.add_s3bucket_file_upload()
                    self.driver.close()
                    assert True
                except NoSuchElementException:
                    book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                    ws = book.active
                    ws['A75'] = "test-delete-campaign"
                    ws['B75'] = "Fail"
                    book.save('Reports/result_admin_cases_output.xlsx')
                    #s3_file_upload = self.add_s3bucket_file_upload()
                    self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                # s3_file_upload = self.add_s3bucket_file_upload()
                self.driver.close()





    def test_delete_common_response(self):
        if self.excelReader_test_cases('test_delete_common_response', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #s3_file_upload = self.add_s3bucket_file_upload()
                self.response_obj = CommonResponseClass(self.driver)
                self.response_obj.clickCommonResponse()
                try:

                    if self.driver.find_element_by_class_name("td_hover").text == 'Automation-common-responses-update':
                        self.response_obj.clickRemoveCommonResponse()
                    else:
                        sleep(4)
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A76'] = "test-delete-common-responses"
                        ws['B76'] = "No data Found"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        self.driver.close()
                    sleep(3)
                    self.response_obj.clickConfirmDelete()
                    sleep(6)
                    book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                    ws = book.active
                    ws['A76'] = "test-delete-common-responses"
                    ws['B76'] = "Pass"
                    book.save('Reports/result_admin_cases_output.xlsx')
                    #s3_file_upload = self.add_s3bucket_file_upload()
                    self.driver.close()
                    assert True
                except NoSuchElementException:
                    book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                    ws = book.active
                    ws['A76'] = "test-delete-common-responses"
                    ws['B76'] = "Fail"
                    book.save('Reports/result_admin_cases_output.xlsx')
                    #s3_file_upload = self.add_s3bucket_file_upload()
                    self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                # s3_file_upload = self.add_s3bucket_file_upload()
                self.driver.close()





    def test_message_merge_delete(self):
        if self.excelReader_test_cases('test_message_merge_delete', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #s3_file_upload = self.add_s3bucket_file_upload()
                self.msgmerge_obj = MessagemergeClass(self.driver)
                self.msgmerge_obj.clickMessageMerge()
                try:
                    if self.driver.find_element_by_class_name("td_hover").text == 'Automation-message-merge-update':
                        self.msgmerge_obj.clickRemoveMessageMerge()
                    else:
                        sleep(4)
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A77'] = "test-message-merge-delete"
                        ws['B77'] = "No data found"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        self.driver.close()
                    sleep(3)
                    self.msgmerge_obj.clickConfirmDelete()
                    sleep(6)
                    book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                    ws = book.active
                    ws['A77'] = "test-message-merge-delete"
                    ws['B77'] = "Pass"
                    book.save('Reports/result_admin_cases_output.xlsx')
                    #s3_file_upload = self.add_s3bucket_file_upload()
                    self.driver.close()
                    assert True
                except NoSuchElementException:
                    book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                    ws = book.active
                    ws['A77'] = "test-message-merge-delete"
                    ws['B77'] = "Fail"
                    book.save('Reports/result_admin_cases_output.xlsx')
                    #s3_file_upload = self.add_s3bucket_file_upload()
                    self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                # s3_file_upload = self.add_s3bucket_file_upload()
                self.driver.close()





    def test_Ignore_link_delete(self):
        if self.excelReader_test_cases('test_Ignore_link_delete', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #self.add_s3bucket_file_upload()
                self.Ignore_obj = IgnoreLinkClass(self.driver)
                self.Ignore_obj.clickIgnoreLink()
                try:
                    if self.driver.find_element_by_class_name("td_hover").text == 'Update-automation.com':
                        self.Ignore_obj.clickRemoveIgnoreLink()
                    else:
                        sleep(4)
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A78'] = "test-ignore-delete"
                        ws['B78'] = "No data found"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        # self.add_s3bucket_file_upload()
                        self.driver.close()
                    sleep(3)
                    self.Ignore_obj.clickConfirmDelete()
                    sleep(6)
                    book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                    ws = book.active
                    ws['A78'] = "test-ignore-delete"
                    ws['B78'] = "Pass"
                    book.save('Reports/result_admin_cases_output.xlsx')
                    #s3_file_upload = self.add_s3bucket_file_upload()
                    self.driver.close()
                    assert True
                except NoSuchElementException:
                    book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                    ws = book.active
                    ws['A78'] = "test-ignore-delete"
                    ws['B78'] = "Fail"
                    book.save('Reports/result_admin_cases_output.xlsx')
                    #s3_file_upload = self.add_s3bucket_file_upload()
                    self.driver.close()
                    assert False
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                # s3_file_upload = self.add_s3bucket_file_upload()
                self.driver.close()








    def test_Organizations_delete(self):
        if self.excelReader_test_cases('test_Organizations_delete', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #self.add_s3bucket_file_upload()
                self.org_obj = OrganizationsClass(self.driver)
                self.org_obj.clickOrganizations()
                try:

                    if self.driver.find_element_by_class_name("custom_org_link").text == 'Automation-organization-updated':
                        self.org_obj.clickRemoveOrgLink()
                    else:
                        sleep(4)
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A79'] = "test-organization-delete"
                        ws['B79'] = "no data found"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        # self.add_s3bucket_file_upload()
                        self.driver.close()
                    sleep(3)
                    self.org_obj.clickConfirmDelete()
                    sleep(4)
                    book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                    ws = book.active
                    ws['A79'] = "test-organization-delete"
                    ws['B79'] = "Pass"
                    book.save('Reports/result_admin_cases_output.xlsx')
                    #s3_file_upload = self.add_s3bucket_file_upload()
                    self.driver.close()
                    assert True
                except NoSuchElementException:
                    book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                    ws = book.active
                    ws['A79'] = "test-organization-delete"
                    ws['B79'] = "Fail"
                    book.save('Reports/result_admin_cases_output.xlsx')
                    #s3_file_upload = self.add_s3bucket_file_upload()
                    self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                # s3_file_upload = self.add_s3bucket_file_upload()
                self.driver.close()





    def test_api_key_delete(self):
        if self.excelReader_test_cases('test_api_key_delete', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                self.apikey_obj = ApiKeyClass(self.driver)
                self.apikey_obj.clickorgbtn()
                try:
                    self.apikey_obj.clickorgapibtn()
                except NoSuchElementException:
                    sleep(4)
                    book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                    ws = book.active
                    ws['A80'] = "test-delete-api-key"
                    ws['B80'] = "Organization not found"
                    book.save('Reports/result_admin_cases_output.xlsx')
                    self.driver.close()
                sleep(2)
                try:
                    sleep(3)
                    if self.driver.find_element_by_xpath("//*[@id='tableData_api']/tbody/tr[1]/td[1]").text == 'Automation-api-key-update':
                        self.apikey_obj.clickRemoveApikeyLink()
                    else:
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A80'] = "test-delete-api-key"
                        ws['B80'] = "No data Found"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        # self.add_s3bucket_file_upload()
                        self.driver.close()
                    self.apikey_obj.clickConfirmDelete()
                    sleep(4)
                    book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                    ws = book.active
                    ws['A80'] = "test-delete-api-key"
                    ws['B80'] = "Pass"
                    book.save('Reports/result_admin_cases_output.xlsx')
                    #s3_file_upload = self.add_s3bucket_file_upload()
                    self.driver.close()
                    assert True
                except NoSuchElementException:
                    book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                    ws = book.active
                    ws['A80'] = "test-delete-api-key"
                    ws['B80'] = "Fail"
                    book.save('Reports/result_admin_cases_output.xlsx')
                    #self.add_s3bucket_file_upload()
                    self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                # s3_file_upload = self.add_s3bucket_file_upload()
                self.driver.close()




    def test_Vcard_delete(self):
        if self.excelReader_test_cases('test_Vcard_delete', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #self.add_s3bucket_file_upload()
                self.vcard_obj = VcardClass(self.driver)
                self.vcard_obj.clickVcard()
                try:

                    if self.driver.find_element_by_class_name("td_hover").text == 'Automation-vcard-update':

                        self.vcard_obj.clickRemoveVcardLink()
                    else:
                        sleep(2)
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A81'] = "test-vcard-delete"
                        ws['B81'] = "No data found"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        # self.add_s3bucket_file_upload()
                        self.driver.close()
                    sleep(3)
                    self.vcard_obj.clickConfirmDelete()
                    sleep(6)
                    book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                    ws = book.active
                    ws['A81'] = "test-vcard-delete"
                    ws['B81'] = "Pass"
                    book.save('Reports/result_admin_cases_output.xlsx')
                    #self.add_s3bucket_file_upload()
                    self.driver.close()
                    assert True
                except NoSuchElementException:
                    book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                    ws = book.active
                    ws['A81'] = "test-vcard-delete"
                    ws['B81'] = "Fail"
                    book.save('Reports/result_admin_cases_output.xlsx')
                    #self.add_s3bucket_file_upload()
                    self.driver.close()
                    assert False
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                # s3_file_upload = self.add_s3bucket_file_upload()
                self.driver.close()



    def test_business_unit_delete(self):
        if self.excelReader_test_cases('test_business_unit_delete', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #self.add_s3bucket_file_upload()
                self.unit_obj = BusinessUnitClass(self.driver)
                self.unit_obj.clickBuisnessUnit()
                try:
                    if self.driver.find_element_by_class_name("td_hover").text == 'Automation-Business-unit-update':
                        self.unit_obj.clickRemoveunitLink()
                    else:
                        sleep(4)
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A82'] = "test-delete-business-units"
                        ws['B82'] = "No data found"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        # self.add_s3bucket_file_upload()
                        self.driver.close()
                    sleep(3)
                    self.unit_obj.clickConfirmDelete()
                    sleep(6)
                    book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                    ws = book.active
                    ws['A82'] = "test-delete-business-units"
                    ws['B82'] = "Pass"
                    book.save('Reports/result_admin_cases_output.xlsx')
                    #self.add_s3bucket_file_upload()
                    self.driver.close()
                    assert True
                except NoSuchElementException:
                    book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                    ws = book.active
                    ws['A82'] = "test-delete-business-units"
                    ws['B82'] = "Fail"
                    book.save('Reports/result_admin_cases_output.xlsx')
                    #self.add_s3bucket_file_upload()
                    self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                # s3_file_upload = self.add_s3bucket_file_upload()
                self.driver.close()


    def test_longcodes_delete(self):
        if self.excelReader_test_cases('test_longcodes_delete', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                self.longcode_obj = LongcodesClass(self.driver)
                self.longcode_obj.clickLongcode()
                sleep(2)
                try:
                    self.longcode_obj.clickOrgLongcodeBtn()
                    if self.driver.find_element_by_class_name("td_hover").text == '14346604956':
                        self.longcode_obj.clickRemovelongcodeLink()
                    else:
                        sleep(4)
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A83'] = "test-longcode-delete"
                        ws['B83'] = "No data found"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        # self.add_s3bucket_file_upload()
                        self.driver.close()
                    sleep(3)
                    self.longcode_obj.clickConfirmDelete()
                    sleep(6)
                    book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                    ws = book.active
                    ws['A83'] = "test-longcode-delete"
                    ws['B83'] = "Pass"
                    book.save('Reports/result_admin_cases_output.xlsx')
                    #self.add_s3bucket_file_upload()
                    self.driver.close()
                    assert True
                except NoSuchElementException:
                    book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                    ws = book.active
                    ws['A83'] = "test-longcode-delete"
                    ws['B83'] = "Fail"
                    book.save('Reports/result_admin_cases_output.xlsx')
                    #self.add_s3bucket_file_upload()
                    self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                # s3_file_upload = self.add_s3bucket_file_upload()
                self.driver.close()




    def test_friendly_message_delete(self):
        if self.excelReader_test_cases('test_friendly_message_delete', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #self.add_s3bucket_file_upload()
                self.friendly_message_obj = FriendlyerrorClass(self.driver)
                self.friendly_message_obj.clickfriendlymessagebtn()
                try:
                    if self.driver.find_element_by_xpath("//*[@id='tableData_message']/tbody/tr[1]/td[1]").text == '216033':
                        self.friendly_message_obj.clickRemovefriendlymessageLink()
                    else:
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A85'] = "test-friendly-message-delete"
                        ws['B85'] = "No data found"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        # self.add_s3bucket_file_upload()
                        self.driver.close()
                    sleep(3)
                    self.friendly_message_obj.clickConfirmDelete()
                    sleep(6)
                    book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                    ws = book.active
                    ws['A85'] = "test-friendly-message-delete"
                    ws['B85'] = "Pass"
                    book.save('Reports/result_admin_cases_output.xlsx')
                    #self.add_s3bucket_file_upload()
                    self.driver.close()
                    assert True
                except NoSuchElementException:
                    book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                    ws = book.active
                    ws['A85'] = "test-friendly-message-delete"
                    ws['B85'] = "Fail"
                    book.save('Reports/result_admin_cases_output.xlsx')
                    #self.add_s3bucket_file_upload()
                    self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                # s3_file_upload = self.add_s3bucket_file_upload()
                self.driver.close()




    def test_cron_scheduler_delete(self):
        if self.excelReader_test_cases('test_cron_scheduler_delete', 'test_cases') == "success":
            act_title = self.check_login()
            exp_title = "Admin Dashboard"
            if act_title == exp_title:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_admin_cases_output.xlsx')
                #self.add_s3bucket_file_upload()
                sleep(2)
                self.cron_scheduler_obj = CronSchedulerClass(self.driver)
                self.cron_scheduler_obj.clickCronScheduler()
                try:
                    if self.driver.find_element_by_class_name("td_hover").text == 'Automation-cron-Long Code Cron-update':
                       self.cron_scheduler_obj.clickRemovecronschedulerLink()
                    else:
                        sleep(4)
                        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                        ws = book.active
                        ws['A86'] = "test-cron-scheduler-delete"
                        ws['B86'] = "No data Found"
                        book.save('Reports/result_admin_cases_output.xlsx')
                        # self.add_s3bucket_file_upload()
                        self.driver.close()
                    sleep(3)
                    self.cron_scheduler_obj.clickConfirmDelete()
                    sleep(6)
                    book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                    ws = book.active
                    ws['A86'] = "test-cron-scheduler-delete"
                    ws['B86'] = "Pass"
                    book.save('Reports/result_admin_cases_output.xlsx')
                    #self.add_s3bucket_file_upload()
                    self.driver.close()
                    assert True
                except NoSuchElementException:
                    book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                    ws = book.active
                    ws['A86'] = "test-cron-scheduler-delete"
                    ws['B86'] = "Fail"
                    book.save('Reports/result_admin_cases_output.xlsx')
                    #self.add_s3bucket_file_upload()
                    self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_admin_cases_output.xlsx')
                # s3_file_upload = self.add_s3bucket_file_upload()
                self.driver.close()






    def test_excel_empty_row_remove(self):
        book = openpyxl.load_workbook('Reports/result_admin_cases_output.xlsx')
        ws = book.active
        index_row = []
        # loop each row in column A
        for i in range(1, ws.max_row):
            # define emptiness of cell
            if ws.cell(i, 1).value is None:
                # collect indexes of rows
                index_row.append(i)
        # loop each index value
        for row_del in range(len(index_row)):
            ws.delete_rows(idx=index_row[row_del], amount=1)
            # exclude offset of rows through each iteration
            index_row = list(map(lambda k: k - 1, index_row))
        book.save('Reports/result_admin_cases_output.xlsx')


    def test_s3bucket_file_upload(self):
        client_s3 = boto3.client('s3', aws_access_key_id=aws_access_key_id,
                                 aws_secret_access_key=aws_secret_access_key)
        with open('Reports/result_admin_cases_output.xlsx', 'rb') as data:
            client_s3.upload_fileobj(data, bucket_name, 'Reports/result_' + str(date) + '.xlsx')

