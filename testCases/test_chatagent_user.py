from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.common.exceptions import NoSuchElementException
from pageObjects.Chat_agent_user_LoginPage import ChatagentLoginClass
from pageObjects.chatagent_new_response import chatagentnewresponseClass
from pageObjects.chatagent_past_response import chatagentpastresponseClass
from pageObjects.chatagent_ignore_response import chatagentignoreresponseClass
from pageObjects.chat_agent_user_profile import chatagentUserprofileClass
import boto3
import os
from dotenv import load_dotenv
load_dotenv()
aws_access_key_id = os.getenv('AWS_ACCESS_KEY')
aws_secret_access_key = os.getenv('AWS_SECRET_KEY')
bucket_name = os.getenv('AWS_BUCKET_NAME')
from utilities import XLUtils
from datetime import date
date = date.today()
from openpyxl.workbook import Workbook
from openpyxl.styles import Font
import openpyxl
from time import sleep

class Test_004_chatagent_User:



    def excelReader_test_cases(self, method, sheet):
        path = "assets"
        fileName = "test_cases.xlsx"
        getFolderPath = os.path.join(path, fileName)
        self.rows = XLUtils.getRowCount(getFolderPath, sheet)
        for r in range(2, self.rows + 1):
            self.testCase = XLUtils.readData(getFolderPath, sheet, r, 1)
            self.action = XLUtils.readData(getFolderPath, sheet, r, 3)
            if self.testCase == method and self.action == 'Y':
                return "success"

    def test_excel_file_data_clear(self):
        book = openpyxl.load_workbook('Reports/result_chat_agent_cases_output.xlsx')
        sheet = book['Chatagent-Testcases-Output']
        sheet.delete_rows(2, 20)
        book.save('Reports/result_chat_agent_cases_output.xlsx')



    def check_chat_agent_user_login(self):
        self.driver = webdriver.Chrome(ChromeDriverManager().install())
        self.driver.maximize_window()
        sleep(1)
        path = "assets"
        fileName = "data.xlsx"
        getFolderPath = os.path.join(path, fileName)
        self.lp = ChatagentLoginClass(self.driver)
        self.rows = XLUtils.getRowCount(getFolderPath, 'Chat-Agent')
        for r in range(4, self.rows + 1):
            self.login_url = XLUtils.readData(getFolderPath, 'Chat-Agent', r, 1)
            self.email = XLUtils.readData(getFolderPath, 'Chat-Agent', r, 2)
            self.password = XLUtils.readData(getFolderPath, 'Chat-Agent', r, 3)
            self.driver.get(self.login_url)
            sleep(1)
            self.lp.setchatagentEmailName(self.email)
            sleep(2)
            self.lp.setchatagentPassword(self.password)
            sleep(2)
            self.lp.clickchatagentPLogin()
            sleep(10)
            act_title = self.driver.title
            return act_title

    def test_chat_agent_user_login(self):
        if self.excelReader_test_cases('test_chat_agent_user_login', 'test_cases') == "success":
            act_title = self.check_chat_agent_user_login()
            exp_title = "User Dashboard"
            if act_title == exp_title:
                sleep(8)
                wb = Workbook()
                ws = wb.active
                ws['A1'] = "Test Cases"
                ws['A1'].font = Font(name="Arial", b=True)
                ws['B1'] = "Status"
                ws['B1'].font = Font(name="Arial", b=True)
                ws.title = "Chatagent-Testcases-Output"
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                wb.save('Reports/result_chat_agent_cases_output.xlsx')
                #self.add_s3bucket_file_upload()
                self.driver.close()
                assert True
            else:
                wb = Workbook()
                ws = wb.active
                ws['A1'] = "Test Cases"
                ws['A1'].font = Font(name="Arial", b=True)
                ws['B1'] = "Status"
                ws['B1'].font = Font(name="Arial", b=True)
                ws.title = "Chatagent-Testcases-Output"
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                wb.save('Reports/result_chat_agent_cases_output.xlsx')
                #self.add_s3bucket_file_upload()
                self.driver.close()




    def test_chat_agent_new_response(self):
        if self.excelReader_test_cases('test_chat_agent_new_response', 'test_cases') == "success":
            act_title = self.check_chat_agent_user_login()
            exp_title = "User Dashboard"
            if act_title == exp_title:
                book = openpyxl.load_workbook('Reports/result_chat_agent_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_chat_agent_cases_output.xlsx')
                self.chat_agent_user_obj = chatagentnewresponseClass(self.driver)
                self.chat_agent_user_obj.clickchatagentnewresponse()
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_chat_agent_cases_output.xlsx')
                ws = book.active
                ws['A3'] = "test-new-response"
                ws['B3'] = "Pass"
                book.save('Reports/result_chat_agent_cases_output.xlsx')
                self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_chat_agent_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_chat_agent_cases_output.xlsx')
                self.driver.close()

    def test_chat_agent_past_response(self):
        if self.excelReader_test_cases('test_chat_agent_past_response', 'test_cases') == "success":
            act_title = self.check_chat_agent_user_login()
            exp_title = "User Dashboard"
            if act_title == exp_title:
                book = openpyxl.load_workbook('Reports/result_chat_agent_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_chat_agent_cases_output.xlsx')
                self.chat_agent_user_obj = chatagentpastresponseClass(self.driver)
                self.chat_agent_user_obj.clickchatagentpastresponse()
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_chat_agent_cases_output.xlsx')
                ws = book.active
                ws['A4'] = "test-past-response"
                ws['B4'] = "Pass"
                book.save('Reports/result_chat_agent_cases_output.xlsx')
                self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_chat_agent_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_chat_agent_cases_output.xlsx')
                self.driver.close()






    def test_chat_agent_ignore_response(self):
        if self.excelReader_test_cases('test_chat_agent_ignore_response', 'test_cases') == "success":
            act_title = self.check_chat_agent_user_login()
            exp_title = "User Dashboard"
            if act_title == exp_title:
                book = openpyxl.load_workbook('Reports/result_chat_agent_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_chat_agent_cases_output.xlsx')
                self.chat_agent_user_obj = chatagentignoreresponseClass(self.driver)
                self.chat_agent_user_obj.clickchatagentignoreresponse()
                sleep(2)
                book = openpyxl.load_workbook('Reports/result_chat_agent_cases_output.xlsx')
                ws = book.active
                ws['A5'] = "test-ignore-response"
                ws['B5'] = "Pass"
                book.save('Reports/result_chat_agent_cases_output.xlsx')
                #self.add_s3bucket_file_upload()
                self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_chat_agent_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_chat_agent_cases_output.xlsx')
                self.driver.close()

    def test_chat_agent_profile_update(self):
        if self.excelReader_test_cases('test_chat_agent_profile_update', 'test_cases') == "success":
            act_title = self.check_chat_agent_user_login()
            exp_title = "User Dashboard"
            if act_title == exp_title:
                book = openpyxl.load_workbook('Reports/result_chat_agent_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Pass"
                book.save('Reports/result_chat_agent_cases_output.xlsx')
                path = "assets"
                fileName = "data.xlsx"
                getFolderPath = os.path.join(path, fileName)
                self.rows = XLUtils.getRowCount(getFolderPath, 'Chat-Agent')
                self.chatagent_profile_obj = chatagentUserprofileClass(self.driver)
                for r in range(9, self.rows + 1):
                    self.first_name = XLUtils.readData(getFolderPath, 'Chat-Agent', r, 1)
                    self.last_name = XLUtils.readData(getFolderPath, 'Chat-Agent', r, 2)
                    self.phone_number = XLUtils.readData(getFolderPath, 'Chat-Agent', r, 3)
                    self.group_assignment = XLUtils.readData(getFolderPath, 'Chat-Agent', r, 4)
                    self.current_password = XLUtils.readData(getFolderPath, 'Chat-Agent', r, 5)
                    self.new_password = XLUtils.readData(getFolderPath, 'Chat-Agent', r, 6)
                    self.confirm_password = XLUtils.readData(getFolderPath, 'Chat-Agent', r, 7)
                    self.chatagent_profile_obj.click_chat_agent_profile_page_link()
                    self.chatagent_profile_obj.set_chatagent_First_Name(self.first_name)
                    self.chatagent_profile_obj.set_chatagent_Last_Name(self.last_name)
                    self.chatagent_profile_obj.set_chatagent_Phone_number(self.phone_number)
                    sleep(2)
                    self.chatagent_profile_obj.click_chat_agent_profile_date()
                    sleep(2)
                    self.chatagent_profile_obj.set_chat_agent_group_assignment(self.group_assignment)
                    sleep(2)
                    self.chatagent_profile_obj.chatagent_profile_Selectstate()
                    sleep(2)
                    self.chatagent_profile_obj.setchatagent_current_password(self.current_password)
                    sleep(2)
                    self.chatagent_profile_obj.setchatagent_new_password(self.new_password)
                    sleep(2)
                    self.chatagent_profile_obj.setchatagent_confirm_password(self.confirm_password)
                    sleep(2)
                    self.chatagent_profile_obj.clickchatagentupdatebtn()
                    sleep(3)
                    try:
                        success_toastr = self.driver.find_element_by_class_name("toast-error")
                        if success_toastr:
                            book = openpyxl.load_workbook('Reports/result_chat_agent_cases_output.xlsx')
                            ws = book.active
                            ws['A6'] = "test-profile-update"
                            ws['B6'] = "Fail"
                            book.save('Reports/result_chat_agent_cases_output.xlsx')
                            #self.add_s3bucket_file_upload()
                            self.driver.close()
                    except NoSuchElementException:
                        sleep(4)
                        book = openpyxl.load_workbook('Reports/result_chat_agent_cases_output.xlsx')
                        ws = book.active
                        ws['A6'] = "test-profile-update"
                        ws['B6'] = "Pass"
                        book.save('Reports/result_chat_agent_cases_output.xlsx')
                        #self.add_s3bucket_file_upload()
                        self.driver.close()
            else:
                book = openpyxl.load_workbook('Reports/result_chat_agent_cases_output.xlsx')
                ws = book.active
                ws['A2'] = "test-login"
                ws['B2'] = "Fail"
                book.save('Reports/result_chat_agent_cases_output.xlsx')
                self.driver.close()


    def test_excel_empty_row_remove(self):
        book = openpyxl.load_workbook('Reports/result_chat_agent_cases_output.xlsx')
        ws = book.active
        index_row = []
        # loop each row in column A
        for i in range(1, ws.max_row):
            # define emptiness of cell
            if ws.cell(i, 1).value is None:
                # collect indexes of rows
                index_row.append(i)
        # loop each index value
        for row_del in range(len(index_row)):
            ws.delete_rows(idx=index_row[row_del], amount=1)
            # exclude offset of rows through each iteration
            index_row = list(map(lambda k: k - 1, index_row))
        book.save('Reports/result_chat_agent_cases_output.xlsx')


    def test_s3bucket_file_upload(self):
        client_s3 = boto3.client('s3', aws_access_key_id=aws_access_key_id,
                                 aws_secret_access_key=aws_secret_access_key)
        with open('Reports/result_chat_agent_cases_output.xlsx', 'rb') as data:
            return client_s3.upload_fileobj(data, bucket_name, 'Reports/result_chat_agent_user' + str(date) + '.xlsx')